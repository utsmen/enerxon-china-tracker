#!/usr/bin/env python3
"""
Charlie Tracker — Multi-Project Spool Production Tracking (Generalised)
URL structure: / → /project/<id> → /project/<id>/spool/<spool_id>
All ITP steps, diameters, and production rates are project-specific (no hardcoded constants).
"""
import os, sys, json
from datetime import datetime, date, timedelta
from functools import wraps
from flask import Flask, render_template_string, request, jsonify, send_file, g, session, redirect, url_for

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'charlie-dev-key')
DATABASE_URL = os.environ.get('DATABASE_URL', '')
USE_PG = DATABASE_URL.startswith('postgres')
SITE_PASSWORD = os.environ.get('SITE_PASSWORD', 'Enerxon@china')

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('authenticated'):
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Not authenticated'}), 401
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = ''
    if request.method == 'POST':
        pw = request.form.get('password', '')
        if pw == SITE_PASSWORD:
            session['authenticated'] = True
            session.permanent = True
            app.permanent_session_lifetime = timedelta(days=90)
            return redirect('/')
        error = 'Wrong password / \u5bc6\u7801\u9519\u8bef'
    return render_template_string(LOGIN_HTML, error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

LOGIN_HTML = """<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>ENERXON Tracker — Login</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,'PingFang SC','Microsoft YaHei',sans-serif;background:#f0f2f5;display:flex;align-items:center;justify-content:center;min-height:100vh}
.login-box{background:#fff;border-radius:16px;padding:40px;width:320px;box-shadow:0 4px 20px rgba(0,0,0,0.1);text-align:center}
.login-box h1{color:#2F5496;font-size:20px;margin-bottom:4px}
.login-box .cn{color:#888;font-size:13px;margin-bottom:24px}
.login-box input{width:100%;padding:12px;border:2px solid #ddd;border-radius:8px;font-size:16px;margin-bottom:16px;text-align:center}
.login-box input:focus{border-color:#2F5496;outline:none}
.login-box button{width:100%;padding:12px;background:#2F5496;color:#fff;border:none;border-radius:8px;font-size:16px;font-weight:600;cursor:pointer}
.login-box button:hover{background:#1a3a6e}
.error{color:#e74c3c;font-size:13px;margin-bottom:12px}
</style></head><body>
<div class="login-box">
  <h1>ENERXON Tracker</h1>
  <div class="cn">\u751f\u4ea7\u8fdb\u5ea6\u8ffd\u8e2a\u7cfb\u7edf</div>
  {% if error %}<div class="error">{{ error }}</div>{% endif %}
  <form method="POST">
    <input type="password" name="password" placeholder="Password / \u5bc6\u7801" autofocus>
    <button type="submit">Enter / \u8fdb\u5165</button>
  </form>
</div>
</body></html>"""

# Default step definitions — fallback when project has no steps defined yet
_DEFAULT_STEPS = [
    {'step_number':1,'name_en':'Material Receiving & Traceability','name_cn':'\u6765\u6599\u68c0\u9a8c\u53ca\u53ef\u8ffd\u6eaf\u6027','weight':5,'category':'fab_fixed','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':1,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'fab'},
    {'step_number':2,'name_en':'Documentation Review (WPS/PQR/ITP)','name_cn':'\u6587\u4ef6\u5ba1\u67e5\uff08WPS/PQR/ITP\uff09','weight':3,'category':'fab_fixed','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':2,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'fab'},
    {'step_number':3,'name_en':'Pipe Cutting \u2014 Dimensional Check','name_cn':'\u7ba1\u9053\u5207\u5272 \u2014 \u5c3a\u5bf8\u68c0\u9a8c','weight':8,'category':'fab_fixed','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':3,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'fab'},
    {'step_number':4,'name_en':'End Preparation / Bevelling','name_cn':'\u7ba1\u53e3\u51c6\u5907 / \u5761\u53e3\u52a0\u5de5','weight':5,'category':'fab_fixed','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':4,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'fab'},
    {'step_number':5,'name_en':'Fit-Up & Assembly Inspection','name_cn':'\u7ec4\u5bf9\u53ca\u88c5\u914d\u68c0\u9a8c','weight':10,'category':'fab_fixed','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':5,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'fab'},
    {'step_number':16,'name_en':'Branch Welding','name_cn':'\u652f\u7ba1\u710a\u63a5','weight':5,'category':'branch','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':6,'is_conditional':1,'is_hold_point':0,'is_release':0,'phase':'fab'},
    {'step_number':6,'name_en':'Production Welding as per WPS','name_cn':'\u6309WPS\u4e3b\u7ba1\u710a\u63a5','weight':10,'category':'welding','hours_fixed':0,'hours_variable':'welding','spool_type':'ALL','display_order':7,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'fab'},
    {'step_number':7,'name_en':'Visual Inspection (VT) \u2014 100%','name_cn':'\u76ee\u89c6\u68c0\u9a8cVT\uff08\u5168\u68c0\uff09','weight':8,'category':'fab_fixed','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':8,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'fab'},
    {'step_number':8,'name_en':'Radiographic Test (RT) \u2014 100%','name_cn':'\u5c04\u7ebf\u68c0\u6d4bRT\uff08\u5168\u68c0\uff09\u2605\u505c\u6b62\u70b9','weight':10,'category':'fab_fixed','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':9,'is_conditional':0,'is_hold_point':1,'is_release':0,'phase':'fab'},
    {'step_number':9,'name_en':'Magnetic Particle (MT) \u2014 100%','name_cn':'\u78c1\u7c89\u68c0\u6d4bMT\uff08\u5168\u68c0\uff09','weight':5,'category':'fab_fixed','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':10,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'fab'},
    {'step_number':10,'name_en':'Cleaning Prior to Painting','name_cn':'\u6d82\u88c5\u524d\u6e05\u6d01\u5904\u7406','weight':3,'category':'paint_fixed','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':11,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'paint'},
    {'step_number':11,'name_en':'Surface Preparation \u2014 Blasting','name_cn':'\u8868\u9762\u5904\u7406 \u2014 \u55b7\u7802','weight':8,'category':'surface_treatment','hours_fixed':0,'hours_variable':'surface','spool_type':'ALL','display_order':12,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'paint'},
    {'step_number':12,'name_en':'Painting Application','name_cn':'\u6d82\u88c5\u65bd\u5de5\uff08\u5e95\u6f06/\u4e2d\u95f4\u6f06/\u9762\u6f06\uff09','weight':8,'category':'surface_treatment','hours_fixed':0,'hours_variable':'surface','spool_type':'ALL','display_order':13,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'paint'},
    {'step_number':13,'name_en':'Coating Inspection \u2014 DFT','name_cn':'\u6d82\u5c42\u68c0\u9a8c \u2014 \u819c\u539a\u53ca\u9644\u7740\u529b','weight':4,'category':'paint_fixed','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':14,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'paint'},
    {'step_number':14,'name_en':'Dimensional Inspection & Marking','name_cn':'\u5c3a\u5bf8\u68c0\u9a8c\u53ca\u6807\u8bc6','weight':5,'category':'paint_fixed','hours_fixed':2.0,'hours_variable':'','spool_type':'ALL','display_order':15,'is_conditional':0,'is_hold_point':0,'is_release':0,'phase':'paint'},
    {'step_number':15,'name_en':'Final Inspection \u2014 Released','name_cn':'\u6700\u7ec8\u68c0\u9a8c \u2014 \u53d1\u8d27\u653e\u884c \u2605\u89c1\u8bc1','weight':3,'category':'packing','hours_fixed':3.0,'hours_variable':'','spool_type':'ALL','display_order':16,'is_conditional':0,'is_hold_point':0,'is_release':1,'phase':'paint','counts_for_production':0},
]

# ── DB Layer ──────────────────────────────────────────────────────────────────
def get_db():
    if 'db' not in g:
        if USE_PG:
            import psycopg2, psycopg2.extras
            g.db = psycopg2.connect(DATABASE_URL.replace('postgres://','postgresql://',1), connect_timeout=10)
            g.db.autocommit = False; g.db_type = 'pg'
        else:
            import sqlite3
            g.db = sqlite3.connect(os.environ.get('SQLITE_PATH','tracker.db'))
            g.db.row_factory = sqlite3.Row; g.db_type = 'sqlite'
    return g.db

def db_execute(q, p=None):
    db = get_db()
    if g.db_type == 'pg':
        import psycopg2.extras
        q = q.replace('?','%s')
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(q, p or ()); return cur
    return db.execute(q, p or ())

def db_fetchall(q, p=None):
    cur = db_execute(q, p); rows = cur.fetchall()
    return rows if (USE_PG and g.db_type=='pg') else [dict(r) for r in rows]

def db_fetchone(q, p=None):
    cur = db_execute(q, p); r = cur.fetchone()
    return r if (r and USE_PG and g.db_type=='pg') else (dict(r) if r else None)

def db_commit(): get_db().commit()

@app.teardown_appcontext
def close_db(e):
    db = g.pop('db', None)
    if db: db.close()

def init_db():
    if USE_PG:
        import psycopg2
        c = psycopg2.connect(DATABASE_URL.replace('postgres://','postgresql://',1)); c.autocommit = True
        cur = c.cursor()
        # Execute each statement separately so one failure doesn't block the rest
        pg_statements = [
            "CREATE TABLE IF NOT EXISTS spools (id SERIAL PRIMARY KEY, spool_id TEXT NOT NULL, spool_full TEXT DEFAULT '', iso_no TEXT DEFAULT '', marking TEXT DEFAULT '', mk_number TEXT DEFAULT '', main_diameter TEXT DEFAULT '', line TEXT DEFAULT '', sequence INTEGER DEFAULT 0, project TEXT DEFAULT '', has_branches INTEGER DEFAULT 0, spool_type TEXT DEFAULT 'SPOOL', actual_weight_kg REAL DEFAULT 0, surface_m2 REAL DEFAULT 0, joint_count INTEGER DEFAULT 0, raf_inches REAL DEFAULT 0, created_at TIMESTAMP DEFAULT NOW(), UNIQUE(project, spool_id))",
            "CREATE TABLE IF NOT EXISTS progress (id SERIAL PRIMARY KEY, spool_id TEXT NOT NULL, project TEXT DEFAULT '', step_number INTEGER NOT NULL, completed INTEGER DEFAULT 0, completed_by TEXT DEFAULT '', completed_at TIMESTAMP, remarks TEXT DEFAULT '', UNIQUE(project, spool_id, step_number))",
            "CREATE TABLE IF NOT EXISTS activity_log (id SERIAL PRIMARY KEY, spool_id TEXT NOT NULL, project TEXT DEFAULT '', step_number INTEGER, action TEXT NOT NULL, operator TEXT DEFAULT '', timestamp TIMESTAMP DEFAULT NOW(), details TEXT DEFAULT '')",
            "CREATE INDEX IF NOT EXISTS idx_progress_ps ON progress(project, spool_id)",
            "CREATE INDEX IF NOT EXISTS idx_activity_ps ON activity_log(project, spool_id)",
            "CREATE TABLE IF NOT EXISTS schedule (id SERIAL PRIMARY KEY, project TEXT NOT NULL, diameter TEXT NOT NULL, task_type TEXT NOT NULL, description TEXT DEFAULT '', planned_start DATE NOT NULL, planned_end DATE NOT NULL, spool_count INTEGER DEFAULT 0, UNIQUE(project, diameter, task_type))",
            "CREATE TABLE IF NOT EXISTS project_settings (id SERIAL PRIMARY KEY, project TEXT NOT NULL, key TEXT NOT NULL, value TEXT DEFAULT '', UNIQUE(project, key))",
            "CREATE TABLE IF NOT EXISTS drawings (id SERIAL PRIMARY KEY, project TEXT NOT NULL, spool_id TEXT NOT NULL, pdf_data BYTEA NOT NULL, UNIQUE(project, spool_id))",
            "CREATE TABLE IF NOT EXISTS project_steps (id SERIAL PRIMARY KEY, project TEXT NOT NULL, step_number INTEGER NOT NULL, name_en TEXT NOT NULL, name_cn TEXT NOT NULL DEFAULT '', weight INTEGER DEFAULT 5, category TEXT NOT NULL, hours_fixed REAL DEFAULT 2.0, hours_variable TEXT DEFAULT '', spool_type TEXT DEFAULT 'ALL', display_order INTEGER NOT NULL, is_conditional INTEGER DEFAULT 0, is_hold_point INTEGER DEFAULT 0, is_release INTEGER DEFAULT 0, phase TEXT DEFAULT 'fab', counts_for_production INTEGER DEFAULT 1, UNIQUE(project, step_number))",
            "ALTER TABLE project_steps ADD COLUMN counts_for_production INTEGER DEFAULT 1",
            # Ensure spool_type column exists on old spools tables and backfill NULLs
            "ALTER TABLE spools ADD COLUMN spool_type TEXT DEFAULT 'SPOOL'",
            # Shipment assignment per spool (nullable FK to shipments.shipment_number)
            "ALTER TABLE spools ADD COLUMN shipment_number INTEGER",
            "UPDATE spools SET spool_type = 'SPOOL' WHERE spool_type IS NULL",
            # Backfill: any spool named 'STRAIGHT PIPE ...' is by definition a straight pipe
            "UPDATE spools SET spool_type = 'STRAIGHT' WHERE spool_id LIKE 'STRAIGHT PIPE%' AND spool_type <> 'STRAIGHT'",
            # Migrate legacy schedule task_type names to phase names
            "UPDATE schedule SET task_type = 'fab' WHERE task_type IN ('fabrication', 'Fabricacion')",
            "UPDATE schedule SET task_type = 'paint' WHERE task_type IN ('painting', 'Pintura')",
            # QC Reporting tables
            "CREATE TABLE IF NOT EXISTS qc_reports (id SERIAL PRIMARY KEY, project TEXT NOT NULL, spool_id TEXT NOT NULL, report_type TEXT NOT NULL, report_subtype TEXT DEFAULT '', status TEXT DEFAULT 'draft', inspector_name TEXT DEFAULT '', inspector_date TEXT DEFAULT '', tpi_name TEXT DEFAULT '', tpi_date TEXT DEFAULT '', data TEXT DEFAULT '{}', created_at TIMESTAMP DEFAULT NOW(), updated_at TIMESTAMP DEFAULT NOW(), created_by TEXT DEFAULT '', UNIQUE(project, spool_id, report_type, report_subtype))",
            "CREATE TABLE IF NOT EXISTS qc_images (id SERIAL PRIMARY KEY, project TEXT NOT NULL, spool_id TEXT NOT NULL, report_type TEXT NOT NULL, image_data BYTEA NOT NULL, filename TEXT DEFAULT '', mime_type TEXT DEFAULT 'image/jpeg', caption TEXT DEFAULT '', uploaded_at TIMESTAMP DEFAULT NOW(), uploaded_by TEXT DEFAULT '')",
            "CREATE INDEX IF NOT EXISTS idx_qc_reports_ps ON qc_reports(project, spool_id)",
            "CREATE INDEX IF NOT EXISTS idx_qc_images_ps ON qc_images(project, spool_id, report_type)",
            "CREATE TABLE IF NOT EXISTS qc_inspectors (id SERIAL PRIMARY KEY, name TEXT NOT NULL, role TEXT DEFAULT '', signature_data TEXT DEFAULT '', created_at TIMESTAMP DEFAULT NOW(), UNIQUE(name))",
            "CREATE TABLE IF NOT EXISTS shipments (id SERIAL PRIMARY KEY, project TEXT NOT NULL, shipment_number INTEGER NOT NULL, description TEXT DEFAULT '', etd DATE, transit_days INTEGER DEFAULT 45, notes TEXT DEFAULT '', created_at TIMESTAMP DEFAULT NOW(), UNIQUE(project, shipment_number))",
            # Chat agent log (production & quality assistant)
            "CREATE TABLE IF NOT EXISTS chat_log (id SERIAL PRIMARY KEY, project TEXT NOT NULL, user_msg TEXT NOT NULL, assistant_msg TEXT NOT NULL, tools_used TEXT DEFAULT '[]', feedback TEXT DEFAULT '', created_at TIMESTAMP DEFAULT NOW())",
            "CREATE INDEX IF NOT EXISTS idx_chat_log_project ON chat_log(project, created_at)",
        ]
        for stmt in pg_statements:
            try: cur.execute(stmt)
            except Exception as e: print(f"init_db skip: {e}")
        c.close()
    else:
        import sqlite3
        c = sqlite3.connect(os.environ.get('SQLITE_PATH','tracker.db'))
        c.executescript("""
            CREATE TABLE IF NOT EXISTS spools (id INTEGER PRIMARY KEY AUTOINCREMENT, spool_id TEXT NOT NULL, spool_full TEXT DEFAULT '', iso_no TEXT DEFAULT '', marking TEXT DEFAULT '', mk_number TEXT DEFAULT '', main_diameter TEXT DEFAULT '', line TEXT DEFAULT '', sequence INTEGER DEFAULT 0, project TEXT DEFAULT '', has_branches INTEGER DEFAULT 0, spool_type TEXT DEFAULT 'SPOOL', actual_weight_kg REAL DEFAULT 0, surface_m2 REAL DEFAULT 0, joint_count INTEGER DEFAULT 0, raf_inches REAL DEFAULT 0, created_at TEXT DEFAULT (datetime('now')), UNIQUE(project, spool_id));
            CREATE TABLE IF NOT EXISTS progress (id INTEGER PRIMARY KEY AUTOINCREMENT, spool_id TEXT NOT NULL, project TEXT DEFAULT '', step_number INTEGER NOT NULL, completed INTEGER DEFAULT 0, completed_by TEXT DEFAULT '', completed_at TEXT, remarks TEXT DEFAULT '', UNIQUE(project, spool_id, step_number));
            CREATE TABLE IF NOT EXISTS activity_log (id INTEGER PRIMARY KEY AUTOINCREMENT, spool_id TEXT NOT NULL, project TEXT DEFAULT '', step_number INTEGER, action TEXT NOT NULL, operator TEXT DEFAULT '', timestamp TEXT DEFAULT (datetime('now')), details TEXT DEFAULT '');
            CREATE TABLE IF NOT EXISTS schedule (id INTEGER PRIMARY KEY AUTOINCREMENT, project TEXT NOT NULL, diameter TEXT NOT NULL, task_type TEXT NOT NULL, description TEXT DEFAULT '', planned_start TEXT NOT NULL, planned_end TEXT NOT NULL, spool_count INTEGER DEFAULT 0, UNIQUE(project, diameter, task_type));
            CREATE TABLE IF NOT EXISTS project_settings (id INTEGER PRIMARY KEY AUTOINCREMENT, project TEXT NOT NULL, key TEXT NOT NULL, value TEXT DEFAULT '', UNIQUE(project, key));
            CREATE TABLE IF NOT EXISTS drawings (id INTEGER PRIMARY KEY AUTOINCREMENT, project TEXT NOT NULL, spool_id TEXT NOT NULL, pdf_data BLOB NOT NULL, UNIQUE(project, spool_id));
            CREATE TABLE IF NOT EXISTS project_steps (id INTEGER PRIMARY KEY AUTOINCREMENT, project TEXT NOT NULL, step_number INTEGER NOT NULL, name_en TEXT NOT NULL, name_cn TEXT NOT NULL DEFAULT '', weight INTEGER DEFAULT 5, category TEXT NOT NULL, hours_fixed REAL DEFAULT 2.0, hours_variable TEXT DEFAULT '', spool_type TEXT DEFAULT 'ALL', display_order INTEGER NOT NULL, is_conditional INTEGER DEFAULT 0, is_hold_point INTEGER DEFAULT 0, is_release INTEGER DEFAULT 0, phase TEXT DEFAULT 'fab', counts_for_production INTEGER DEFAULT 1, UNIQUE(project, step_number));
            UPDATE spools SET spool_type = 'STRAIGHT' WHERE spool_id LIKE 'STRAIGHT PIPE%' AND spool_type <> 'STRAIGHT';
            UPDATE schedule SET task_type = 'fab' WHERE task_type IN ('fabrication', 'Fabricacion');
            UPDATE schedule SET task_type = 'paint' WHERE task_type IN ('painting', 'Pintura');
            CREATE TABLE IF NOT EXISTS qc_reports (id INTEGER PRIMARY KEY AUTOINCREMENT, project TEXT NOT NULL, spool_id TEXT NOT NULL, report_type TEXT NOT NULL, report_subtype TEXT DEFAULT '', status TEXT DEFAULT 'draft', inspector_name TEXT DEFAULT '', inspector_date TEXT DEFAULT '', tpi_name TEXT DEFAULT '', tpi_date TEXT DEFAULT '', data TEXT DEFAULT '{}', created_at TEXT DEFAULT (datetime('now')), updated_at TEXT DEFAULT (datetime('now')), created_by TEXT DEFAULT '', UNIQUE(project, spool_id, report_type, report_subtype));
            CREATE TABLE IF NOT EXISTS qc_images (id INTEGER PRIMARY KEY AUTOINCREMENT, project TEXT NOT NULL, spool_id TEXT NOT NULL, report_type TEXT NOT NULL, image_data BLOB NOT NULL, filename TEXT DEFAULT '', mime_type TEXT DEFAULT 'image/jpeg', caption TEXT DEFAULT '', uploaded_at TEXT DEFAULT (datetime('now')), uploaded_by TEXT DEFAULT '');
            CREATE TABLE IF NOT EXISTS qc_inspectors (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE, role TEXT DEFAULT '', signature_data TEXT DEFAULT '', created_at TEXT DEFAULT (datetime('now')));
            CREATE TABLE IF NOT EXISTS shipments (id INTEGER PRIMARY KEY AUTOINCREMENT, project TEXT NOT NULL, shipment_number INTEGER NOT NULL, description TEXT DEFAULT '', etd TEXT, transit_days INTEGER DEFAULT 45, notes TEXT DEFAULT '', created_at TEXT DEFAULT (datetime('now')), UNIQUE(project, shipment_number));
            CREATE TABLE IF NOT EXISTS chat_log (id INTEGER PRIMARY KEY AUTOINCREMENT, project TEXT NOT NULL, user_msg TEXT NOT NULL, assistant_msg TEXT NOT NULL, tools_used TEXT DEFAULT '[]', feedback TEXT DEFAULT '', created_at TEXT DEFAULT (datetime('now')));
            CREATE INDEX IF NOT EXISTS idx_chat_log_project ON chat_log(project, created_at);
        """); c.close()

# ── QC Report Type Registry ──────────────────────────────────────────────────
# Each report maps to a specific ITP step. 'rec_seq' is the record number suffix.
# 'spool_only' means the report only applies to welded spools (not straight pipes).
# Reports are ordered by ITP step sequence within each project.
# Print format: English-only EN 10204 3.1 certificate. Web forms: bilingual EN/CN.
# QC category labels — universal, not project-specific
QC_CATEGORY_LABELS = {
    'fab': ('Fabrication / \u5236\u9020', '#2F5496'),
    'weld': ('Welding / \u710a\u63a5', '#C0392B'),
    'ndt': ('NDT / \u65e0\u635f\u68c0\u6d4b', '#8E44AD'),
    'measurement': ('Measurement / \u6d4b\u91cf', '#27AE60'),
    'material': ('Material / \u6750\u6599', '#E67E22'),
    'surface': ('Surface Treatment / \u8868\u9762\u5904\u7406', '#16A085'),
}

def get_qc_setting(project, key, default=None):
    """Get a QC-related project setting, parsing JSON if needed."""
    row = db_fetchone("SELECT value FROM project_settings WHERE project=? AND key=?", (project, key))
    if not row:
        return default
    val = row['value']
    if val and val.startswith(('[', '{')):
        return json.loads(val)
    return val

def get_qc_report_defs(project):
    """Load QC report definitions from project_settings."""
    return get_qc_setting(project, 'qc_report_defs', [])

def get_qc_project_info(project):
    """Load QC project info (contract, client, material, ITP) from project_settings."""
    return get_qc_setting(project, 'qc_project_info', {})

def get_wps_registry(project):
    """Load WPS registry from project_settings."""
    return get_qc_setting(project, 'wps_registry', {})

def get_qc_reports_for_spool(project, spool_type='SPOOL'):
    """Return list of report defs applicable to this project+spool_type."""
    defs = get_qc_report_defs(project)
    is_straight = (spool_type or '').upper() == 'STRAIGHT'
    return [d for d in defs if not (is_straight and d.get('spool_only', False))]

def get_record_number(project, spool_id, rec_seq):
    """Generate record number: ENJOB25011423-REC-SPL-001-001"""
    return f"{project}-REC-{spool_id}-{rec_seq}"

# ── Helpers: Project Steps & Settings ─────────────────────────────────────────
def get_project_steps(project):
    """Load step definitions for a project from DB. Cached per request."""
    cache_key = f'_steps_{project}'
    if hasattr(g, cache_key):
        return getattr(g, cache_key)
    rows = db_fetchall("SELECT * FROM project_steps WHERE project=? ORDER BY display_order", (project,))
    if not rows:
        # Fallback: seed legacy 424 steps for this project
        rows = _DEFAULT_STEPS
    setattr(g, cache_key, rows)
    return rows

def get_project_settings(project):
    """Get project settings with defaults."""
    rows = db_fetchall("SELECT key, value FROM project_settings WHERE project=?", (project,))
    settings = {r['key']: r['value'] for r in rows}
    # Migrate old setting names to new ones
    key_migrations = {'painting_capability_m2d': 'surface_capability_m2d', 'painting_days': 'secondary_phase_days'}
    for old_key, new_key in key_migrations.items():
        if old_key in settings and new_key not in settings:
            settings[new_key] = settings[old_key]
    defaults = {'committed_weeks_saved':'0', 'committed_days_saved':'0', 'sea_transit_days':'45',
                'standard_weeks':'9', 'welding_capability_ipd':'1000', 'surface_capability_m2d':'91',
                'spools_per_day':'{}', 'secondary_phase_days':'13'}
    for k,v in defaults.items():
        if k not in settings: settings[k] = v
    return settings

def get_diameter_order(project):
    """Get diameter order for a project, sorted descending by numeric value."""
    rows = db_fetchall("SELECT DISTINCT main_diameter FROM spools WHERE project=? AND main_diameter != '' AND main_diameter != ?", (project, '?'))
    diameters = []
    for r in rows:
        d = (r['main_diameter'] or '').replace('"', '')
        try: diameters.append((float(d), d))
        except ValueError: diameters.append((0, d))
    diameters.sort(reverse=True)
    return [d[1] for d in diameters]

# ── Helpers: Progress Calculation ─────────────────────────────────────────────
def get_phase_order(steps_def):
    """Get ordered list of distinct phases from step definitions (by first appearance in display_order)."""
    seen = set(); phases = []
    for s in steps_def:
        p = s['phase']
        if p not in seen: seen.add(p); phases.append(p)
    return phases

def step_applies_to_spool(step, spool):
    """Single applicability rule — does this step definition apply to this spool?
    Used by every consumer of steps_def to decide inclusion. The rule is:
      - If step.spool_type != 'ALL', must match spool.spool_type
      - If step.is_conditional, spool must have branches
    No project-specific logic. Any consumer that wants a step-per-spool
    filter MUST call this helper instead of duplicating the rule."""
    st_type = step.get('spool_type', 'ALL') or 'ALL'
    if st_type != 'ALL' and (spool.get('spool_type', 'SPOOL') or 'SPOOL') != st_type:
        return False
    if step.get('is_conditional') and not spool.get('has_branches'):
        return False
    return True

def get_post_production_steps(steps_def):
    """Return post-production steps (counts_for_production=0) sorted by display_order.
    The single authoritative definition of post-production ordering. Every consumer
    that needs post-production steps MUST call this helper instead of filtering
    and sorting independently."""
    return sorted(
        [s for s in steps_def if not s.get('counts_for_production', 1)],
        key=lambda s: s.get('display_order') or s.get('step_number', 0))

def spool_hours(spool_row, completed_steps, settings, steps_def):
    """Calculate hours-based progress for a single spool. Phases are dynamic — derived from step definitions."""
    raf = float(spool_row.get('raf_inches') or 0)
    surface = float(spool_row.get('surface_m2') or 0)
    spool_type = spool_row.get('spool_type', 'SPOOL') or 'SPOOL'
    weld_cap = float(settings.get('welding_capability_ipd', '552'))
    surface_cap = float(settings.get('surface_capability_m2d', '91'))

    # Pre-count surface steps per phase for even splitting
    surface_count_by_phase = {}
    for step in steps_def:
        if step['hours_variable'] == 'surface' and step['phase']:
            surface_count_by_phase[step['phase']] = surface_count_by_phase.get(step['phase'], 0) + 1

    # Accumulate per phase
    phase_totals = {}  # {phase: {total, done}}
    weld_hrs = 0.0; surface_hrs = 0.0

    for step in steps_def:
        sn = step['step_number']
        phase = step['phase']
        if not step_applies_to_spool(step, spool_row): continue
        if not step.get('counts_for_production', 1): continue

        # Calculate hours for this step
        if step['hours_variable'] == 'welding':
            hrs = (raf / weld_cap * 8) if weld_cap > 0 and raf > 0 else 0
            weld_hrs = hrs
        elif step['hours_variable'] == 'surface':
            if surface <= 0: continue
            n_surface = surface_count_by_phase.get(phase, 1)
            total_surface_hrs = (surface * 0.98 / surface_cap * 8) if surface_cap > 0 else 0
            hrs = total_surface_hrs / n_surface if n_surface > 0 else 0
            surface_hrs += hrs
        else:
            hrs = float(step.get('hours_fixed', 2.0) or 2.0)

        if phase not in phase_totals: phase_totals[phase] = {'total': 0.0, 'done': 0.0}
        phase_totals[phase]['total'] += hrs
        if sn in completed_steps: phase_totals[phase]['done'] += hrs

    # Calculate percentages per phase
    phases = {}
    total = 0.0; done = 0.0
    for phase, pt in phase_totals.items():
        pct = round(pt['done'] / pt['total'] * 100, 1) if pt['total'] > 0 else 0.0
        phases[phase] = {'total': pt['total'], 'done': pt['done'], 'pct': pct}
        total += pt['total']; done += pt['done']

    pct = round(done / total * 100, 1) if total > 0 else 0.0

    return {
        'phases': phases, 'total': total, 'done': done, 'pct': pct,
        'raf_inches': raf, 'surface_m2': surface, 'weld_hrs': weld_hrs, 'surface_hrs': surface_hrs,
    }

def bulk_spool_progress(project, settings=None):
    """Calculate hours-based progress for ALL spools at once."""
    if not settings: settings = get_project_settings(project)
    steps_def = get_project_steps(project)
    spools = db_fetchall("SELECT * FROM spools WHERE project=? ORDER BY sequence", (project,))
    all_progress = db_fetchall("SELECT spool_id, step_number, completed FROM progress WHERE project=?", (project,))
    completed_map = {}
    for r in all_progress:
        if r['completed']:
            if r['spool_id'] not in completed_map: completed_map[r['spool_id']] = set()
            completed_map[r['spool_id']].add(r['step_number'])
    result = {}
    for s in spools:
        sid = s['spool_id']
        done_steps = completed_map.get(sid, set())
        result[sid] = spool_hours(s, done_steps, settings, steps_def)
        result[sid]['spool'] = s
        result[sid]['done_steps'] = done_steps
    return result

def spool_progress(project, spool_id):
    """Hours-based progress for a single spool. Returns percentage."""
    settings = get_project_settings(project)
    steps_def = get_project_steps(project)
    sp = db_fetchone("SELECT * FROM spools WHERE project=? AND spool_id=?", (project, spool_id))
    if not sp: return 0.0
    rows = db_fetchall("SELECT step_number, completed FROM progress WHERE project=? AND spool_id=?", (project, spool_id))
    done_steps = set(r['step_number'] for r in rows if r['completed'])
    return spool_hours(sp, done_steps, settings, steps_def)['pct']

def project_stats(project):
    settings = get_project_settings(project)
    steps_def = get_project_steps(project)
    phase_order = get_phase_order(steps_def)
    bulk = bulk_spool_progress(project, settings)
    spools = [v for v in bulk.values()]
    total = len(spools)
    st = {'total':total,'completed':0,'in_progress':0,'not_started':0,'overall_pct':0.0,'by_diameter':{},'by_line':{},'phase_order':phase_order}
    tp = 0
    for v in spools:
        p = v['pct']; tp += p; s = v['spool']
        if p>=100: st['completed']+=1
        elif p>0: st['in_progress']+=1
        else: st['not_started']+=1
        d = s['main_diameter'] or '?'
        if d not in st['by_diameter']:
            st['by_diameter'][d] = {'total':0,'pct_sum':0,'phase_pct_sums':{ph:0 for ph in phase_order}}
        st['by_diameter'][d]['total']+=1; st['by_diameter'][d]['pct_sum']+=p
        for ph in phase_order:
            st['by_diameter'][d]['phase_pct_sums'][ph] += v['phases'].get(ph, {}).get('pct', 0)
        l = s['line'] or '?'
        if l not in st['by_line']: st['by_line'][l] = {'total':0,'pct_sum':0}
        st['by_line'][l]['total']+=1; st['by_line'][l]['pct_sum']+=p
    if total: st['overall_pct'] = round(tp/total,1)
    for v in st['by_diameter'].values():
        v['avg_pct'] = round(v['pct_sum']/v['total'],1) if v['total'] else 0
        v['phase_avgs'] = {}
        for ph in phase_order:
            v['phase_avgs'][ph] = round(v['phase_pct_sums'][ph]/v['total'],1) if v['total'] else 0
    for v in st['by_line'].values(): v['avg_pct'] = round(v['pct_sum']/v['total'],1) if v['total'] else 0
    return st

def fix_timestamps(rows):
    for r in rows:
        for k in ('completed_at','timestamp'):
            if r.get(k) and not isinstance(r[k], str): r[k] = str(r[k])
    return rows

def parse_date(val):
    """Parse date from various formats (ISO, PG timestamp, HTTP date)."""
    if not val: return None
    s = str(val).strip()[:10]
    try: return date.fromisoformat(s)
    except: pass
    for fmt in ('%a, %d %b %Y', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y'):
        try: return datetime.strptime(str(val).strip()[:30], fmt + '%f' if '.' in str(val) else fmt).date()
        except: pass
    try: return datetime.strptime(str(val).strip().split('.')[0].split('+')[0], '%Y-%m-%d %H:%M:%S').date()
    except: pass
    try:
        from email.utils import parsedate_to_datetime
        return parsedate_to_datetime(str(val)).date()
    except: pass
    return None

def schedule_status(project, bulk=None):
    """Calculate on-track/danger/delay status per diameter."""
    sched = db_fetchall("SELECT * FROM schedule WHERE project=? ORDER BY planned_start", (project,))
    if not sched: return None
    if not bulk: bulk = bulk_spool_progress(project)
    steps_def = get_project_steps(project)
    phase_order = get_phase_order(steps_def)
    settings = get_project_settings(project)
    diam_order = get_diameter_order(project)
    std_weeks = int(settings.get('standard_weeks', '9'))
    wks_saved = int(settings.get('committed_weeks_saved', '0'))
    days_saved = int(settings.get('committed_days_saved', '0'))
    total_saved = wks_saved * 7 + days_saved
    has_expediting = total_saved > 0
    spools = db_fetchall("SELECT * FROM spools WHERE project=? ORDER BY sequence", (project,))
    today = date.today()
    # Find production start
    prod_start = None
    for sc in sched:
        sd = parse_date(sc['planned_start'])
        if sd and (prod_start is None or sd < prod_start): prod_start = sd
    if prod_start and has_expediting:
        std_end = prod_start + timedelta(days=std_weeks * 7 - 1)
        committed_end = std_end - timedelta(days=total_saved)
    else:
        committed_end = None
    # Group spools by diameter
    by_diam = {}
    for s in spools:
        d = s['main_diameter'] or '?'
        if d not in by_diam: by_diam[d] = []
        by_diam[d].append(s)
    # Build schedule map
    sched_map = {}
    for sc in sched:
        d = sc['diameter']
        if d not in sched_map: sched_map[d] = {}
        sd = parse_date(sc['planned_start']); ed = parse_date(sc['planned_end'])
        sched_map[d][sc['task_type']] = {'start': str(sd) if sd else '', 'end': str(ed) if ed else '', 'spool_count': sc.get('spool_count',0), 'description': sc.get('description','')}
    fc = forecast_production(project, bulk)
    fc_diams = fc['diameters'] if fc else {}
    result = []
    overall_expected = 0; overall_actual = 0; overall_count = 0
    for diam in diam_order:
        dk = f'{diam}"'
        if dk not in sched_map and diam not in sched_map: continue
        sm = sched_map.get(dk, sched_map.get(diam, {}))
        # Build phase_dates from schedule map using phase names
        phase_dates = {}
        for ph in phase_order:
            ph_entry = sm.get(ph, {})
            if ph_entry:
                ps = parse_date(ph_entry.get('start')); pe = parse_date(ph_entry.get('end'))
                if ps and pe:
                    phase_dates[ph] = {'start': str(ps), 'end': str(pe)}
        if not phase_dates: continue
        first_phase = phase_order[0] if phase_order else None
        if not first_phase or first_phase not in phase_dates: continue
        try:
            total_start_d = min(parse_date(pd['start']) for pd in phase_dates.values())
            total_end_d = max(parse_date(pd['end']) for pd in phase_dates.values())
            target_end = committed_end if committed_end else total_end_d
            total_days = max(1, (target_end - total_start_d).days)
            elapsed = max(0, min((today - total_start_d).days, total_days))
            expected_pct = round(elapsed / total_days * 100, 1) if total_days > 0 else 0
        except: continue
        diam_spools = by_diam.get(dk, by_diam.get(f'{diam}"', []))
        if not diam_spools: continue
        actual_sum = sum(bulk.get(s['spool_id'], {}).get('pct', 0) for s in diam_spools)
        actual_pct = round(actual_sum / len(diam_spools), 1)
        # Per-phase averages (dynamic)
        phase_avgs = {}
        for ph in phase_order:
            ph_sum = sum(bulk.get(s['spool_id'], {}).get('phases', {}).get(ph, {}).get('pct', 0) for s in diam_spools)
            phase_avgs[ph] = round(ph_sum / len(diam_spools), 1)
        diff = actual_pct - expected_pct
        diam_fc = fc_diams.get(dk, {})
        fc_end = parse_date(diam_fc.get('forecast_end')) if diam_fc.get('forecast_end') else None
        if diam_fc.get('completed'):
            status = 'completed'
            diff = 0
        elif not diam_fc.get('started') and actual_pct == 0:
            status = 'not_started'
        elif fc_end and target_end:
            days_diff = (target_end - fc_end).days
            diff = days_diff
            if days_diff >= 0: status = 'on_time'
            elif days_diff >= -7: status = 'at_risk'
            else: status = 'delayed'
        else:
            if diff >= -5: status = 'on_time'
            elif diff >= -15: status = 'at_risk'
            else: status = 'delayed'
        overall_expected += expected_pct; overall_actual += actual_pct; overall_count += 1
        # First phase = fabrication for remaining RAF, last phase for remaining surface
        first_phase = phase_order[0] if phase_order else 'fab'
        remaining_raf = sum(bulk.get(s['spool_id'], {}).get('raf_inches', 0) for s in diam_spools if bulk.get(s['spool_id'], {}).get('phases', {}).get(first_phase, {}).get('pct', 0) < 100)
        remaining_m2 = sum(float(s.get('surface_m2') or 0) * 0.98 for s in diam_spools if bulk.get(s['spool_id'], {}).get('pct', 0) < 100)
        result.append({
            'diameter': dk, 'spool_count': len(diam_spools),
            'expected_pct': expected_pct, 'actual_pct': actual_pct, 'diff': round(diff, 1), 'status': status,
            'phase_avgs': phase_avgs,
            'remaining_raf': round(remaining_raf, 0), 'remaining_m2': round(remaining_m2, 1),
            'phase_dates': phase_dates,
            'total_start': str(total_start_d), 'total_end': str(total_end_d),
        })
    statuses = [r['status'] for r in result if r['status'] != 'not_started']
    if 'delayed' in statuses: overall_status = 'delayed'
    elif 'at_risk' in statuses: overall_status = 'at_risk'
    elif statuses and all(s == 'completed' for s in statuses): overall_status = 'completed'
    elif statuses: overall_status = 'on_time'
    else: overall_status = 'not_started'
    return {'diameters': result, 'overall_status': overall_status, 'today': str(today), 'phase_order': phase_order}

def daily_activity(project, day=None):
    if not day: day = date.today().strftime('%Y-%m-%d')
    if USE_PG:
        rows = db_fetchall("SELECT * FROM activity_log WHERE project=? AND timestamp::date = ?::date ORDER BY timestamp DESC", (project, day))
    else:
        rows = db_fetchall("SELECT * FROM activity_log WHERE project=? AND timestamp LIKE ? ORDER BY timestamp DESC", (project, f"{day}%"))
    return fix_timestamps(rows)

def forecast_production(project, bulk=None):
    """Forecast per diameter using actual throughput rates."""
    sched = db_fetchall("SELECT * FROM schedule WHERE project=? ORDER BY planned_start", (project,))
    if not sched: return None
    settings = get_project_settings(project)
    steps_def = get_project_steps(project)
    phase_order = get_phase_order(steps_def)
    diam_order = get_diameter_order(project)
    weld_cap = float(settings.get('welding_capability_ipd', '1000'))
    paint_cap = float(settings.get('surface_capability_m2d', '91'))
    if not bulk: bulk = bulk_spool_progress(project, settings)
    today = date.today()

    # Find welding and surface step numbers dynamically
    welding_steps = {s['step_number'] for s in steps_def if s['hours_variable'] == 'welding'}
    surface_steps = [s['step_number'] for s in steps_def if s['hours_variable'] == 'surface']
    last_surface_step = surface_steps[-1] if surface_steps else None

    prod_start = None
    for sc in sched:
        sd = parse_date(sc['planned_start'])
        if sd and (prod_start is None or sd < prod_start): prod_start = sd
    if not prod_start: prod_start = today
    global_days_elapsed = max(1, (today - prod_start).days)
    std_weeks = int(settings.get('standard_weeks', '9'))
    wks_saved = int(settings.get('committed_weeks_saved', '0'))
    days_saved = int(settings.get('committed_days_saved', '0'))
    total_saved = wks_saved * 7 + days_saved
    std_end = prod_start + timedelta(days=std_weeks * 7 - 1)
    committed_end = std_end - timedelta(days=total_saved) if total_saved > 0 else std_end

    # Per-diameter elapsed days
    if USE_PG:
        earliest_rows = db_fetchall(
            "SELECT s.main_diameter, MIN(p.completed_at) as first_activity "
            "FROM progress p JOIN spools s ON p.spool_id=s.spool_id AND p.project=s.project "
            "WHERE p.project=%s AND p.completed=1 GROUP BY s.main_diameter", (project,))
    else:
        earliest_rows = db_fetchall(
            "SELECT s.main_diameter, MIN(p.completed_at) as first_activity "
            "FROM progress p JOIN spools s ON p.spool_id=s.spool_id AND p.project=s.project "
            "WHERE p.project=? AND p.completed=1 GROUP BY s.main_diameter", (project,))
    diam_first_activity = {}
    for r in earliest_rows:
        d = r['main_diameter']; fa = parse_date(r['first_activity'])
        if d and fa: diam_first_activity[d] = fa

    by_diam = {}
    for sid, info in bulk.items():
        s = info['spool']; d = s['main_diameter'] or '?'
        if d not in by_diam: by_diam[d] = []
        by_diam[d].append(info)

    # Actual rates
    welded_raf_total = sum(i['raf_inches'] for i in bulk.values() if welding_steps & i.get('done_steps', set()))
    actual_weld_ipd = round(welded_raf_total / global_days_elapsed, 1) if global_days_elapsed > 0 else 0
    painted_m2_total = sum(i['surface_m2'] * 0.98 for i in bulk.values() if last_surface_step and last_surface_step in i.get('done_steps', set()))
    actual_paint_m2d = round(painted_m2_total / global_days_elapsed, 1) if global_days_elapsed > 0 else 0

    started_rates = []
    for diam in diam_order:
        dk = f'{diam}"'
        diam_infos = by_diam.get(dk, [])
        if not diam_infos: continue
        done_hrs = sum(i['done'] for i in diam_infos)
        if done_hrs > 0:
            first_activity = diam_first_activity.get(dk, prod_start)
            diam_elapsed = max(1, (today - first_activity).days)
            started_rates.append(done_hrs / diam_elapsed)
    avg_rate = sum(started_rates) / len(started_rates) if started_rates else None

    fc_result = {}
    overall_forecast = None
    for diam in diam_order:
        dk = f'{diam}"'
        diam_infos = by_diam.get(dk, [])
        if not diam_infos: continue
        total_hrs = sum(i['total'] for i in diam_infos)
        done_hrs = sum(i['done'] for i in diam_infos)
        remaining_hrs = total_hrs - done_hrs
        # Per-phase averages (dynamic)
        phase_avgs = {}
        for ph in phase_order:
            ph_sum = sum(i['phases'].get(ph, {}).get('pct', 0) for i in diam_infos)
            phase_avgs[ph] = round(ph_sum / len(diam_infos), 1)
        overall_pct = done_hrs / total_hrs * 100 if total_hrs > 0 else 0
        started = done_hrs > 0
        first_phase = phase_order[0] if phase_order else 'fab'
        remaining_raf = sum(i['raf_inches'] for i in diam_infos if i['phases'].get(first_phase, {}).get('pct', 0) < 100)
        remaining_m2 = sum(i['surface_m2'] * 0.98 for i in diam_infos if i['pct'] < 100)
        total_raf = sum(i['raf_inches'] for i in diam_infos)
        total_m2 = sum(i['surface_m2'] for i in diam_infos)

        is_completed = False
        if remaining_hrs <= 0 and started:
            # Diameter is production-complete — end date is a measured fact, not a forecast.
            # Query MAX(completed_at) from progress for this diameter's spools, restricted
            # to production-counting steps (counts_for_production=1).
            prod_step_nums = [s['step_number'] for s in steps_def if s.get('counts_for_production', 1)]
            spool_ids = [i['spool']['spool_id'] for i in diam_infos if i.get('spool')]
            last_date = None
            if spool_ids and prod_step_nums:
                ph_s = ','.join(['?'] * len(spool_ids))
                ph_p = ','.join(['?'] * len(prod_step_nums))
                row = db_fetchone(
                    f"SELECT MAX(completed_at) as last FROM progress "
                    f"WHERE project=? AND completed=1 AND spool_id IN ({ph_s}) AND step_number IN ({ph_p})",
                    tuple([project] + spool_ids + prod_step_nums))
                if row and row.get('last'):
                    last_date = parse_date(row['last'])
            forecast_end = last_date or today
            forecast_days = 0
            is_completed = True
        elif started:
            first_activity = diam_first_activity.get(dk, prod_start)
            diam_elapsed = max(1, (today - first_activity).days)
            actual_rate = done_hrs / diam_elapsed
            forecast_days = remaining_hrs / actual_rate if actual_rate > 0 else 999
            forecast_end = today + timedelta(days=max(1, int(forecast_days + 0.5)))
        elif avg_rate:
            forecast_days = total_hrs / avg_rate
            latest_start = committed_end - timedelta(days=int(forecast_days + 0.5))
            if today <= latest_start:
                forecast_end = committed_end; forecast_days = (committed_end - today).days
            else:
                forecast_end = today + timedelta(days=max(1, int(forecast_days + 0.5)))
        else:
            forecast_end = committed_end; forecast_days = (committed_end - today).days

        # Only floor to committed_end for in-progress forecasts. Completed diameters
        # report their actual completion date even if it precedes committed_end.
        if not is_completed and forecast_end < committed_end:
            forecast_end = committed_end; forecast_days = (committed_end - today).days

        fc_result[dk] = {
            'phase_avgs': phase_avgs,
            'overall_pct': round(overall_pct, 1), 'forecast_end': str(forecast_end),
            'forecast_days': round(forecast_days, 1),
            'total_hrs': round(total_hrs, 1), 'done_hrs': round(done_hrs, 1),
            'remaining_hrs': round(remaining_hrs, 1),
            'remaining_raf': round(remaining_raf, 0), 'remaining_m2': round(remaining_m2, 1),
            'total_raf': round(total_raf, 0), 'total_m2': round(total_m2, 1),
            'spool_count': len(diam_infos), 'started': started, 'completed': is_completed,
        }
        if forecast_end and (overall_forecast is None or forecast_end > overall_forecast):
            overall_forecast = forecast_end

    return {
        'diameters': fc_result, 'overall_forecast_end': str(overall_forecast) if overall_forecast else None,
        'today': str(today), 'welding_capability': weld_cap, 'painting_capability': paint_cap,
        'actual_weld_ipd': actual_weld_ipd, 'actual_paint_m2d': actual_paint_m2d,
        'days_elapsed': global_days_elapsed,
    }

def post_production_status(project, bulk=None, steps_def=None):
    """Per post-production step (counts_for_production=0): applicable spool count,
    completed count, completion %, last completion date. Generalised — no hardcoded
    step numbers or names. Returns list ordered by display_order."""
    if bulk is None: bulk = bulk_spool_progress(project)
    if steps_def is None: steps_def = get_project_steps(project)
    result = []
    for step in get_post_production_steps(steps_def):
        sn = step['step_number']
        applicable = [sid for sid, info in bulk.items()
                      if step_applies_to_spool(step, info.get('spool', {}))]
        done_ids = [sid for sid in applicable if sn in bulk.get(sid, {}).get('done_steps', set())]
        last_date = None
        if done_ids:
            ph = ','.join(['?'] * len(done_ids))
            row = db_fetchone(
                f"SELECT MAX(completed_at) as last FROM progress "
                f"WHERE project=? AND step_number=? AND completed=1 AND spool_id IN ({ph})",
                tuple([project, sn] + done_ids))
            if row and row.get('last'):
                last_date = str(row['last'])[:10]
        total = len(applicable)
        done = len(done_ids)
        result.append({
            'step_number': sn,
            'name_en': step.get('name_en', ''), 'name_cn': step.get('name_cn', ''),
            'total': total, 'completed': done,
            'pct': round(done / total * 100, 1) if total else 0.0,
            'last_date': last_date,
        })
    return result


def shipment_status(project, bulk=None, steps_def=None):
    """Per shipment: assigned spool count + per post-production step progress + derived
    status. The 'shipped' gate by convention is the LAST post-production step by
    display_order; 'packed' is the aggregate of all preceding pp steps. Zero hardcoded
    step numbers or names. Empty list if no shipments configured."""
    if bulk is None: bulk = bulk_spool_progress(project)
    if steps_def is None: steps_def = get_project_steps(project)
    pp_steps = get_post_production_steps(steps_def)
    shipments = db_fetchall("SELECT * FROM shipments WHERE project=? ORDER BY shipment_number", (project,))
    if not shipments: return []
    # Last pp step = "shipped" gate; all prior pp steps = "packed" gates.
    # When only one pp step exists, it serves as both pack and ship (a single gate).
    shipped_step = pp_steps[-1]['step_number'] if pp_steps else None
    pack_steps = ([s['step_number'] for s in pp_steps[:-1]] if len(pp_steps) > 1
                  else [s['step_number'] for s in pp_steps])
    result = []
    for ship in shipments:
        num = ship.get('shipment_number')
        desc = ship.get('description', '') or ''
        etd = str(ship.get('etd') or '')[:10] if ship.get('etd') else None
        transit = int(ship.get('transit_days') or 0)
        eta = None
        if etd and transit:
            try:
                etd_dt = parse_date(etd)
                if etd_dt:
                    eta = (etd_dt + timedelta(days=transit)).strftime('%Y-%m-%d')
            except Exception: pass
        # Assigned spools = spools with shipment_number matching this row
        assigned = [sid for sid, info in bulk.items()
                    if info.get('spool', {}).get('shipment_number') == num]
        assigned_count = len(assigned)
        # Packed = all pack_steps complete for that spool (or shipped-only model if only one pp step)
        packed_ids = []
        shipped_ids = []
        for sid in assigned:
            done = bulk.get(sid, {}).get('done_steps', set())
            if pack_steps and all(ps in done for ps in pack_steps):
                packed_ids.append(sid)
            if shipped_step and shipped_step in done:
                shipped_ids.append(sid)
        packed_count = len(packed_ids)
        shipped_count = len(shipped_ids)
        # Shipped date = MAX(completed_at) of shipped_step for shipped spools in this shipment
        shipped_date = None
        if shipped_ids and shipped_step:
            ph = ','.join(['?'] * len(shipped_ids))
            row = db_fetchone(
                f"SELECT MAX(completed_at) as last FROM progress "
                f"WHERE project=? AND step_number=? AND completed=1 AND spool_id IN ({ph})",
                tuple([project, shipped_step] + shipped_ids))
            if row and row.get('last'):
                shipped_date = str(row['last'])[:10]
        # Derived status
        if assigned_count == 0:
            status = 'unassigned'
        elif shipped_count == assigned_count:
            status = 'shipped'
        elif shipped_count > 0:
            status = 'partial'
        elif packed_count == assigned_count:
            status = 'ready'
        else:
            status = 'pending'
        result.append({
            'shipment_number': num,
            'description': desc,
            'etd': etd, 'transit_days': transit, 'eta': eta,
            'assigned': assigned_count,
            'packed': packed_count,
            'shipped': shipped_count,
            'shipped_date': shipped_date,
            'status': status,
        })
    return result


def past_hold_point_count(project):
    """Count spools that have passed the hold point (RT or equivalent)."""
    steps_def = get_project_steps(project)
    hold_steps = [s['step_number'] for s in steps_def if s.get('is_hold_point')]
    if not hold_steps: return 0
    hs = hold_steps[0]
    row = db_fetchone("SELECT COUNT(*) as cnt FROM progress WHERE project=? AND step_number=? AND completed=1", (project, hs))
    return row['cnt'] if row else 0

def daily_production_rate(project):
    """Calculate 7-day average production rate and today's steps."""
    today = date.today()
    week_ago = today - timedelta(days=7)
    two_weeks_ago = today - timedelta(days=14)
    if USE_PG:
        this_week_spools = db_fetchall("SELECT COUNT(DISTINCT spool_id) as cnt FROM activity_log WHERE project=? AND action='completed' AND timestamp::date >= ?::date AND timestamp::date <= ?::date", (project, str(week_ago), str(today)))
        last_week_spools = db_fetchall("SELECT COUNT(DISTINCT spool_id) as cnt FROM activity_log WHERE project=? AND action='completed' AND timestamp::date >= ?::date AND timestamp::date < ?::date", (project, str(two_weeks_ago), str(week_ago)))
        today_steps = db_fetchall("SELECT COUNT(*) as cnt FROM activity_log WHERE project=? AND action='completed' AND timestamp::date = ?::date", (project, str(today)))
        today_spools = db_fetchall("SELECT COUNT(DISTINCT spool_id) as cnt FROM activity_log WHERE project=? AND action='completed' AND timestamp::date = ?::date", (project, str(today)))
    else:
        this_week_spools = db_fetchall("SELECT COUNT(DISTINCT spool_id) as cnt FROM activity_log WHERE project=? AND action='completed' AND timestamp >= ? AND timestamp <= ?", (project, str(week_ago), str(today) + ' 23:59:59'))
        last_week_spools = db_fetchall("SELECT COUNT(DISTINCT spool_id) as cnt FROM activity_log WHERE project=? AND action='completed' AND timestamp >= ? AND timestamp < ?", (project, str(two_weeks_ago), str(week_ago)))
        today_steps = db_fetchall("SELECT COUNT(*) as cnt FROM activity_log WHERE project=? AND action='completed' AND timestamp LIKE ?", (project, str(today) + '%'))
        today_spools = db_fetchall("SELECT COUNT(DISTINCT spool_id) as cnt FROM activity_log WHERE project=? AND action='completed' AND timestamp LIKE ?", (project, str(today) + '%'))
    this_wk = this_week_spools[0]['cnt'] if this_week_spools else 0
    last_wk = last_week_spools[0]['cnt'] if last_week_spools else 0
    today_cnt = today_steps[0]['cnt'] if today_steps else 0
    today_sp = today_spools[0]['cnt'] if today_spools else 0
    avg_7day = round(this_wk / 7, 1)
    avg_prev = round(last_wk / 7, 1)
    trend = round(avg_7day - avg_prev, 1)
    return {'avg_7day': avg_7day, 'avg_prev_week': avg_prev, 'trend': trend, 'today_steps': today_cnt, 'today_spools': today_sp}

def generate_report_data(project):
    """Build full report data for a project."""
    settings = get_project_settings(project)
    steps_def = get_project_steps(project)
    bulk = bulk_spool_progress(project, settings)
    st = project_stats(project)
    sched = schedule_status(project, bulk)
    forecast = forecast_production(project, bulk)
    post_prod = post_production_status(project, bulk, steps_def)
    ship_data = shipment_status(project, bulk, steps_def)
    today = date.today().strftime('%Y-%m-%d')
    today_act = daily_activity(project, today)
    steps_today = len([a for a in today_act if a.get('action') == 'completed'])
    # Dynamic release step detection
    release_steps = {s['step_number'] for s in steps_def if s.get('is_release')}
    hold_steps = {s['step_number'] for s in steps_def if s.get('is_hold_point')}
    released_today = len([a for a in today_act if a.get('action') == 'completed' and a.get('step_number') in release_steps])
    rt_count = past_hold_point_count(project)
    prod_rate = daily_production_rate(project)
    spool_pcts = {sid: info['pct'] for sid, info in bulk.items()}
    step_names = {s['step_number']: s['name_en'] for s in steps_def}
    completed_spools = [sid for sid, info in bulk.items() if info['pct'] >= 100]
    return {
        'project': project, 'date': today, 'stats': st, 'schedule': sched,
        'today_activity': today_act, 'steps_completed_today': steps_today,
        'completed_spools': completed_spools, 'total_spools': len(bulk),
        'settings': settings, 'forecast': forecast, 'past_rt': rt_count,
        'production_rate': prod_rate, 'spool_progress': spool_pcts,
        'step_names': step_names, 'released_today': released_today,
        'hold_steps': list(hold_steps), 'release_steps': list(release_steps),
        'phase_order': get_phase_order(steps_def),
        'post_production': post_prod, 'shipment_status': ship_data,
    }

# ── Chat Agent — ENERXON Production & Quality Assistant ─────────────────────
# Generalised: every tool takes `project` as parameter. Knowledge file path is
# derived from project id — new projects drop in knowledge/<project>_qc_knowledge.md
# with zero code change. No hardcoded project IDs anywhere in this section.

_KNOWLEDGE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'knowledge')

def _load_knowledge_file(project, lang='en'):
    """Load the project's QC knowledge markdown file(s).

    Generalised bilingual loader:
    - For lang='zh', try `{project}_qc_knowledge_cn.md` first, then append
      `{project}_qc_knowledge.md` as fallback for sections not yet translated.
    - For lang='en' (default), load only `{project}_qc_knowledge.md`.
    - Returns None if no file exists for the project.

    Any project can add a `_cn.md` file at any time with zero code change.
    Untranslated sections automatically fall back to the English file, so a
    partial translation is always better than none.
    """
    def _read(path):
        if not os.path.exists(path):
            return None
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception:
            return None

    en_path = os.path.join(_KNOWLEDGE_DIR, f'{project}_qc_knowledge.md')
    if lang == 'zh':
        cn_path = os.path.join(_KNOWLEDGE_DIR, f'{project}_qc_knowledge_cn.md')
        cn = _read(cn_path)
        en = _read(en_path)
        # Put CN first so keyword search returns Chinese sections before
        # falling through to the English equivalents.
        if cn and en:
            return cn + '\n\n' + en
        return cn or en
    return _read(en_path)

def _split_sections(md):
    """Split a markdown string into (heading, body) tuples at every ## or ### heading."""
    if not md:
        return []
    lines = md.split('\n')
    sections = []
    cur_head = None; cur_body = []
    for line in lines:
        if line.startswith('## '):
            if cur_head is not None:
                sections.append((cur_head, '\n'.join(cur_body).strip()))
            cur_head = line.lstrip('# ').strip()
            cur_body = []
        else:
            cur_body.append(line)
    if cur_head is not None:
        sections.append((cur_head, '\n'.join(cur_body).strip()))
    return sections

def _knowledge_section(project, *keywords, limit=2, max_chars_per_section=4000):
    """Return sections of the project knowledge file whose heading contains any keyword.
    Case-insensitive substring match. Limits to `limit` sections and caps each
    section to `max_chars_per_section` chars to stay within API token budgets.
    Language for the knowledge file is read from Flask request-local g.chat_lang
    (set by _run_chat_turn). Defaults to 'en' if not set (e.g. when called outside
    a chat request)."""
    lang = getattr(g, 'chat_lang', 'en')
    md = _load_knowledge_file(project, lang=lang)
    if md is None:
        return f"No QC knowledge base is configured for project {project}. Ask the QC manager to add knowledge/{project}_qc_knowledge.md."
    sections = _split_sections(md)
    if not sections:
        return "Knowledge file is empty or malformed."
    kws = [k.lower() for k in keywords if k]

    def _pack(head, body):
        # Cap each section so one matched ## block cannot consume the entire
        # token budget. Large sections are trimmed with an explicit note.
        if len(body) > max_chars_per_section:
            body = body[:max_chars_per_section] + f"\n\n[... section truncated at {max_chars_per_section} chars — ask a more specific follow-up for additional detail ...]"
        return f"## {head}\n\n{body}"

    matched = []
    for head, body in sections:
        hl = head.lower()
        if not kws or any(k in hl for k in kws):
            matched.append(_pack(head, body))
            if len(matched) >= limit:
                break
    if not matched:
        # Fallback: search body text as well (covers Q&A and glossary)
        for head, body in sections:
            bl = body.lower()
            if any(k in bl for k in kws):
                matched.append(_pack(head, body))
                if len(matched) >= limit:
                    break
    if not matched:
        return f"No section matching {list(keywords)} in the {project} knowledge base."
    return '\n\n---\n\n'.join(matched)

# ── Chat tool implementations — all read-only, all take `project` ────────────

def tool_get_material_and_metallurgy(project, **_):
    return _knowledge_section(project, 'material', 'metallurgy', 'acceptance')

def tool_get_welding_procedure(project, wps_number=None, **_):
    if wps_number:
        return _knowledge_section(project, wps_number, 'welding')
    return _knowledge_section(project, 'welding', 'wps')

def tool_get_ndt_procedure(project, ndt_type=None, **_):
    if not ndt_type:
        return _knowledge_section(project, 'ndt', 'non-destructive')
    return _knowledge_section(project, ndt_type, 'ndt')

def tool_get_acceptance_criteria(project, topic=None, **_):
    if not topic:
        return _knowledge_section(project, 'acceptance')
    return _knowledge_section(project, topic, 'acceptance')

def tool_get_itp_flow(project, **_):
    return _knowledge_section(project, 'itp', 'inspection and test plan', 'hold point')

def tool_get_finishing_process(project, **_):
    return _knowledge_section(project, 'finishing', 'surface', 'station')

def tool_get_cutting_and_fitup(project, **_):
    return _knowledge_section(project, 'cutting', 'fit-up', 'fitup', 'bevel')

def tool_get_report_requirements(project, report_type=None, **_):
    if report_type:
        return _knowledge_section(project, report_type, 'report')
    return _knowledge_section(project, 'documentation', 'report', 'certification')

def tool_get_standards_and_codes(project, **_):
    return _knowledge_section(project, 'project overview', 'standard', 'code')

def tool_get_qc_knowledge(project, topic=None, **_):
    if not topic:
        return "Please provide a topic keyword (e.g. 'ferrite', 'RT', 'WPS-001', 'ferroxyl')."
    return _knowledge_section(project, topic, limit=2)

def tool_get_project_stats(project, **_):
    try:
        st = project_stats(project)
    except Exception as e:
        return f"Could not load stats for {project}: {e}"
    out = {
        'project': project,
        'total_spools': st['total'],
        'completed': st['completed'],
        'in_progress': st['in_progress'],
        'not_started': st['not_started'],
        'overall_pct': st['overall_pct'],
        'phase_order': st['phase_order'],
        'by_diameter': {d: {'spools': v['total'], 'avg_pct': v['avg_pct'], 'phase_avgs': v['phase_avgs']}
                        for d, v in st['by_diameter'].items()},
    }
    return json.dumps(out, ensure_ascii=False)

def tool_get_spool_status(project, spool_id=None, **_):
    if not spool_id:
        return "Please provide a spool_id (e.g. 'SPL-045')."
    sp = db_fetchone("SELECT * FROM spools WHERE project=? AND spool_id=?", (project, spool_id))
    if not sp:
        return f"Spool {spool_id} not found in project {project}."
    prog = db_fetchall("SELECT p.step_number, p.completed, p.completed_at, p.completed_by, s.name_en, s.name_cn, s.phase FROM progress p LEFT JOIN project_steps s ON s.project=p.project AND s.step_number=p.step_number WHERE p.project=? AND p.spool_id=? ORDER BY p.step_number", (project, spool_id))
    qc = db_fetchall("SELECT report_type, status, inspector_name, inspector_date FROM qc_reports WHERE project=? AND spool_id=?", (project, spool_id))
    return json.dumps({
        'spool_id': spool_id,
        'marking': sp.get('marking',''),
        'iso_no': sp.get('iso_no',''),
        'main_diameter': sp.get('main_diameter',''),
        'spool_type': sp.get('spool_type','SPOOL'),
        'steps': [{'step': r['step_number'], 'name_en': r.get('name_en',''), 'name_cn': r.get('name_cn',''), 'phase': r.get('phase',''), 'completed': bool(r['completed']), 'completed_at': str(r.get('completed_at') or '')[:19], 'completed_by': r.get('completed_by','')} for r in fix_timestamps(prog)],
        'qc_reports': [{'type': r['report_type'], 'status': r['status'], 'inspector': r.get('inspector_name',''), 'date': r.get('inspector_date','')} for r in qc],
    }, ensure_ascii=False)

def tool_get_qc_report_status(project, report_type=None, status=None, **_):
    q = "SELECT spool_id, report_type, status, inspector_name, inspector_date FROM qc_reports WHERE project=?"
    params = [project]
    if report_type:
        q += " AND report_type=?"; params.append(report_type)
    if status:
        q += " AND status=?"; params.append(status)
    q += " ORDER BY spool_id LIMIT 200"
    rows = db_fetchall(q, tuple(params))
    if not rows:
        return f"No QC reports found in {project}" + (f" for type={report_type}" if report_type else "") + (f" with status={status}" if status else "") + "."
    return json.dumps([{'spool': r['spool_id'], 'type': r['report_type'], 'status': r['status'], 'inspector': r.get('inspector_name','') or '', 'date': r.get('inspector_date','') or ''} for r in rows], ensure_ascii=False)

def tool_get_recent_activity(project, days=3, **_):
    try:
        days = int(days)
    except Exception:
        days = 3
    days = max(1, min(days, 30))
    rows = db_fetchall("SELECT spool_id, step_number, action, operator, timestamp, details FROM activity_log WHERE project=? ORDER BY timestamp DESC LIMIT 100", (project,))
    # Filter in Python to avoid PG/SQLite date math differences
    from datetime import datetime as _dt
    cutoff = _dt.now() - timedelta(days=days)
    out = []
    for r in fix_timestamps(rows):
        ts = r.get('timestamp') or ''
        try:
            t = _dt.strptime(str(ts)[:19], '%Y-%m-%d %H:%M:%S')
            if t < cutoff: continue
        except Exception:
            pass
        out.append({'spool': r['spool_id'], 'step': r.get('step_number'), 'action': r['action'], 'operator': r.get('operator',''), 'timestamp': str(ts)[:19], 'details': r.get('details','')})
        if len(out) >= 50: break
    if not out:
        return f"No activity recorded in project {project} in the last {days} days."
    return json.dumps(out, ensure_ascii=False)

# ── Chat tool registry: name → (function, Anthropic schema) ──────────────────
# Single source of truth. Adding a tool = one entry. No hardcoded per-project logic.

CHAT_TOOLS = {
    'get_material_and_metallurgy': (tool_get_material_and_metallurgy, {
        'name': 'get_material_and_metallurgy',
        'description': 'Get the project material, metallurgy rules, and acceptance ranges (base/filler material, phase balance, ferrite band, PMI criteria, preheat, interpass, heat input limits, intermetallic screen, chloride-free chain).',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string', 'description': 'Project ID, e.g. ENJOB25011423'}}, 'required': ['project']},
    }),
    'get_welding_procedure': (tool_get_welding_procedure, {
        'name': 'get_welding_procedure',
        'description': 'Get the welding procedures (WPS) for the project, including selection rules, pass tables, welder qualification (WPQ), and weld counting rules. Optionally filter by a specific WPS number.',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}, 'wps_number': {'type': 'string', 'description': 'Optional WPS identifier, e.g. WPS-001 or WPS-002'}}, 'required': ['project']},
    }),
    'get_ndt_procedure': (tool_get_ndt_procedure, {
        'name': 'get_ndt_procedure',
        'description': 'Get the full NDT procedure and acceptance criteria for a given test type. Use this for questions about VT, RT, PT, MT, PMI, Ferrite, Ferroxyl, Dimensional inspection, or Metallographic (A923) testing.',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}, 'ndt_type': {'type': 'string', 'description': 'Test type keyword: vt, rt, pt, mt, pmi, ferrite, ferroxyl, dimensional, metallographic'}}, 'required': ['project', 'ndt_type']},
    }),
    'get_acceptance_criteria': (tool_get_acceptance_criteria, {
        'name': 'get_acceptance_criteria',
        'description': 'Look up accept/reject acceptance criteria for any QC topic (ferrite band, PMI Cr/Mo, DFT, hi-lo, back-purge O2, interpass, etc).',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}, 'topic': {'type': 'string', 'description': 'Keyword e.g. ferrite, pmi, dft, hi-lo, interpass'}}, 'required': ['project', 'topic']},
    }),
    'get_itp_flow': (tool_get_itp_flow, {
        'name': 'get_itp_flow',
        'description': 'Get the Inspection and Test Plan (ITP) step flow for the project, including hold points, witness points, and review points.',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}}, 'required': ['project']},
    }),
    'get_finishing_process': (tool_get_finishing_process, {
        'name': 'get_finishing_process',
        'description': 'Get the post-weld finishing and surface treatment process, including station flow, sequence, and key technical decisions (pickling, passivation, ferroxyl, photo, marking, packing).',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}}, 'required': ['project']},
    }),
    'get_cutting_and_fitup': (tool_get_cutting_and_fitup, {
        'name': 'get_cutting_and_fitup',
        'description': 'Get the cutting, end-preparation (beveling), and fit-up rules for the project, including back-purge O2 limit and hi-lo tolerance.',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}}, 'required': ['project']},
    }),
    'get_report_requirements': (tool_get_report_requirements, {
        'name': 'get_report_requirements',
        'description': 'Get the mandatory fields, signatures, and format requirements for QC reports (or a specific report type).',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}, 'report_type': {'type': 'string', 'description': 'Optional: cutting, fitup, welding, vt, rt, pt, pmi, ferrite, dimensional, ferroxyl, dft'}}, 'required': ['project']},
    }),
    'get_standards_and_codes': (tool_get_standards_and_codes, {
        'name': 'get_standards_and_codes',
        'description': 'List the applicable standards and codes for the project (ASME B31.3, Section V, A923, A380, NORSOK M-601/M-630, etc) and what each covers.',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}}, 'required': ['project']},
    }),
    'get_qc_knowledge': (tool_get_qc_knowledge, {
        'name': 'get_qc_knowledge',
        'description': 'Fallback keyword search across the full project QC knowledge base. Use only if no other tool fits the topic.',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}, 'topic': {'type': 'string', 'description': 'Keyword to search for'}}, 'required': ['project', 'topic']},
    }),
    'get_project_stats': (tool_get_project_stats, {
        'name': 'get_project_stats',
        'description': 'Get overall production progress for the project: total spools, completed, in-progress, overall %, and per-diameter breakdown with per-phase averages.',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}}, 'required': ['project']},
    }),
    'get_spool_status': (tool_get_spool_status, {
        'name': 'get_spool_status',
        'description': 'Get the step-by-step status of a specific spool, including all checklist steps, QC reports, and completion dates.',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}, 'spool_id': {'type': 'string', 'description': 'Spool identifier, e.g. SPL-045'}}, 'required': ['project', 'spool_id']},
    }),
    'get_qc_report_status': (tool_get_qc_report_status, {
        'name': 'get_qc_report_status',
        'description': 'List QC reports across the project. Filter by report_type or status (draft, submitted, accepted, rejected) to find e.g. which spools failed RT.',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}, 'report_type': {'type': 'string'}, 'status': {'type': 'string'}}, 'required': ['project']},
    }),
    'get_recent_activity': (tool_get_recent_activity, {
        'name': 'get_recent_activity',
        'description': 'Get the recent activity log for the project (step completions, QC updates) within the last N days (default 3, max 30).',
        'input_schema': {'type': 'object', 'properties': {'project': {'type': 'string'}, 'days': {'type': 'integer'}}, 'required': ['project']},
    }),
}

CHAT_SYSTEM_PROMPT = """You are the ENERXON Production and Quality Assistant for the ENERXON China Tracker website.

YOU HELP with:
- Production progress (spools, diameters, phase progress, recent activity)
- Quality control procedures (WPS/PQR, NDT: VT/RT/PT/MT/PMI, ferrite, ferroxyl, dimensional, cutting, fit-up, metallographic)
- Testing and acceptance criteria per the project's standards (ASME B31.3, ASME Section V, ASTM A923/A380, NORSOK, etc.)
- ITP flow, hold/witness points, finishing process, report requirements

RULES:
1. ONLY use the provided tools to answer. Never invent data, numbers, or procedures.
2. For standards, procedures, acceptance criteria, welding, NDT, materials, ITP, finishing - CALL a knowledge tool (get_material_and_metallurgy, get_welding_procedure, get_ndt_procedure, get_acceptance_criteria, get_itp_flow, get_finishing_process, get_cutting_and_fitup, get_report_requirements, get_standards_and_codes, or get_qc_knowledge as fallback).
3. For live progress, spools, QC report status, or recent activity - CALL a data tool (get_project_stats, get_spool_status, get_qc_report_status, get_recent_activity).
4. Always pass the current project ID to every tool call. The project ID is provided in the first user message.
5. Be CONCISE. Cite spool IDs, report types, ITP steps, or standards when relevant.
6. If a tool returns empty, "No section matching...", or "No QC knowledge base...", you MUST tell the user that no data is available in the project knowledge base. Do NOT guess, do NOT provide values from general engineering knowledge, do NOT cite typical industry ranges. If the knowledge base is missing, tell them to contact the QC manager and stop.
7. If the question is outside production or quality (pricing, contracts, suppliers, customers, HR, logistics costs), politely decline in the same language the user used and explain you only answer production and quality questions.
8. Do not expose internal tool names to the user in the answer text.
9. Match the exact language of the user's most recent message. A language instruction is embedded at the top of each user message - follow it exactly."""

def _detect_language(text):
    """Return 'zh' if the text contains any CJK character, else 'en'. Deterministic, no LLM."""
    if not text:
        return 'zh'
    for ch in text:
        if '\u4e00' <= ch <= '\u9fff':
            return 'zh'
    return 'en'

def _strip_chinese(text):
    """Remove CJK characters from text and clean up stray punctuation/whitespace left by
    bilingual labels like 'Ferrite Measurement (铁素体测量)'. Used to remove Chinese
    anchoring context from tool responses when the user wrote in English, so the model
    doesn't drift to Chinese. Pure string operation, no cost."""
    if not text:
        return text
    import re
    s = re.sub(r'[\u4e00-\u9fff\u3000-\u303f\uff00-\uffef]+', '', text)  # CJK + fullwidth
    s = re.sub(r'\(\s*[,，、]?\s*\)', '', s)      # empty parens left by bilingual labels
    s = re.sub(r'（\s*[,，、]?\s*）', '', s)       # empty fullwidth parens
    s = re.sub(r'[ \t]+', ' ', s)                 # collapse horizontal whitespace
    s = re.sub(r' +\n', '\n', s)                  # trailing spaces on lines
    s = re.sub(r'\n{3,}', '\n\n', s)              # collapse blank lines
    return s.strip()

def _run_chat_turn(project, message, history):
    """Run one chat turn through Anthropic with tool use. Returns (reply_text, tools_used_list)."""
    try:
        import anthropic
    except ImportError:
        return ("服务未就绪：Anthropic SDK 未安装。请联系管理员。\n(Chat service not ready: Anthropic SDK missing.)", [])
    api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    if not api_key:
        return ("服务未就绪：ANTHROPIC_API_KEY 未配置。请联系管理员。\n(Chat service not ready: API key not configured.)", [])

    model = get_qc_setting(project, 'chat_model') or 'claude-haiku-4-5'
    client = anthropic.Anthropic(api_key=api_key)

    tool_schemas = [schema for (_fn, schema) in CHAT_TOOLS.values()]

    # Detect language and make it available to the knowledge tools via g.
    # With pre-localised knowledge files (<project>_qc_knowledge_cn.md),
    # tool responses come back in the correct language already, so we only
    # need a light language instruction to nudge the model; the heavy
    # strip-Chinese + strong-directive machinery is kept as a fallback for
    # projects that have not added a CN file yet.
    lang = _detect_language(message)
    g.chat_lang = lang
    if lang == 'en':
        lang_instruction = "[Reply in ENGLISH ONLY. Do not use any Chinese characters. English only.]"
    else:
        lang_instruction = "[使用简体中文回答。]"

    # Build message list: trimmed history + project-tagged user message
    msgs = []
    for h in (history or [])[-10:]:
        role = h.get('role'); content = h.get('content')
        if role in ('user', 'assistant') and isinstance(content, str) and content.strip():
            msgs.append({'role': role, 'content': content})
    msgs.append({'role': 'user', 'content': f"[Current project: {project}]\n{lang_instruction}\n\n{message}"})

    tools_used = []
    for _iter in range(8):  # safety cap on tool-use loops
        try:
            resp = client.messages.create(
                model=model,
                max_tokens=1500,
                system=CHAT_SYSTEM_PROMPT,
                tools=tool_schemas,
                messages=msgs,
            )
        except anthropic.RateLimitError:
            msg_zh = "系统繁忙：Anthropic API 速率限制。请等待 1 分钟后重试。"
            msg_en = "System busy: Anthropic API rate limit reached. Please wait 1 minute and try again."
            return (msg_zh + "\n\n" + msg_en, tools_used)
        except anthropic.APIStatusError as e:
            status = getattr(e, 'status_code', 'unknown')
            return (f"API error ({status}): please try again. / 服务异常（{status}），请稍后重试。", tools_used)
        except anthropic.APIConnectionError:
            return ("Network error connecting to Anthropic. / 连接 Anthropic 服务失败，请稍后重试。", tools_used)
        except Exception as e:
            return (f"Unexpected error: {type(e).__name__}. Please try again. / 系统异常（{type(e).__name__}），请稍后重试。", tools_used)
        if resp.stop_reason != 'tool_use':
            # Collect final text blocks
            text_out = ''.join([b.text for b in resp.content if getattr(b, 'type', None) == 'text'])
            return (text_out.strip() or '(no response)', tools_used)
        # Execute tool calls
        msgs.append({'role': 'assistant', 'content': resp.content})
        tool_results = []
        for block in resp.content:
            if getattr(block, 'type', None) == 'tool_use':
                tool_name = block.name
                tool_input = block.input or {}
                tools_used.append(tool_name)
                tool_input['project'] = project  # enforce project scoping, ignore any model-supplied project
                fn_schema = CHAT_TOOLS.get(tool_name)
                if not fn_schema:
                    result = f"Unknown tool: {tool_name}"
                else:
                    try:
                        result = fn_schema[0](**tool_input)
                    except Exception as e:
                        result = f"Tool error: {e}"
                result_str = str(result)
                if lang == 'en':
                    result_str = _strip_chinese(result_str)
                tool_results.append({'type': 'tool_result', 'tool_use_id': block.id, 'content': result_str[:20000]})
        msgs.append({'role': 'user', 'content': tool_results})
    return ("(Tool-use loop limit reached. Please rephrase your question.)", tools_used)

@app.route('/api/chat', methods=['POST'])
@login_required
def api_chat():
    body = request.get_json(silent=True) or {}
    project = (body.get('project') or '').strip()
    message = (body.get('message') or '').strip()
    history = body.get('history') or []
    if not project or not message:
        return jsonify({'error': 'project and message are required'}), 400
    reply, tools_used = _run_chat_turn(project, message, history)
    # Log to DB
    log_id = None
    try:
        db_execute("INSERT INTO chat_log (project, user_msg, assistant_msg, tools_used) VALUES (?, ?, ?, ?)",
                   (project, message, reply, json.dumps(tools_used, ensure_ascii=False)))
        db_commit()
        row = db_fetchone("SELECT id FROM chat_log WHERE project=? ORDER BY id DESC LIMIT 1", (project,))
        if row: log_id = row['id']
    except Exception as e:
        print(f"chat_log insert failed: {e}")
    return jsonify({'reply': reply, 'tools_used': tools_used, 'log_id': log_id})

@app.route('/api/chat/feedback', methods=['POST'])
@login_required
def api_chat_feedback():
    body = request.get_json(silent=True) or {}
    log_id = body.get('log_id')
    fb = (body.get('feedback') or '').strip()
    if not log_id or fb not in ('up', 'down', ''):
        return jsonify({'error': 'log_id and feedback in (up|down|) required'}), 400
    try:
        db_execute("UPDATE chat_log SET feedback=? WHERE id=?", (fb, int(log_id)))
        db_commit()
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    return jsonify({'ok': True})

# ── API: Projects ─────────────────────────────────────────────────────────────
@app.route('/')
@login_required
def home():
    return render_template_string(HOME_HTML)

@app.route('/api/projects')
@login_required
def api_projects():
    rows = db_fetchall("SELECT project, COUNT(*) as spool_count FROM spools GROUP BY project ORDER BY project")
    result = []
    for r in rows:
        p = r['project']
        st = project_stats(p)
        result.append({'project': p, 'spool_count': r['spool_count'], 'overall_pct': st['overall_pct'],
                        'completed': st['completed'], 'in_progress': st['in_progress'], 'not_started': st['not_started']})
    return jsonify(result)

# ── API: Project Dashboard ────────────────────────────────────────────────────
@app.route('/project/<project>')
@login_required
def project_page(project):
    return render_template_string(PROJECT_HTML, project=project)

@app.route('/api/project/<project>/dashboard')
@login_required
def api_project_dashboard(project):
    import traceback as tb
    try:
        settings = get_project_settings(project)
        bulk = bulk_spool_progress(project, settings)
        st = project_stats(project)
        act = db_fetchall("SELECT * FROM activity_log WHERE project=? ORDER BY timestamp DESC LIMIT 20", (project,))
        st['recent_activity'] = fix_timestamps(act)
        st['settings'] = settings
        st['forecast'] = forecast_production(project, bulk)
        st['past_rt'] = past_hold_point_count(project)
        st['production_rate'] = daily_production_rate(project)
        st['schedule_data'] = schedule_status(project, bulk)
        return jsonify(st)
    except Exception as e:
        print(f"Dashboard error: {e}\n{tb.format_exc()}")
        return jsonify({'error': str(e), 'trace': tb.format_exc()}), 500

@app.route('/api/project/<project>/spools')
@login_required
def api_project_spools(project):
    spools = db_fetchall("SELECT * FROM spools WHERE project=? ORDER BY sequence", (project,))
    result = []
    for s in spools:
        p = spool_progress(project, s['spool_id'])
        result.append({'spool': s, 'progress_pct': p})
    return jsonify(result)

# ── API: Spool Detail ─────────────────────────────────────────────────────────
@app.route('/project/<project>/spool/<spool_id>')
@login_required
def spool_page(project, spool_id):
    return render_template_string(SPOOL_HTML, project=project, spool_id=spool_id)

@app.route('/api/project/<project>/spool/<spool_id>')
@login_required
def api_spool(project, spool_id):
    sp = db_fetchone("SELECT * FROM spools WHERE project=? AND spool_id=?", (project, spool_id))
    if not sp: return jsonify({'error':'Not found'}), 404
    steps_def = get_project_steps(project)
    steps = fix_timestamps(db_fetchall("SELECT * FROM progress WHERE project=? AND spool_id=? ORDER BY step_number", (project, spool_id)))
    act = fix_timestamps(db_fetchall("SELECT * FROM activity_log WHERE project=? AND spool_id=? ORDER BY timestamp DESC LIMIT 10", (project, spool_id)))
    spool_type = sp.get('spool_type', 'SPOOL') or 'SPOOL'
    step_definitions = [
        {'number':s['step_number'],'name_en':s['name_en'],'name_cn':s['name_cn'],'weight':s['weight'],
         'is_hold_point':s.get('is_hold_point',0),'is_release':s.get('is_release',0)}
        for s in steps_def
        if (not s['is_conditional'] or sp.get('has_branches'))
        and ((s.get('spool_type','ALL') or 'ALL') in ('ALL', spool_type))
    ]
    return jsonify({'spool':sp, 'progress_pct': spool_progress(project, spool_id), 'steps':steps, 'activity':act,
        'step_definitions': step_definitions,
        'has_branches': bool(sp.get('has_branches')) if sp else False})

@app.route('/api/project/<project>/spool/<spool_id>/step/<int:step>', methods=['POST'])
@login_required
def api_update_step(project, spool_id, step):
    d = request.get_json() or {}
    comp = 1 if d.get('completed') else 0
    op = d.get('operator',''); rem = d.get('remarks','')
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if USE_PG:
        db_execute("INSERT INTO progress (project,spool_id,step_number,completed,completed_by,completed_at,remarks) VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (project,spool_id,step_number) DO UPDATE SET completed=EXCLUDED.completed, completed_by=EXCLUDED.completed_by, completed_at=CASE WHEN EXCLUDED.completed=1 THEN EXCLUDED.completed_at ELSE NULL END, remarks=EXCLUDED.remarks",
            (project, spool_id, step, comp, op, now if comp else None, rem))
    else:
        db_execute("INSERT INTO progress (project,spool_id,step_number,completed,completed_by,completed_at,remarks) VALUES (?,?,?,?,?,?,?) ON CONFLICT(project,spool_id,step_number) DO UPDATE SET completed=excluded.completed, completed_by=excluded.completed_by, completed_at=CASE WHEN excluded.completed=1 THEN excluded.completed_at ELSE NULL END, remarks=excluded.remarks",
            (project, spool_id, step, comp, op, now if comp else None, rem))
    steps_def = get_project_steps(project)
    sn = next((s['name_en'] for s in steps_def if s['step_number']==step), f"Step {step}")
    act = "completed" if comp else "unchecked"
    db_execute("INSERT INTO activity_log (project,spool_id,step_number,action,operator,details) VALUES (?,?,?,?,?,?)",
        (project, spool_id, step, act, op, f"{sn}: {act} by {op}"))
    db_commit()
    return jsonify({'ok':True, 'progress_pct': spool_progress(project, spool_id)})

# ── API: QC Reports ──────────────────────────────────────────────────────────
@app.route('/project/<project>/spool/<spool_id>/qc/<report_type>')
@login_required
def qc_report_page(project, spool_id, report_type):
    subtype = request.args.get('sub', '')
    return render_template_string(QC_REPORT_HTML, project=project, spool_id=spool_id,
                                  report_type=report_type, report_subtype=subtype)

@app.route('/api/project/<project>/spool/<spool_id>/qc')
@login_required
def api_qc_list(project, spool_id):
    """List all QC reports for a spool with status, filtered by spool_type."""
    # Get spool type (SPOOL or STRAIGHT)
    sp = db_fetchone("SELECT spool_type FROM spools WHERE project=? AND spool_id=?", (project, spool_id))
    spool_type = sp.get('spool_type', 'SPOOL') if sp else 'SPOOL'
    report_defs = get_qc_reports_for_spool(project, spool_type)
    # Get existing report statuses
    existing = {}
    rows = db_fetchall("SELECT report_type, report_subtype, status, inspector_name, tpi_name, updated_at FROM qc_reports WHERE project=? AND spool_id=?", (project, spool_id))
    for r in rows:
        key = f"{r['report_type']}|{r.get('report_subtype','')}"
        existing[key] = {'status': r['status'], 'inspector': r.get('inspector_name',''), 'tpi': r.get('tpi_name',''), 'updated': r.get('updated_at','')}
    # Get step completion dates from progress table (report date = step completion date)
    step_dates = {}
    step_rows = db_fetchall("SELECT step_number, completed_at FROM progress WHERE project=? AND spool_id=? AND completed=1", (project, spool_id))
    for sr in step_rows:
        step_dates[sr['step_number']] = str(sr.get('completed_at','') or '')[:10]  # YYYY-MM-DD
    result = []
    proj_info = get_qc_project_info(project)
    for d in report_defs:
        subtype = d.get('report_subtype', '')
        key = f"{d['type']}|{subtype}"
        ex = existing.get(key, {})
        rec_no = get_record_number(project, spool_id, d['rec_seq'])
        result.append({
            'type': d['type'], 'subtype': subtype, 'sub_label': d.get('sub_label',''),
            'rec_seq': d['rec_seq'], 'rec_no': rec_no,
            'itp_step': d.get('itp_step',0),
            'name_en': d['name_en'], 'name_cn': d['name_cn'],
            'category': d['category'], 'icon': d.get('icon',''),
            'standard': d.get('standard',''),
            'insp_code': d.get('insp_code',''),
            'is_hold': d.get('is_hold', False),
            'has_images': d.get('has_images', False),
            'per_weld': d.get('per_weld', False),
            'status': ex.get('status', 'not_started'),
            'inspector': ex.get('inspector', ''),
            'tpi': ex.get('tpi', ''),
            'updated': ex.get('updated', ''),
            'step_date': step_dates.get(d.get('itp_step',0), ''),
            'contract': proj_info.get('contract',''),
            'client': proj_info.get('client',''),
            'material': proj_info.get('material',''),
            'itp': proj_info.get('itp',''),
        })
    return jsonify(result)

@app.route('/api/project/<project>/spool/<spool_id>/qc/<report_type>', methods=['GET'])
@login_required
def api_qc_get(project, spool_id, report_type):
    subtype = request.args.get('sub', '')
    # Find the ITP step number for this report type to get step completion date
    defs = get_qc_report_defs(project)
    itp_step = 0
    rec_no = ''
    for d in defs:
        if d['type'] == report_type and d.get('report_subtype','') == subtype:
            itp_step = d.get('itp_step', 0)
            rec_no = get_record_number(project, spool_id, d['rec_seq'])
            break
    step_date = ''
    if itp_step:
        sr = db_fetchone("SELECT completed_at FROM progress WHERE project=? AND spool_id=? AND step_number=? AND completed=1", (project, spool_id, itp_step))
        if sr and sr.get('completed_at'):
            step_date = str(sr['completed_at'] or '')[:10]
    row = db_fetchone("SELECT * FROM qc_reports WHERE project=? AND spool_id=? AND report_type=? AND report_subtype=?",
                      (project, spool_id, report_type, subtype))
    if row:
        r = dict(row)
        r['data'] = json.loads(r.get('data','{}'))
        r['step_date'] = step_date
        r['rec_no'] = rec_no
        proj_info = get_qc_project_info(project)
        r['itp'] = proj_info.get('itp', '')
        r['material_type'] = proj_info.get('material_type', '')
        img_count = db_fetchone("SELECT COUNT(*) as cnt FROM qc_images WHERE project=? AND spool_id=? AND report_type=?",
                                (project, spool_id, report_type))
        r['image_count'] = img_count['cnt'] if img_count else 0
        return jsonify(r)
    # Return empty template
    return jsonify({
        'project': project, 'spool_id': spool_id, 'report_type': report_type,
        'report_subtype': subtype, 'status': 'not_started',
        'inspector_name': '', 'inspector_date': '', 'tpi_name': '', 'tpi_date': '',
        'data': {}, 'image_count': 0, 'step_date': step_date, 'rec_no': rec_no,
        'itp': get_qc_project_info(project).get('itp', ''),
        'material_type': get_qc_project_info(project).get('material_type', ''),
    })

@app.route('/api/project/<project>/spool/<spool_id>/qc/<report_type>', methods=['POST'])
@login_required
def api_qc_save(project, spool_id, report_type):
    d = request.get_json() or {}
    subtype = d.get('report_subtype', '')
    status = d.get('status', 'draft')
    inspector = d.get('inspector_name', '')
    inspector_date = d.get('inspector_date', '')
    tpi = d.get('tpi_name', '')
    tpi_date = d.get('tpi_date', '')
    data_json = json.dumps(d.get('data', {}))
    created_by = d.get('created_by', inspector)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if USE_PG:
        db_execute("""INSERT INTO qc_reports (project,spool_id,report_type,report_subtype,status,inspector_name,inspector_date,tpi_name,tpi_date,data,created_by,updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (project,spool_id,report_type,report_subtype) DO UPDATE SET
            status=EXCLUDED.status, inspector_name=EXCLUDED.inspector_name, inspector_date=EXCLUDED.inspector_date,
            tpi_name=EXCLUDED.tpi_name, tpi_date=EXCLUDED.tpi_date, data=EXCLUDED.data, updated_at=EXCLUDED.updated_at""",
            (project, spool_id, report_type, subtype, status, inspector, inspector_date, tpi, tpi_date, data_json, created_by, now))
    else:
        db_execute("""INSERT INTO qc_reports (project,spool_id,report_type,report_subtype,status,inspector_name,inspector_date,tpi_name,tpi_date,data,created_by,updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(project,spool_id,report_type,report_subtype) DO UPDATE SET
            status=excluded.status, inspector_name=excluded.inspector_name, inspector_date=excluded.inspector_date,
            tpi_name=excluded.tpi_name, tpi_date=excluded.tpi_date, data=excluded.data, updated_at=excluded.updated_at""",
            (project, spool_id, report_type, subtype, status, inspector, inspector_date, tpi, tpi_date, data_json, created_by, now))
    db_commit()
    return jsonify({'ok': True})

@app.route('/api/project/<project>/spool/<spool_id>/qc/<report_type>/image', methods=['POST'])
@login_required
def api_qc_image_upload(project, spool_id, report_type):
    if 'file' in request.files:
        f = request.files['file']
        img_data = f.read()
        fname = f.filename or 'image.jpg'
        mime = f.content_type or 'image/jpeg'
    else:
        img_data = request.get_data()
        fname = request.headers.get('X-Filename', 'image.jpg')
        mime = request.content_type or 'image/jpeg'
    if not img_data:
        return jsonify({'error': 'No image data'}), 400
    caption = request.form.get('caption', '') if 'file' in request.files else request.headers.get('X-Caption', '')
    operator = request.form.get('operator', '') if 'file' in request.files else request.headers.get('X-Operator', '')
    db_execute("INSERT INTO qc_images (project,spool_id,report_type,image_data,filename,mime_type,caption,uploaded_by) VALUES (?,?,?,?,?,?,?,?)",
        (project, spool_id, report_type, img_data, fname, mime, caption, operator))
    db_commit()
    return jsonify({'ok': True, 'size_kb': round(len(img_data)/1024, 1)})

@app.route('/api/project/<project>/spool/<spool_id>/qc/<report_type>/images')
@login_required
def api_qc_images_list(project, spool_id, report_type):
    rows = db_fetchall("SELECT id, filename, mime_type, caption, uploaded_by, uploaded_at FROM qc_images WHERE project=? AND spool_id=? AND report_type=? ORDER BY id",
        (project, spool_id, report_type))
    return jsonify([dict(r) for r in rows])

@app.route('/api/project/<project>/qc/image/<int:image_id>')
@login_required
def api_qc_image_serve(project, image_id):
    row = db_fetchone("SELECT image_data, mime_type, filename FROM qc_images WHERE id=? AND project=?", (image_id, project))
    if not row:
        return 'Not found', 404
    from io import BytesIO
    return send_file(BytesIO(row['image_data']), mimetype=row.get('mime_type','image/jpeg'),
                     download_name=row.get('filename','image.jpg'))

@app.route('/api/project/<project>/qc/image/<int:image_id>', methods=['DELETE'])
@login_required
def api_qc_image_delete(project, image_id):
    db_execute("DELETE FROM qc_images WHERE id=? AND project=?", (image_id, project))
    db_commit()
    return jsonify({'ok': True})

# ── API: QC Seed from DXF SQLite ─────────────────────────────────────────────
@app.route('/api/project/<project>/spool/<spool_id>/qc/seed', methods=['POST'])
@login_required
def api_qc_seed(project, spool_id):
    """Read DXF SQLite and pre-populate QC report data for this spool."""
    import sqlite3 as sqlite3_mod
    # Accept seed data via POST body (for remote seeding from local machine)
    post_data = request.get_json(silent=True)
    if post_data:
        seed = post_data
        # Pre-populate report-specific data
        for d in get_qc_report_defs(project):
            rt = d['type']
            subtype = d.get('report_subtype', '')
            existing = db_fetchone("SELECT data FROM qc_reports WHERE project=? AND spool_id=? AND report_type=? AND report_subtype=?",
                                   (project, spool_id, rt, subtype))
            if existing and existing.get('data','{}') != '{}':
                continue
            data = seed.get(rt, {})
            if data:
                data_json = json.dumps(data)
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                if USE_PG:
                    db_execute("INSERT INTO qc_reports (project,spool_id,report_type,report_subtype,status,data,updated_at) VALUES (%s,%s,%s,%s,'not_started',%s,%s) ON CONFLICT (project,spool_id,report_type,report_subtype) DO UPDATE SET data=EXCLUDED.data, updated_at=EXCLUDED.updated_at WHERE qc_reports.data = '{}'",
                        (project, spool_id, rt, subtype, data_json, now))
                else:
                    db_execute("INSERT INTO qc_reports (project,spool_id,report_type,report_subtype,status,data,updated_at) VALUES (?,?,?,?,'not_started',?,?) ON CONFLICT(project,spool_id,report_type,report_subtype) DO UPDATE SET data=excluded.data, updated_at=excluded.updated_at WHERE qc_reports.data = '{}'",
                        (project, spool_id, rt, subtype, data_json, now))
        db_commit()
        return jsonify({'ok': True, 'source': 'post_body'})
    # Auto-seed for straight pipes — derive data from spool name (no DXF needed)
    sp_row = db_fetchone("SELECT * FROM spools WHERE project=? AND spool_id=?", (project, spool_id))
    if sp_row and (sp_row.get('spool_type') or '').upper() == 'STRAIGHT':
        import re as re_mod
        # Extract length from spool name (e.g. "STRAIGHT PIPE L=9697mm" → 9697)
        name = sp_row.get('spool_id','') + ' ' + (sp_row.get('marking') or '')
        m = re_mod.search(r'L=(\d+)', name)
        length_mm = int(m.group(1)) if m else None
        diameter = sp_row.get('main_diameter','')
        drawing_ref = f"{project}-{spool_id}"
        seed = {
            'cutting': {'drawing_ref': drawing_ref, 'pieces': [{'mark': 'Cut 1', 'description': f'STRAIGHT PIPE {diameter}', 'nominal': length_mm, 'size': diameter}] if length_mm else []},
            'dimensional': {'nominal_l_mm': length_mm, 'drawing_ref': drawing_ref, 'flanges': [], 'fittings': []},
        }
        for d in get_qc_report_defs(project):
            rt = d['type']
            if rt not in seed: continue
            subtype = d.get('report_subtype', '')
            existing = db_fetchone("SELECT data FROM qc_reports WHERE project=? AND spool_id=? AND report_type=? AND report_subtype=?",
                                   (project, spool_id, rt, subtype))
            if existing and existing.get('data','{}') != '{}':
                continue
            data_json = json.dumps(seed[rt])
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if USE_PG:
                db_execute("INSERT INTO qc_reports (project,spool_id,report_type,report_subtype,status,data,updated_at) VALUES (%s,%s,%s,%s,'not_started',%s,%s) ON CONFLICT (project,spool_id,report_type,report_subtype) DO UPDATE SET data=EXCLUDED.data, updated_at=EXCLUDED.updated_at WHERE qc_reports.data = '{}'",
                    (project, spool_id, rt, subtype, data_json, now))
            else:
                db_execute("INSERT INTO qc_reports (project,spool_id,report_type,report_subtype,status,data,updated_at) VALUES (?,?,?,?,'not_started',?,?) ON CONFLICT(project,spool_id,report_type,report_subtype) DO UPDATE SET data=excluded.data, updated_at=excluded.updated_at WHERE qc_reports.data = '{}'",
                    (project, spool_id, rt, subtype, data_json, now))
        db_commit()
        return jsonify({'ok': True, 'source': 'straight_pipe', 'length_mm': length_mm})

    # Fall back to DXF file (local only)
    db_path = get_qc_setting(project, 'dxf_sqlite_path')
    if not db_path or not os.path.exists(db_path):
        return jsonify({'ok': True, 'source': 'no_dxf', 'seed': {}})
    conn = sqlite3_mod.connect(db_path)
    conn.row_factory = sqlite3_mod.Row
    # Find spool
    sp = conn.execute("SELECT * FROM spools WHERE spool_number=?", (spool_id,)).fetchone()
    if not sp:
        conn.close()
        return jsonify({'error': f'Spool {spool_id} not found in DXF database'}), 404
    sid = sp['spool_id']
    attrs = json.loads(sp['attrs'] or '{}')
    # Get welds
    welds_raw = conn.execute("SELECT weld_tag, weld_type, size FROM welds WHERE spool_id=? ORDER BY weld_tag", (sid,)).fetchall()
    welds = [{'weld_tag': w['weld_tag'], 'weld_type': w['weld_type'] or 'butt', 'size': w['size'] or ''} for w in welds_raw]
    # Get parts (pipes, flanges, fittings)
    parts_raw = [dict(r) for r in conn.execute("SELECT * FROM parts WHERE spool_id=? ORDER BY item_number", (sid,)).fetchall()]
    pipes = [{'mark': f"Cut {p['item_number']}", 'description': p['description'], 'nominal': p['cut_length_mm'], 'size': p['nominal_size']}
             for p in parts_raw if p['part_type'] == 'pipe' and p.get('cut_length_mm') and p['cut_length_mm'] > 0]
    flanges = [{'size': p['nominal_size'], 'description': p['description'], 'rating': p.get('rating','')}
               for p in parts_raw if p['part_type'] == 'flange']
    fittings = [{'size': p['nominal_size'], 'description': p['description'], 'part_type': p['part_type']}
                for p in parts_raw if p['part_type'] not in ('pipe', 'flange', 'repad')]
    # Envelope dimensions
    envelope = attrs.get('envelope', {})
    conn.close()
    # Build drawing reference
    spool_data = db_fetchone("SELECT marking, iso_no FROM spools WHERE project=? AND spool_id=?", (project, spool_id))
    drawing_ref = f"{project}-{spool_id}"
    marking = spool_data.get('marking', '') if spool_data else ''
    # Build seed data per report type
    seed = {
        'welds': welds,
        'pipes': pipes,
        'flanges': flanges,
        'fittings': fittings,
        'envelope': envelope,
        'drawing_ref': drawing_ref,
        'marking': marking,
        'wps_options': list((get_wps_registry(project)).keys()),
    }
    # Pre-populate report-specific data in qc_reports
    for d in get_qc_report_defs(project):
        rt = d['type']
        subtype = d.get('report_subtype', '')
        existing = db_fetchone("SELECT data FROM qc_reports WHERE project=? AND spool_id=? AND report_type=? AND report_subtype=?",
                               (project, spool_id, rt, subtype))
        if existing and existing.get('data','{}') != '{}':
            continue  # Don't overwrite existing data
        data = {}
        if rt == 'cutting':
            data = {'drawing_ref': drawing_ref, 'pieces': pipes}
        elif rt in ('fitup', 'vt', 'pt', 'mt', 'rt', 'welding_log'):
            data = {'welds': [dict(w) for w in welds]}
        elif rt == 'ferrite':
            data = {'readings': [{'weld_tag': w['weld_tag'], 'size': w['size']} for w in welds]}
        elif rt == 'pmi':
            all_items = [{'item': f"{p['description'][:30]} {p['nominal_size']}", 'location': 'Base metal'} for p in parts_raw if p['part_type'] in ('pipe','flange','elbow','tee')]
            weld_items = [{'item': f"Weld {w['weld_tag']} ({w['size']})", 'location': 'Weld'} for w in welds]
            data = {'items': all_items[:5] + weld_items[:5]}  # Sample of items
        elif rt == 'dimensional':
            data = {
                'nominal_l_mm': envelope.get('length_mm'),
                'nominal_w_mm': envelope.get('width_mm'),
                'nominal_h_mm': envelope.get('height_mm'),
                'flanges': [{'size': f['size'], 'description': f['description']} for f in flanges],
                'fittings': fittings,
                'drawing_ref': drawing_ref,
            }
        elif rt == 'ferroxyl':
            areas = [{'location': f"Weld {w['weld_tag']} + HAZ / 焊缝{w['weld_tag']}+热影响区", 'copper_deposit': None} for w in welds]
            areas.append({'location': 'Ground/repaired areas / 打磨/修复区域', 'copper_deposit': None})
            data = {'areas': areas}
        if data:
            data_json = json.dumps(data)
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            if USE_PG:
                db_execute("""INSERT INTO qc_reports (project,spool_id,report_type,report_subtype,status,data,updated_at)
                    VALUES (%s,%s,%s,%s,'not_started',%s,%s)
                    ON CONFLICT (project,spool_id,report_type,report_subtype) DO UPDATE SET data=EXCLUDED.data, updated_at=EXCLUDED.updated_at
                    WHERE qc_reports.data = '{}'""",
                    (project, spool_id, rt, subtype, data_json, now))
            else:
                db_execute("""INSERT INTO qc_reports (project,spool_id,report_type,report_subtype,status,data,updated_at)
                    VALUES (?,?,?,?,'not_started',?,?)
                    ON CONFLICT(project,spool_id,report_type,report_subtype) DO UPDATE SET data=excluded.data, updated_at=excluded.updated_at
                    WHERE qc_reports.data = '{}'""",
                    (project, spool_id, rt, subtype, data_json, now))
    db_commit()
    return jsonify({'ok': True, 'seed': seed})

# ── API: Shipments ────────────────────────────────────────────────────────────
@app.route('/api/project/<project>/shipments')
@login_required
def api_shipments_list(project):
    """List all shipments for a project with calculated ETA."""
    rows = db_fetchall("SELECT * FROM shipments WHERE project=? ORDER BY shipment_number", (project,))
    result = []
    for r in rows:
        d = dict(r)
        etd = d.get('etd', '')
        transit = int(d.get('transit_days', 0) or 0)
        eta = ''
        if etd and transit:
            try:
                from datetime import timedelta
                etd_dt = parse_date(etd) if callable(parse_date) else None
                if etd_dt:
                    eta = (etd_dt + timedelta(days=transit)).strftime('%Y-%m-%d')
            except: pass
        d['eta'] = eta
        result.append(d)
    return jsonify(result)

@app.route('/api/project/<project>/shipments', methods=['POST'])
@login_required
def api_shipments_save(project):
    """Add or update a shipment. Only updates fields present in the request."""
    d = request.get_json() or {}
    num = d.get('shipment_number')
    if not num: return jsonify({'error': 'shipment_number required'}), 400
    # Check if shipment exists
    existing = db_fetchone("SELECT * FROM shipments WHERE project=? AND shipment_number=?", (project, num))
    if existing:
        # Update only fields that are in the request
        updates = []
        params = []
        for col in ['description', 'etd', 'transit_days', 'notes']:
            if col in d:
                updates.append(f"{col}=?")
                params.append(d[col] if d[col] != '' or col not in ('etd',) else None)
        if updates:
            params.extend([project, num])
            q = f"UPDATE shipments SET {','.join(updates)} WHERE project=? AND shipment_number=?"
            db_execute(q, tuple(params))
    else:
        # Insert new with defaults
        desc = d.get('description', '')
        etd = d.get('etd', '') or None
        transit = d.get('transit_days', 45)
        notes = d.get('notes', '')
        db_execute("INSERT INTO shipments (project,shipment_number,description,etd,transit_days,notes) VALUES (?,?,?,?,?,?)",
            (project, num, desc, etd, transit, notes))
    db_commit()
    return jsonify({'ok': True})

@app.route('/api/project/<project>/shipments/<int:num>', methods=['DELETE'])
@login_required
def api_shipments_delete(project, num):
    db_execute("DELETE FROM shipments WHERE project=? AND shipment_number=?", (project, num))
    # Clear any spools assigned to the deleted shipment
    db_execute("UPDATE spools SET shipment_number=NULL WHERE project=? AND shipment_number=?", (project, num))
    db_commit()
    return jsonify({'ok': True})

@app.route('/api/project/<project>/spool/<spool_id>/shipment', methods=['POST'])
@login_required
def api_spool_set_shipment(project, spool_id):
    """Assign a single spool to a shipment, or unassign when shipment_number is null."""
    d = request.get_json() or {}
    num = d.get('shipment_number')
    if num == '' or num == 0: num = None
    db_execute("UPDATE spools SET shipment_number=? WHERE project=? AND spool_id=?",
               (num, project, spool_id))
    db_commit()
    return jsonify({'ok': True})

@app.route('/api/project/<project>/spools/assign-shipment', methods=['POST'])
@login_required
def api_spools_bulk_assign_shipment(project):
    """Bulk-assign multiple spools to a shipment (or unassign with null)."""
    d = request.get_json() or {}
    spool_ids = d.get('spool_ids') or []
    num = d.get('shipment_number')
    if num == '' or num == 0: num = None
    if not spool_ids:
        return jsonify({'error': 'spool_ids required'}), 400
    ph = ','.join(['?'] * len(spool_ids))
    db_execute(
        f"UPDATE spools SET shipment_number=? WHERE project=? AND spool_id IN ({ph})",
        tuple([num, project] + list(spool_ids)))
    db_commit()
    return jsonify({'ok': True, 'updated': len(spool_ids)})

@app.route('/api/project/<project>/shipment/<int:num>/mark-shipped', methods=['POST'])
@login_required
def api_shipment_mark_shipped(project, num):
    """Bulk-tick the 'shipped' gate (last counts_for_production=0 step by display_order)
    for every spool assigned to the given shipment. Generalised — no hardcoded step.
    Body: {date: 'YYYY-MM-DD', operator: '...'} (both optional; defaults to today)."""
    d = request.get_json() or {}
    op = d.get('operator', 'bulk-ship')
    ship_date = d.get('date')
    if ship_date:
        ts = f"{ship_date} 12:00:00"
    else:
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    steps_def = get_project_steps(project)
    pp_steps = get_post_production_steps(steps_def)
    if not pp_steps:
        return jsonify({'error': 'No post-production step defined for this project'}), 400
    shipped_step = pp_steps[-1]['step_number']
    spools = db_fetchall("SELECT spool_id FROM spools WHERE project=? AND shipment_number=?",
                         (project, num))
    updated = 0
    for s in spools:
        db_execute(
            "INSERT INTO progress (project,spool_id,step_number,completed,completed_by,completed_at) "
            "VALUES (?,?,?,1,?,?) "
            "ON CONFLICT (project,spool_id,step_number) DO UPDATE SET "
            "completed=1, completed_by=EXCLUDED.completed_by, completed_at=EXCLUDED.completed_at",
            (project, s['spool_id'], shipped_step, op, ts))
        db_execute(
            "INSERT INTO activity_log (project,spool_id,step_number,action,operator,details) "
            "VALUES (?,?,?,?,?,?)",
            (project, s['spool_id'], shipped_step, 'completed', op,
             f"Shipped via shipment {num} on {ts[:10]}"))
        updated += 1
    db_commit()
    return jsonify({'ok': True, 'updated': updated, 'shipped_step': shipped_step, 'date': ts[:10]})

# ── API: Inspector Registry ───────────────────────────────────────────────────
@app.route('/api/inspectors')
@login_required
def api_inspectors_list():
    """List all saved inspectors with their roles (signature_data excluded for speed)."""
    rows = db_fetchall("SELECT id, name, role, CASE WHEN signature_data != '' THEN 1 ELSE 0 END as has_signature FROM qc_inspectors ORDER BY name")
    return jsonify([dict(r) for r in rows])

@app.route('/api/inspectors', methods=['POST'])
@login_required
def api_inspectors_save():
    """Add or update an inspector. If name exists, update role/signature."""
    d = request.get_json() or {}
    name = (d.get('name','') or '').strip()
    if not name: return jsonify({'error':'Name required'}), 400
    role = d.get('role', '')
    sig = d.get('signature_data', '')
    if USE_PG:
        db_execute("INSERT INTO qc_inspectors (name,role,signature_data) VALUES (%s,%s,%s) ON CONFLICT (name) DO UPDATE SET role=EXCLUDED.role, signature_data=CASE WHEN EXCLUDED.signature_data != '' THEN EXCLUDED.signature_data ELSE qc_inspectors.signature_data END",
            (name, role, sig))
    else:
        db_execute("INSERT INTO qc_inspectors (name,role,signature_data) VALUES (?,?,?) ON CONFLICT(name) DO UPDATE SET role=excluded.role, signature_data=CASE WHEN excluded.signature_data != '' THEN excluded.signature_data ELSE qc_inspectors.signature_data END",
            (name, role, sig))
    db_commit()
    return jsonify({'ok': True})

@app.route('/api/inspectors/<int:inspector_id>/signature')
@login_required
def api_inspector_signature(inspector_id):
    """Get an inspector's signature as base64 data URI."""
    row = db_fetchone("SELECT signature_data FROM qc_inspectors WHERE id=?", (inspector_id,))
    if not row or not row.get('signature_data'):
        return jsonify({'signature_data': ''})
    return jsonify({'signature_data': row['signature_data']})

@app.route('/api/inspectors/<int:inspector_id>', methods=['DELETE'])
@login_required
def api_inspector_delete(inspector_id):
    db_execute("DELETE FROM qc_inspectors WHERE id=?", (inspector_id,))
    db_commit()
    return jsonify({'ok': True})

# ── API: WPS Registry ────────────────────────────────────────────────────────
@app.route('/api/project/<project>/wps')
@login_required
def api_wps_list(project):
    """Return WPS options for this project."""
    return jsonify(get_wps_registry(project))

# ── QC PDF Generator (reportlab) — EN 10204 Type 3.1 Certificate ─────────────
def build_qc_pdf(project, spool_id, report_def, report_row, proj_info, step_date):
    """Generate a single QC report PDF. Returns bytes. Generalised — works for any report type."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor, black, white
    from reportlab.pdfgen import canvas as pdfcanvas
    from io import BytesIO
    buf = BytesIO()
    W, H = A4
    c = pdfcanvas.Canvas(buf, pagesize=A4)
    margin = 20*mm
    cw = W - 2*margin  # content width
    y = H - 15*mm
    blue = HexColor('#2F5496')
    grey = HexColor('#666666')
    light = HexColor('#f0f2f5')
    green = HexColor('#27ae60')
    red = HexColor('#e74c3c')
    data = json.loads(report_row['data']) if report_row and report_row.get('data') else {}
    rec_no = get_record_number(project, spool_id, report_def['rec_seq'])
    sub_label = f" ({report_def.get('sub_label','')})" if report_def.get('sub_label') else ''
    inspector = report_row.get('inspector_name', '') if report_row else ''
    overall = data.get('overall_result', '')

    def draw_text(x, yy, text, size=9, color=black, bold=False, align='left', max_w=None):
        font = 'Helvetica-Bold' if bold else 'Helvetica'
        c.setFont(font, size); c.setFillColor(color)
        if align == 'center' and max_w:
            tw = c.stringWidth(text, font, size)
            x = x + (max_w - tw) / 2
        elif align == 'right' and max_w:
            tw = c.stringWidth(text, font, size)
            x = x + max_w - tw
        c.drawString(x, yy, str(text)[:120])

    def draw_line(yy, color=black, width=0.5):
        c.setStrokeColor(color); c.setLineWidth(width)
        c.line(margin, yy, W-margin, yy)

    def draw_rect(x, yy, w, h, fill=None, stroke=None):
        if fill: c.setFillColor(fill)
        if stroke: c.setStrokeColor(stroke)
        c.rect(x, yy, w, h, fill=bool(fill), stroke=bool(stroke))

    def new_page_if_needed(yy, need=40*mm):
        if yy < margin + need:
            draw_footer()
            c.showPage()
            return H - 15*mm
        return yy

    def draw_footer():
        fy = margin - 5*mm
        draw_line(fy + 8*mm, grey, 0.3)
        draw_text(margin, fy + 2*mm, 'ENERXON (Cangzhou) Pipeline Co., Ltd. · ISO 9001:2015 · IQNet', 7, grey)
        draw_text(margin, fy - 2*mm, 'EN 10204 Type 3.1 Inspection Certificate', 7, grey)
        draw_text(W-margin, fy, '3.1', 8, grey, align='right', max_w=0)

    # ── LOGO + TITLE ──
    logo_path = os.path.join(os.path.expanduser('~'), 'Library', 'CloudStorage', 'Dropbox', 'OB', 'Logos', 'Enerxon Logo Especifications 1.png')
    if os.path.exists(logo_path):
        try: c.drawImage(logo_path, margin, y - 14*mm, width=35*mm, height=14*mm, preserveAspectRatio=True, mask='auto')
        except: pass
    draw_text(margin + 40*mm, y - 5*mm, report_def['name_en'] + ' Report' + sub_label, 13, blue, bold=True, align='center', max_w=cw - 80*mm)
    draw_text(W - margin, y - 3*mm, 'Page 1 of 1', 8, grey, align='right', max_w=0)
    y -= 18*mm
    draw_line(y, black, 1.5)
    y -= 3*mm

    # ── INFO GRID ──
    info_fields = [
        ('Report No.', rec_no), ('Date', step_date),
        ('Project', project), ('Contract', proj_info.get('contract', '')),
        ('Client', proj_info.get('client', '')), ('ITP', proj_info.get('itp', '')),
        ('Spool No.', spool_id), ('Material', proj_info.get('material', '')),
    ]
    col_w = cw / 2
    for i, (label, val) in enumerate(info_fields):
        col = i % 2
        if col == 0 and i > 0: y -= 5*mm
        x = margin + col * col_w
        draw_text(x, y, label + ':', 8, grey, bold=True)
        draw_text(x + 25*mm, y, str(val), 9, black)
    y -= 8*mm
    draw_line(y, grey, 0.3)
    y -= 3*mm

    # ── STANDARD / CRITERIA BOX ──
    standard = report_def.get('standard', '')
    if standard:
        draw_rect(margin, y - 8*mm, cw, 8*mm, fill=light)
        draw_text(margin + 3*mm, y - 5*mm, f"Standard: {standard}", 9, grey)
        y -= 12*mm

    # ── DATA SECTION — generalised table renderer ──
    def draw_table(headers, rows, col_widths=None, yy=None):
        if yy is None: yy = y
        n = len(headers)
        if not col_widths:
            col_widths = [cw / n] * n
        # Header
        yy = new_page_if_needed(yy, 15*mm)
        hx = margin
        draw_rect(margin, yy - 5*mm, cw, 6*mm, fill=light)
        for j, h in enumerate(headers):
            draw_text(hx + 2*mm, yy - 3.5*mm, h, 7, black, bold=True)
            hx += col_widths[j]
        yy -= 7*mm
        # Rows
        for row in rows:
            yy = new_page_if_needed(yy, 8*mm)
            hx = margin
            for j, cell in enumerate(row):
                color = green if cell == 'ACC' else (red if cell == 'REJ' else black)
                bold = cell in ('ACC', 'REJ')
                draw_text(hx + 2*mm, yy - 3*mm, str(cell), 8, color, bold=bold)
                hx += col_widths[j]
            draw_line(yy - 5*mm, HexColor('#eeeeee'), 0.3)
            yy -= 6*mm
        return yy

    def section_title(title, yy):
        yy = new_page_if_needed(yy, 15*mm)
        draw_text(margin, yy, title, 10, blue, bold=True)
        yy -= 2*mm
        draw_line(yy, blue, 0.5)
        yy -= 4*mm
        return yy

    # Render data section based on report type
    rt = report_def['type']
    if rt == 'cutting':
        y = section_title('Pipe Cut Inspection', y)
        pieces = data.get('pieces', [])
        rows = [[p.get('mark',''), p.get('size',''), str(int(p['nominal'])) if p.get('nominal') else '', str(int(p['actual'])) if p.get('actual') else '', 'ACC' if p.get('pass') else ('REJ' if p.get('pass') is False else '—')] for p in pieces]
        y = draw_table(['Piece', 'Size', 'Nominal (mm)', 'Actual (mm)', 'Result'], rows, yy=y)

    elif rt == 'fitup':
        y = section_title('Joint Geometric Check', y)
        welds = data.get('welds', [])
        rows = [[w.get('weld_tag',''), w.get('size',''), w.get('weld_type',''), 'ACC' if w.get('geometric_ok') else ('REJ' if w.get('geometric_ok') is False else '—')] for w in welds]
        y = draw_table(['Joint', 'Size', 'Type', 'Result'], rows, yy=y)

    elif rt == 'welding_log':
        y = section_title('WPS Applied', y)
        wps_nums = data.get('wps_numbers', [])
        wps_reg = get_wps_registry(project)
        for wn in wps_nums:
            wps = wps_reg.get(wn, {})
            if wps:
                draw_text(margin, y, f"{wn} — {wps.get('label','')}", 9, blue, bold=True); y -= 5*mm
                for k in ['processes','filler','shielding','backing','preheat','interpass','heat_input']:
                    if wps.get(k):
                        draw_text(margin + 5*mm, y, f"{k.title()}: {wps[k]}", 8, grey); y -= 4*mm
                y -= 2*mm
        if data.get('wps_followed'):
            draw_text(margin, y, 'WPS Followed: YES', 9, green, bold=True); y -= 6*mm
        y = section_title('Weld Data', y)
        welds = data.get('welds', [])
        rows = [[w.get('weld_tag',''), w.get('size',''), w.get('welder_id',''), w.get('date',''), 'ACC' if w.get('done') else '—'] for w in welds]
        y = draw_table(['Weld', 'Size', 'Welder ID', 'Date', 'Done'], rows, yy=y)

    elif rt == 'vt':
        y = section_title('VT Checklist', y)
        vt_items = ['Identification correct','Surface condition suitable','Workmanship acceptable','Dimensions/alignment correct','Weld profile acceptable','No visible cracks','No incomplete fusion/penetration','No undercut/overlap/arc strikes','No surface defects','Attachments/end prep acceptable','ASME B31.3 Table 341.3.2 verified','VT per ASME V Art. 9 recorded']
        for i, item in enumerate(vt_items):
            y = new_page_if_needed(y, 7*mm)
            key = ['id_correct','surface_suitable','workmanship','dimensions_align','weld_profile','no_cracks','no_fusion_defect','no_surface_defect','no_damage','attachments_ok','acceptance_verified','vt_recorded'][i]
            val = data.get(key)
            result = 'ACC' if val is True else ('REJ' if val is False else '—')
            color = green if val is True else (red if val is False else grey)
            draw_text(margin, y, f"{i+1}. {item}", 8, black)
            draw_text(W - margin - 15*mm, y, result, 9, color, bold=True)
            y -= 5*mm

    elif rt == 'rt':
        y = section_title('RT Equipment', y)
        for k, label in [('instrument','Instrument'),('source_type','Source'),('focus_size','Focus'),('sfd','SFD'),('iqi_type','IQI'),('tube_voltage','Voltage (kV)'),('tube_current','Current (mA)'),('density_range','Density Range'),('technique','Technique')]:
            if data.get(k):
                draw_text(margin, y, f"{label}: {data[k]}", 8, grey); y -= 4*mm
        y -= 2*mm
        y = section_title('Film Results (3 films per weld)', y)
        welds = data.get('welds', [])
        headers = ['Weld','Film','Bar','Circ.','Crack','LOP','LOF','Density','Result']
        for w in welds:
            for fi, f in enumerate(w.get('films', [])):
                y = new_page_if_needed(y, 6*mm)
                tag = w.get('weld_tag','') if fi == 0 else ''
                defects = ['X' if f.get(k) else '—' for k in ['bar','circular','crack','lop','lof']]
                result = f.get('result', '—')
                row_data = [tag, str(fi+1)] + defects + [str(f.get('density','')), result]
                hx = margin
                cws = [18*mm, 12*mm, 12*mm, 12*mm, 12*mm, 12*mm, 12*mm, 20*mm, 15*mm]
                for j, cell in enumerate(row_data):
                    color = green if cell == 'ACC' else (red if cell in ('REJ','X') else black)
                    draw_text(hx + 1*mm, y, cell, 7, color, bold=(cell in ('ACC','REJ')))
                    hx += cws[j]
                draw_line(y - 2*mm, HexColor('#eee'), 0.2)
                y -= 4*mm

    elif rt in ('pt', 'mt'):
        y = section_title('Weld Examination', y)
        welds = data.get('welds', [])
        rows = [[w.get('weld_tag',''), w.get('size',''), w.get('indication','NRI'), w.get('result','—')] for w in welds]
        y = draw_table(['Weld', 'Size', 'Indication', 'Result'], rows, yy=y)

    elif rt == 'pmi':
        y = section_title('PMI Test Points', y)
        items = data.get('items', [])
        rows = [[it.get('item',''), it.get('location',''), str(it.get('cr','')), str(it.get('ni','')), str(it.get('mo','')), 'S32205' if it.get('grade_ok') else '—', it.get('result','—')] for it in items]
        y = draw_table(['Item', 'Location', 'Cr%', 'Ni%', 'Mo%', 'Grade', 'Result'], rows, yy=y)

    elif rt == 'ferrite':
        y = section_title('Ferrite Readings', y)
        readings = data.get('readings', [])
        rows = []
        for rd in readings:
            r1, r2, r3 = rd.get('r1'), rd.get('r2'), rd.get('r3')
            avg = f"{((r1+r2+r3)/3):.1f}%" if all(v is not None for v in [r1,r2,r3]) else '—'
            avg_val = (r1+r2+r3)/3 if all(v is not None for v in [r1,r2,r3]) else None
            result = 'ACC' if (avg_val and 30 <= avg_val <= 65) else ('REJ' if avg_val else '—')
            rows.append([rd.get('weld_tag',''), str(r1 or ''), str(r2 or ''), str(r3 or ''), avg, result])
        y = draw_table(['Weld', 'R1 (%)', 'R2 (%)', 'R3 (%)', 'Average', 'Result'], rows, yy=y)

    elif rt == 'dimensional':
        y = section_title('Overall Dimensions', y)
        rows = []
        for k, label in [('l','Length'),('w','Width'),('h','Height')]:
            nom = data.get(f'nominal_{k}_mm')
            act = data.get(f'actual_{k}_mm')
            dev = f"{(act-nom):.1f}mm" if nom and act else '—'
            ok = 'ACC' if (nom and act and abs(act-nom) <= 5) else ('REJ' if nom and act else '—')
            rows.append([label, str(int(nom)) if nom else '—', str(int(act)) if act else '—', dev, ok])
        y = draw_table(['Dimension', 'Nominal (mm)', 'Actual (mm)', 'Dev.', 'Result'], rows, yy=y)
        flanges = data.get('flanges', [])
        if flanges:
            y = section_title('Flange Alignment', y)
            rows = [[f.get('description',''), f.get('size',''), 'ACC' if f.get('bolt_hole_ok') else '—', 'ACC' if f.get('perp_ok') else '—'] for f in flanges]
            y = draw_table(['Flange', 'Size', 'Bolt-Hole', 'Perp.'], rows, yy=y)

    elif rt == 'ferroxyl':
        y = section_title('Test Results', y)
        areas = data.get('areas', [])
        rows = [[a.get('location',''), a.get('surface_cond',''), 'No' if a.get('copper_deposit') is False else ('Yes' if a.get('copper_deposit') else '—'), 'ACC' if a.get('copper_deposit') is False else ('REJ' if a.get('copper_deposit') else '—')] for a in areas]
        y = draw_table(['Area', 'Surface', 'Cu Deposit', 'Result'], rows, yy=y)

    elif rt == 'dft':
        y = section_title('DFT Readings', y)
        readings = data.get('readings', [])
        rows = [[r.get('point',''), str(r.get('value','')) + ' um' if r.get('value') else '—', '—'] for r in readings]
        y = draw_table(['Point', 'Reading', 'Result'], rows, yy=y)

    # ── REMARKS ──
    remarks = data.get('remarks', '')
    if remarks:
        y -= 3*mm
        y = new_page_if_needed(y, 15*mm)
        draw_text(margin, y, 'Remarks:', 9, grey, bold=True); y -= 5*mm
        draw_text(margin, y, remarks[:200], 8, black); y -= 6*mm

    # ── OVERALL RESULT ──
    y -= 3*mm
    y = new_page_if_needed(y, 25*mm)
    draw_line(y, black, 1.5); y -= 8*mm
    draw_text(margin, y, 'Inspection Result:', 11, black, bold=True)
    result_text = 'CONFORMING' if overall == 'ACC' else ('NON-CONFORMING' if overall == 'REJ' else 'PENDING')
    result_color = green if overall == 'ACC' else (red if overall == 'REJ' else grey)
    draw_text(W - margin - 50*mm, y, result_text, 12, result_color, bold=True)
    y -= 10*mm

    # ── SIGNATURES ──
    y = new_page_if_needed(y, 30*mm)
    sig_w = cw / 3 - 3*mm
    # Load inspector signature if available
    inspector_sig = None
    if inspector:
        sig_row = db_fetchone("SELECT signature_data FROM qc_inspectors WHERE name=?", (inspector,))
        if sig_row and sig_row.get('signature_data','').startswith('data:image'):
            inspector_sig = sig_row['signature_data']
    for i, label in enumerate(['QC Inspector', 'QM', 'TPI']):
        sx = margin + i * (sig_w + 4*mm)
        draw_rect(sx, y - 22*mm, sig_w, 22*mm, stroke=grey)
        draw_text(sx + 2*mm, y - 4*mm, label, 7, grey)
        draw_line(y - 18*mm, grey, 0.3)
        if i == 0 and inspector:
            # Draw signature image if available
            if inspector_sig:
                try:
                    import base64
                    sig_b64 = inspector_sig.split(',')[1]
                    sig_bytes = base64.b64decode(sig_b64)
                    sig_buf = BytesIO(sig_bytes)
                    from reportlab.lib.utils import ImageReader
                    sig_img = ImageReader(sig_buf)
                    c.drawImage(sig_img, sx + 5*mm, y - 17*mm, width=sig_w - 10*mm, height=10*mm, preserveAspectRatio=True, mask='auto')
                except: pass
            draw_text(sx + 5*mm, y - 15*mm, inspector, 9, blue)
            draw_text(sx + 2*mm, y - 21*mm, step_date, 7, grey)

    # ── FOOTER ──
    draw_footer()
    c.save()
    buf.seek(0)
    return buf.getvalue()

# ── API: QC Export (ZIP of all reports for a spool) ──────────────────────────
@app.route('/api/project/<project>/spool/<spool_id>/qc/export')
@login_required
def api_qc_export(project, spool_id):
    """Export all QC reports as PDFs in a ZIP, organized by spool folder."""
    import zipfile
    from io import BytesIO
    sp = db_fetchone("SELECT spool_type FROM spools WHERE project=? AND spool_id=?", (project, spool_id))
    spool_type = sp.get('spool_type', 'SPOOL') if sp else 'SPOOL'
    report_defs = get_qc_reports_for_spool(project, spool_type)
    proj_info = get_qc_project_info(project)
    buf = BytesIO()
    folder = spool_id.replace('/', '-').replace(' ', '_')
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for d in report_defs:
            rt = d['type']
            subtype = d.get('report_subtype', '')
            row = db_fetchone("SELECT * FROM qc_reports WHERE project=? AND spool_id=? AND report_type=? AND report_subtype=?",
                              (project, spool_id, rt, subtype))
            itp_step = d.get('itp_step', 0)
            sr = db_fetchone("SELECT completed_at FROM progress WHERE project=? AND spool_id=? AND step_number=? AND completed=1",
                             (project, spool_id, itp_step)) if itp_step else None
            step_date = str(sr['completed_at'] or '')[:10] if sr and sr.get('completed_at') else ''
            try:
                pdf_bytes = build_qc_pdf(project, spool_id, d, row, proj_info, step_date)
                sub = f"_{d.get('sub_label','').split('/')[0].strip().replace(' ','_')}" if d.get('sub_label') else ''
                fname = f"{d['rec_seq']}_{d['name_en'].replace(' ','_').replace('(','').replace(')','')}{sub}.pdf"
                zf.writestr(f"{folder}/{fname}", pdf_bytes)
            except Exception as e:
                print(f"PDF error {d['type']}: {e}")
                zf.writestr(f"{folder}/{d['rec_seq']}_{d['type']}_ERROR.txt", str(e))
    buf.seek(0)
    return send_file(buf, mimetype='application/zip',
                     download_name=f"{project}_{folder}_QC_Reports.zip", as_attachment=True)

# ── API: Import & Export ──────────────────────────────────────────────────────
@app.route('/api/import', methods=['POST'])
def api_import():
    data = request.get_json()
    if not data or not isinstance(data, list): return jsonify({'error':'Expected JSON array'}), 400
    count = 0
    for s in data:
        proj = s.get('project','')
        try:
            spool_type = s.get('spool_type', 'SPOOL') or 'SPOOL'
            if USE_PG:
                db_execute("INSERT INTO spools (project,spool_id,spool_full,iso_no,marking,mk_number,main_diameter,line,sequence,has_branches,spool_type) VALUES (?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT (project,spool_id) DO UPDATE SET has_branches=EXCLUDED.has_branches, spool_type=EXCLUDED.spool_type",
                    (proj, s['spool_id'], s.get('spool_full',''), s.get('iso_no',''), s.get('marking',''), s.get('mk_number',''), s.get('main_diameter',''), s.get('line',''), s.get('sequence',0), 1 if s.get('has_branches') else 0, spool_type))
            else:
                db_execute("INSERT INTO spools (project,spool_id,spool_full,iso_no,marking,mk_number,main_diameter,line,sequence,has_branches,spool_type) VALUES (?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(project,spool_id) DO UPDATE SET has_branches=excluded.has_branches, spool_type=excluded.spool_type",
                    (proj, s['spool_id'], s.get('spool_full',''), s.get('iso_no',''), s.get('marking',''), s.get('mk_number',''), s.get('main_diameter',''), s.get('line',''), s.get('sequence',0), 1 if s.get('has_branches') else 0, spool_type))
            has_br = 1 if s.get('has_branches') else 0
            steps_def = get_project_steps(proj)
            for step in steps_def:
                if step['is_conditional'] and not has_br:
                    continue
                st = step.get('spool_type', 'ALL') or 'ALL'
                if st != 'ALL' and st != spool_type:
                    continue
                if USE_PG:
                    db_execute("INSERT INTO progress (project,spool_id,step_number,completed) VALUES (?,?,?,0) ON CONFLICT (project,spool_id,step_number) DO NOTHING", (proj, s['spool_id'], step['step_number']))
                else:
                    db_execute("INSERT OR IGNORE INTO progress (project,spool_id,step_number,completed) VALUES (?,?,?,0)", (proj, s['spool_id'], step['step_number']))
            count += 1
        except Exception as e: print(f"Import error {s.get('spool_id','?')}: {e}")
    db_commit()
    return jsonify({'ok':True, 'imported':count})

@app.route('/api/project/<project>/export')
@login_required
def api_export(project):
    import openpyxl; from openpyxl.styles import Font, PatternFill, Alignment; import tempfile
    steps_def = get_project_steps(project)
    spools = db_fetchall("SELECT * FROM spools WHERE project=? ORDER BY sequence", (project,))
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = project
    hdr = ['#','Spool','Diameter','Line','Progress %'] + [f"S{s['step_number']}" for s in steps_def]
    hf = Font(bold=True,size=9,color='FFFFFF'); hfill = PatternFill(start_color='2F5496',end_color='2F5496',fill_type='solid')
    for c,h in enumerate(hdr,1):
        cl = ws.cell(1,c,h); cl.font=hf; cl.fill=hfill; cl.alignment=Alignment(horizontal='center',wrap_text=True)
    for i,s in enumerate(spools,2):
        p = spool_progress(project, s['spool_id'])
        steps = db_fetchall("SELECT step_number,completed FROM progress WHERE project=? AND spool_id=?", (project, s['spool_id']))
        sm = {st['step_number']:st['completed'] for st in steps}
        ws.cell(i,1,i-1); ws.cell(i,2,s['spool_id']); ws.cell(i,3,s['main_diameter']); ws.cell(i,4,s['line']); ws.cell(i,5,p)
        for j,sd in enumerate(steps_def):
            c = 6+j; done = sm.get(sd['step_number'],0)
            ws.cell(i,c,'\u2713' if done else '')
            if done: ws.cell(i,c).fill = PatternFill(start_color='C6EFCE',end_color='C6EFCE',fill_type='solid')
    ws.column_dimensions['B'].width=15; ws.freeze_panes='A2'
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False); wb.save(tmp.name)
    return send_file(tmp.name, as_attachment=True, download_name=f"{project}_progress_{date.today().strftime('%Y%m%d')}.xlsx")

@app.route('/api/project/<project>/spool/<spool_id>/weight', methods=['POST'])
@login_required
def api_update_weight(project, spool_id):
    d = request.get_json() or {}
    weight = d.get('weight_kg', 0); op = d.get('operator', '')
    db_execute("UPDATE spools SET actual_weight_kg=? WHERE project=? AND spool_id=?", (weight, project, spool_id))
    db_execute("INSERT INTO activity_log (project,spool_id,step_number,action,operator,details) VALUES (?,?,?,?,?,?)",
        (project, spool_id, 0, 'weight', op, f"Weight recorded: {weight} kg by {op}"))
    db_commit()
    return jsonify({'ok': True, 'weight_kg': weight})

@app.route('/api/project/<project>/surface', methods=['POST'])
def api_bulk_surface(project):
    data = request.get_json()
    if not data: return jsonify({'error': 'No JSON'}), 400
    count = 0
    for spool_id, surface in data.items():
        db_execute("UPDATE spools SET surface_m2=? WHERE project=? AND spool_id=?", (float(surface), project, spool_id))
        count += 1
    db_commit()
    return jsonify({'ok': True, 'updated': count})

@app.route('/api/project/<project>/joints', methods=['POST'])
def api_bulk_joints(project):
    data = request.get_json()
    if not data: return jsonify({'error': 'No JSON'}), 400
    count = 0
    for spool_id, vals in data.items():
        jc = vals.get('joint_count', 0) if isinstance(vals, dict) else int(vals)
        ri = vals.get('raf_inches', 0) if isinstance(vals, dict) else 0
        db_execute("UPDATE spools SET joint_count=?, raf_inches=? WHERE project=? AND spool_id=?", (jc, ri, project, spool_id))
        count += 1
    db_commit()
    return jsonify({'ok': True, 'updated': count})

@app.route('/api/project/<project>/spool/<spool_id>/drawing', methods=['POST'])
def api_upload_drawing(project, spool_id):
    if 'file' in request.files: pdf_data = request.files['file'].read()
    else: pdf_data = request.get_data()
    if not pdf_data: return jsonify({'error': 'No data'}), 400
    if USE_PG:
        import psycopg2
        db_execute("DELETE FROM drawings WHERE project=%s AND spool_id=%s", (project, spool_id))
        cur = get_db().cursor()
        cur.execute("INSERT INTO drawings (project, spool_id, pdf_data) VALUES (%s, %s, %s)", (project, spool_id, psycopg2.Binary(pdf_data)))
    else:
        db_execute("DELETE FROM drawings WHERE project=? AND spool_id=?", (project, spool_id))
        db_execute("INSERT INTO drawings (project, spool_id, pdf_data) VALUES (?,?,?)", (project, spool_id, pdf_data))
    db_commit()
    return jsonify({'ok': True, 'size_kb': round(len(pdf_data)/1024, 1)})

@app.route('/api/project/<project>/spool/<spool_id>/drawing')
@login_required
def api_get_drawing(project, spool_id):
    if USE_PG:
        cur = get_db().cursor()
        cur.execute("SELECT pdf_data FROM drawings WHERE project=%s AND spool_id=%s", (project, spool_id))
        row = cur.fetchone()
        if not row: return jsonify({'error': 'No drawing'}), 404
        pdf_data = bytes(row[0])
    else:
        row = db_fetchone("SELECT pdf_data FROM drawings WHERE project=? AND spool_id=?", (project, spool_id))
        if not row: return jsonify({'error': 'No drawing'}), 404
        pdf_data = row['pdf_data']
    from flask import Response
    return Response(pdf_data, mimetype='application/pdf', headers={'Content-Disposition': f'inline; filename={spool_id}.pdf'})

@app.route('/api/project/<project>/drawings/list')
@login_required
def api_list_drawings(project):
    if USE_PG:
        cur = get_db().cursor()
        cur.execute("SELECT spool_id, length(pdf_data) as size_bytes FROM drawings WHERE project=%s", (project,))
        rows = cur.fetchall()
        return jsonify([{'spool_id': r[0], 'size_kb': round(r[1]/1024,1)} for r in rows])
    else:
        rows = db_fetchall("SELECT spool_id, length(pdf_data) as size_bytes FROM drawings WHERE project=?", (project,))
        return jsonify([{'spool_id': r['spool_id'], 'size_kb': round(r['size_bytes']/1024,1)} for r in rows])

@app.route('/api/migrate', methods=['POST'])
def api_migrate():
    results = []
    try:
        if USE_PG:
            # Add columns/tables that might be missing on older deployments
            for col, tbl, default in [
                ('has_branches', 'spools', '0'), ('actual_weight_kg', 'spools', '0'),
                ('surface_m2', 'spools', '0'), ('joint_count', 'spools', '0'),
                ('raf_inches', 'spools', '0'), ('spool_type', 'spools', "'SPOOL'"),
            ]:
                try:
                    db_execute(f"ALTER TABLE {tbl} ADD COLUMN {col} {'INTEGER' if col in ('has_branches','joint_count') else 'REAL' if col in ('actual_weight_kg','surface_m2','raf_inches') else 'TEXT'} DEFAULT {default}")
                    results.append(f"added {col}")
                except: get_db().rollback(); results.append(f"{col} exists")
        db_commit(); results.append("migration complete")
    except Exception as e: results.append(str(e))
    return jsonify({'ok': True, 'results': results})

@app.route('/api/project/<project>/spool/<spool_id>/delete', methods=['POST'])
def api_delete_spool(project, spool_id):
    db_execute("DELETE FROM activity_log WHERE project=? AND spool_id=?", (project, spool_id))
    db_execute("DELETE FROM progress WHERE project=? AND spool_id=?", (project, spool_id))
    db_execute("DELETE FROM drawings WHERE project=? AND spool_id=?", (project, spool_id))
    db_execute("DELETE FROM spools WHERE project=? AND spool_id=?", (project, spool_id))
    db_commit()
    return jsonify({'ok': True, 'deleted': spool_id})

@app.route('/api/cleanup', methods=['POST'])
def api_cleanup():
    db_execute("DELETE FROM activity_log WHERE project=''")
    db_execute("DELETE FROM progress WHERE project=''")
    db_execute("DELETE FROM spools WHERE project=''")
    db_execute("DELETE FROM drawings WHERE NOT EXISTS (SELECT 1 FROM spools WHERE spools.project=drawings.project AND spools.spool_id=drawings.spool_id)")
    db_commit()
    return jsonify({'ok': True})

@app.route('/healthz')
def healthz():
    try: db_execute("SELECT 1"); return jsonify({'status':'ok','db':'connected'})
    except Exception as e: return jsonify({'status':'error','db':str(e)}), 500

# ── API: Project Steps ────────────────────────────────────────────────────────
@app.route('/api/project/<project>/steps', methods=['GET', 'POST'])
def api_project_steps(project):
    """Get or set ITP step definitions for a project."""
    if request.method == 'GET':
        steps = get_project_steps(project)
        return jsonify(steps)
    data = request.get_json()
    if not data or not isinstance(data, list): return jsonify({'error': 'Expected JSON array of step definitions'}), 400
    db_execute("DELETE FROM project_steps WHERE project=?", (project,))
    for s in data:
        cols = 'project,step_number,name_en,name_cn,weight,category,hours_fixed,hours_variable,spool_type,display_order,is_conditional,is_hold_point,is_release,phase,counts_for_production'
        vals = (project, s['step_number'], s['name_en'], s.get('name_cn',''), s.get('weight',5), s['category'],
                s.get('hours_fixed',2.0), s.get('hours_variable',''), s.get('spool_type','ALL'),
                s['display_order'], s.get('is_conditional',0), s.get('is_hold_point',0), s.get('is_release',0), s.get('phase','fab'), s.get('counts_for_production',1))
        ph = ','.join(['%s']*15) if USE_PG else ','.join(['?']*15)
        db_execute(f"INSERT INTO project_steps ({cols}) VALUES ({ph})", vals)
    db_commit()
    # Clear cache
    cache_key = f'_steps_{project}'
    if hasattr(g, cache_key): delattr(g, cache_key)
    return jsonify({'ok': True, 'steps': len(data)})

# ── API: Schedule & Reports ──────────────────────────────────────────────────
@app.route('/api/project/<project>/schedule', methods=['POST'])
def api_set_schedule(project):
    """Import schedule data. Accepts JSON array or {start: 'YYYY-MM-DD'} for auto-calculation."""
    data = request.get_json()
    if not data: return jsonify({'error': 'No JSON data'}), 400
    count = 0
    if isinstance(data, list):
        for item in data:
            if USE_PG:
                db_execute("INSERT INTO schedule (project,diameter,task_type,description,planned_start,planned_end,spool_count) VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (project,diameter,task_type) DO UPDATE SET description=EXCLUDED.description,planned_start=EXCLUDED.planned_start,planned_end=EXCLUDED.planned_end,spool_count=EXCLUDED.spool_count",
                    (project, item['diameter'], item['task_type'], item.get('description',''), item['planned_start'], item['planned_end'], item.get('spool_count',0)))
            else:
                db_execute("INSERT INTO schedule (project,diameter,task_type,description,planned_start,planned_end,spool_count) VALUES (?,?,?,?,?,?,?) ON CONFLICT(project,diameter,task_type) DO UPDATE SET description=excluded.description,planned_start=excluded.planned_start,planned_end=excluded.planned_end,spool_count=excluded.spool_count",
                    (project, item['diameter'], item['task_type'], item.get('description',''), item['planned_start'], item['planned_end'], item.get('spool_count',0)))
            count += 1
    elif isinstance(data, dict) and ('start' in data or 'fab_start' in data):
        start_date = date.fromisoformat(data.get('start', data.get('fab_start')))
        settings = get_project_settings(project)
        steps_def = get_project_steps(project)
        phase_order = get_phase_order(steps_def)
        std_weeks = int(settings.get('standard_weeks', '9'))
        secondary_phase_days = int(settings.get('secondary_phase_days', '13'))
        diam_order = get_diameter_order(project)
        spools_rows = db_fetchall("SELECT main_diameter FROM spools WHERE project=?", (project,))
        diam_counts = {}
        for s in spools_rows:
            d = (s['main_diameter'] or '?').replace('"','')
            if d == '?' or d == '0' or d == '': continue
            if d not in diam_counts: diam_counts[d] = 0
            diam_counts[d] += 1
        total_spools = sum(diam_counts.values())
        # Standard schedule: proportional distribution across standard_weeks
        # One diameter at a time, largest first, each gets (count/total) share of total days
        total_std_days = std_weeks * 7
        # For multi-phase: first phase gets (total_std_days - secondary_phase_days), second phase gets secondary_phase_days per diameter
        if len(phase_order) > 1:
            first_phase_total_days = total_std_days - secondary_phase_days  # last diameter's paint must finish by end
        else:
            first_phase_total_days = total_std_days  # single phase uses all days
        current_start = start_date
        remaining_days = first_phase_total_days
        first_phase = phase_order[0] if phase_order else 'fab'
        for i, diam in enumerate(diam_order):
            if diam not in diam_counts: continue
            cnt = diam_counts[diam]
            # Proportional: this diameter gets its share of total days
            if total_spools > 0:
                if i == len([d for d in diam_order if d in diam_counts]) - 1:
                    fab_days = remaining_days  # last diameter gets remainder
                else:
                    fab_days = max(1, round(first_phase_total_days * cnt / total_spools))
            else:
                fab_days = 1
            remaining_days -= fab_days
            fab_end = current_start + timedelta(days=fab_days)
            dk = f'{diam}"'
            phase_entries = []
            phase_entries.append((first_phase, current_start, fab_end, f'{first_phase.capitalize()} {dk} ({cnt} spools)'))
            # Secondary phases follow immediately after first phase for this diameter
            prev_end = fab_end
            for ph in phase_order[1:]:
                ph_start = prev_end + timedelta(days=1)
                ph_end = ph_start + timedelta(days=secondary_phase_days)
                phase_entries.append((ph, ph_start, ph_end, f'{ph.capitalize()} {dk}'))
                prev_end = ph_end
            current_start = fab_end  # next diameter starts when this one's first phase ends
            for tt, sd, ed, desc in phase_entries:
                if USE_PG:
                    db_execute("INSERT INTO schedule (project,diameter,task_type,description,planned_start,planned_end,spool_count) VALUES (%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (project,diameter,task_type) DO UPDATE SET description=EXCLUDED.description,planned_start=EXCLUDED.planned_start,planned_end=EXCLUDED.planned_end,spool_count=EXCLUDED.spool_count",
                        (project, dk, tt, desc, str(sd), str(ed), cnt))
                else:
                    db_execute("INSERT INTO schedule (project,diameter,task_type,description,planned_start,planned_end,spool_count) VALUES (?,?,?,?,?,?,?) ON CONFLICT(project,diameter,task_type) DO UPDATE SET description=excluded.description,planned_start=excluded.planned_start,planned_end=excluded.planned_end,spool_count=excluded.spool_count",
                        (project, dk, tt, desc, str(sd), str(ed), cnt))
                count += 1
    db_commit()
    return jsonify({'ok': True, 'entries': count})

@app.route('/api/project/<project>/schedule')
@login_required
def api_get_schedule(project):
    rows = db_fetchall("SELECT * FROM schedule WHERE project=? ORDER BY planned_start", (project,))
    return jsonify(fix_timestamps(rows))

@app.route('/api/project/<project>/settings', methods=['GET','POST'])
def api_project_settings(project):
    if request.method == 'GET':
        return jsonify(get_project_settings(project))
    data = request.get_json()
    if not data: return jsonify({'error':'No JSON'}), 400
    for k, v in data.items():
        if USE_PG:
            db_execute("INSERT INTO project_settings (project,key,value) VALUES (%s,%s,%s) ON CONFLICT (project,key) DO UPDATE SET value=EXCLUDED.value", (project, k, str(v)))
        else:
            db_execute("INSERT INTO project_settings (project,key,value) VALUES (?,?,?) ON CONFLICT(project,key) DO UPDATE SET value=excluded.value", (project, k, str(v)))
    db_commit()
    return jsonify({'ok': True, 'settings': get_project_settings(project)})

@app.route('/api/project/<project>/report')
@login_required
def api_report_data(project):
    import traceback as tb
    try:
        return jsonify(generate_report_data(project))
    except Exception as e:
        print(f"Report error: {e}\n{tb.format_exc()}")
        return jsonify({'error': str(e), 'trace': tb.format_exc()}), 500

@app.route('/project/<project>/report')
@login_required
def report_page(project):
    return render_template_string(REPORT_HTML, project=project)

@app.route('/api/project/<project>/report/download')
@login_required
def api_report_download(project):
    """Download a professional Excel production report."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile
    rpt = generate_report_data(project)
    st = rpt['stats']; sched = rpt.get('schedule'); sett = rpt.get('settings', {})
    fc_data = rpt.get('forecast', {}) or {}
    fc_diams = fc_data.get('diameters', {}) if fc_data else {}
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Production Report"
    hf = Font(bold=True, size=11, color='FFFFFF'); hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    dark_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')
    bf = Font(size=10); bfb = Font(bold=True, size=10)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    orange_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    thin = Border(left=Side('thin',color='C0C0C0'),right=Side('thin',color='C0C0C0'),top=Side('thin',color='C0C0C0'),bottom=Side('thin',color='C0C0C0'))
    center = Alignment(horizontal='center', vertical='center')
    std_weeks = int(sett.get('standard_weeks', '9'))
    wks_saved = int(sett.get('committed_weeks_saved', '0'))
    days_saved = int(sett.get('committed_days_saved', '0'))
    total_saved = wks_saved * 7 + days_saved
    has_expediting = total_saved > 0
    transit_days = int(sett.get('sea_transit_days', '45'))
    today = date.today()
    ws.merge_cells('A1:H1')
    ws['A1'] = f"ENERXON \u2014 Production Report / \u751f\u4ea7\u62a5\u544a"; ws['A1'].font = Font(bold=True, size=16, color='2F5496')
    ws['A2'] = f"Project: {project}"; ws['A2'].font = bfb
    ws['A3'] = f"Date: {rpt['date']}"; ws['A3'].font = bf
    ws['A4'] = f"Overall Progress: {st['overall_pct']}%"; ws['A4'].font = Font(bold=True, size=14, color='2F5496')
    ws['A5'] = f"Total: {st['total']} spools | Done: {st['completed']} | In Progress: {st['in_progress']} | Pending: {st['not_started']}"; ws['A5'].font = bf
    row = 7
    if sched and sched.get('diameters'):
        status_label = {'on_time': 'ON TIME \u2713', 'at_risk': 'AT RISK \u26a0', 'delayed': 'DELAYED \u2717', 'not_started': 'NOT STARTED'}
        ws.cell(row, 1, "SCHEDULE STATUS BY DIAMETER / \u6309\u7ba1\u5f84\u8ba1\u5212\u72b6\u6001").font = Font(bold=True, size=12, color='2F5496')
        row += 1
        phases = sched.get('phase_order', ['fab'])
        phase_colors = ['4472C4', 'ED7D31', '8E44AD', '27AE60']
        phase_headers = [f'{ph.capitalize()} %' for ph in phases]
        headers = ['Diameter','Spools'] + phase_headers + ['Overall %','Diff (days)','Status','Start','End','Forecast End']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row, col, h); c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = thin
        row += 1
        for d in sched['diameters']:
            dk = d['diameter']; fcd = fc_diams.get(dk, {})
            ws.cell(row, 1, dk).font = bfb; ws.cell(row, 1).border = thin
            ws.cell(row, 2, d['spool_count']).font = bf; ws.cell(row, 2).alignment = center; ws.cell(row, 2).border = thin
            col_idx = 3
            for ph in phases:
                ph_val = (d.get('phase_avgs') or {}).get(ph, 0)
                cp = ws.cell(row, col_idx, ph_val); cp.font = bfb; cp.alignment = center; cp.border = thin
                if ph_val >= 100: cp.fill = green_fill
                col_idx += 1
            ws.cell(row, col_idx, d['actual_pct']).font = bfb; ws.cell(row, col_idx).alignment = center; ws.cell(row, col_idx).border = thin
            col_idx += 1
            diff_val = d['diff']; diff_label = f"+{diff_val}d" if diff_val > 0 else f"{diff_val}d" if diff_val < 0 else "0d"
            c6 = ws.cell(row, col_idx, diff_label); c6.alignment = center; c6.border = thin
            c6.font = Font(size=10, color='27AE60' if diff_val > 0 else 'E74C3C' if diff_val < 0 else '333333')
            col_idx += 1
            sc = ws.cell(row, col_idx, status_label.get(d['status'], d['status']))
            sc.font = bfb; sc.alignment = center; sc.border = thin
            if d['status'] == 'on_time': sc.fill = green_fill
            elif d['status'] == 'at_risk': sc.fill = orange_fill
            elif d['status'] == 'delayed': sc.fill = red_fill
            col_idx += 1
            ws.cell(row, col_idx, d.get('total_start','')).font = bf; ws.cell(row, col_idx).border = thin
            col_idx += 1
            ws.cell(row, col_idx, d.get('total_end','')).font = bf; ws.cell(row, col_idx).border = thin
            col_idx += 1
            ws.cell(row, col_idx, fcd.get('forecast_end', '')).font = bf; ws.cell(row, col_idx).border = thin
            row += 1
        row += 1
        # Expediting + Gantt + Rate + Results + Transit — same as before
        prod_start = None
        starts = [d['total_start'] for d in sched['diameters'] if d.get('total_start')]
        if starts: prod_start = date.fromisoformat(min(starts))
        if prod_start and has_expediting:
            std_end = prod_start + timedelta(days=std_weeks * 7 - 1)
            commit_end = std_end - timedelta(days=total_saved)
            fc_overall_end = fc_data.get('overall_forecast_end')
            fc_end_d = date.fromisoformat(fc_overall_end) if fc_overall_end else None
            fc_diff_commit = (commit_end - fc_end_d).days if fc_end_d else 0
            commit_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            ws.cell(row, 1, "\u26a1 EXPEDITING COMMITMENT / \u52a0\u6025\u627f\u8bfa").font = Font(bold=True, size=12, color='4472C4')
            row += 1
            commit_items = [
                ('Start / \u5f00\u59cb', str(prod_start), '2F5496'),
                ('Standard End / \u6807\u51c6\u5b8c\u5de5', str(std_end), '888888'),
                ('Committed End / \u627f\u8bfa\u5b8c\u5de5', str(commit_end), '4472C4'),
                ('Saved / \u8282\u7701', f"{total_saved} days ({wks_saved} weeks)", '4472C4'),
                ('Forecast / \u9884\u6d4b', "{} {}".format(fc_end_d or '\u2014', '\u2713' if fc_diff_commit >= 0 else '\u2717'), '27AE60' if fc_diff_commit >= 0 else 'E74C3C'),
            ]
            for ci, (label, val, color) in enumerate(commit_items):
                c1 = ws.cell(row, ci*2 + 1, label); c1.font = Font(size=9, color='666666'); c1.fill = commit_fill; c1.border = thin
                c2 = ws.cell(row, ci*2 + 2, val); c2.font = Font(bold=True, size=11, color=color); c2.fill = commit_fill; c2.border = thin
            row += 2
        elif prod_start:
            std_end = prod_start + timedelta(days=std_weeks * 7 - 1)
            commit_end = std_end; fc_end_d = None
        else:
            std_end = None; commit_end = None; fc_end_d = None
        # Gantt
        ws.cell(row, 1, "PRODUCTION GANTT / \u751f\u4ea7\u7518\u7279\u56fe").font = Font(bold=True, size=12, color='2F5496')
        row += 1
        if prod_start:
            fc_overall_end = fc_data.get('overall_forecast_end')
            if not fc_end_d and fc_overall_end: fc_end_d = date.fromisoformat(fc_overall_end)
            if has_expediting: num_weeks = std_weeks
            else:
                max_d = std_end or prod_start + timedelta(days=62)
                if fc_end_d and fc_end_d > max_d: max_d = fc_end_d
                num_weeks = max(std_weeks, ((max_d - prod_start).days // 7) + 2)
            weeks = []
            for i in range(num_weeks):
                ws_date = prod_start + timedelta(days=i * 7); we_date = ws_date + timedelta(days=6)
                weeks.append((ws_date, we_date))
            ratio = (std_weeks - wks_saved) / std_weeks if has_expediting else 1.0
            gantt_hdr_row = row
            ws.cell(row, 1, 'Diameter').font = Font(bold=True, size=9, color='FFFFFF'); ws.cell(row, 1).fill = dark_fill; ws.cell(row, 1).border = thin
            ws.cell(row, 2, 'Phase').font = Font(bold=True, size=9, color='FFFFFF'); ws.cell(row, 2).fill = dark_fill; ws.cell(row, 2).border = thin
            ws.cell(row, 3, '%').font = Font(bold=True, size=9, color='FFFFFF'); ws.cell(row, 3).fill = dark_fill; ws.cell(row, 3).alignment = center; ws.cell(row, 3).border = thin
            today_col = None
            for i, (ws_d, we_d) in enumerate(weeks):
                col = 4 + i; wk_label = f"W{i+1}\n{ws_d.strftime('%d/%m')}"
                c = ws.cell(row, col, wk_label)
                is_current = ws_d <= today <= we_d
                if is_current: c.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'); today_col = col
                else: c.fill = dark_fill
                c.font = Font(bold=True, size=7, color='FFFFFF'); c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = thin
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 5.5
            row += 1
            phase_fills = [PatternFill(start_color=c, end_color=c, fill_type='solid') for c in phase_colors[:len(phases)]]
            saved_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

            today_border = Border(left=Side('thick',color='FF0000'),right=Side('thick',color='FF0000'),top=Side('thin',color='C0C0C0'),bottom=Side('thin',color='C0C0C0'))
            today_bar_border = Border(left=Side('thick',color='FF0000'),right=Side('thick',color='FF0000'),top=Side('medium',color='333333'),bottom=Side('medium',color='333333'))
            for d in sched['diameters']:
                dk = d['diameter']; fcd = fc_diams.get(dk, {})
                overall_pct = fcd.get('overall_pct', 0) or 0; dm_started = fcd.get('started', False)
                dm_fc_end_str = fcd.get('forecast_end', ''); dm_fc_end = date.fromisoformat(dm_fc_end_str) if dm_fc_end_str else None
                # Get phase percentages from phase_avgs
                dm_phase_avgs = fcd.get('phase_avgs', {}) or d.get('phase_avgs', {}) or {}
                # Parse schedule dates from phase_dates
                phase_dates = {}
                d_phase_dates = d.get('phase_dates', {})
                for ph in phases:
                    pd = d_phase_dates.get(ph, {})
                    if pd.get('start') and pd.get('end'):
                        phase_dates[ph] = (date.fromisoformat(pd['start']), date.fromisoformat(pd['end']))
                # Compute expedited dates per phase
                exp_phase_dates = {}
                prev_exp_end = None
                for pi, ph in enumerate(phases):
                    if ph not in phase_dates: continue
                    ph_ps, ph_pe = phase_dates[ph]
                    if has_expediting and commit_end:
                        exp_s = prod_start + timedelta(days=(ph_ps - prod_start).days * ratio)
                        exp_e = prod_start + timedelta(days=(ph_pe - prod_start).days * ratio)
                        if exp_e > commit_end: exp_e = commit_end
                        if prev_exp_end and exp_s < prev_exp_end: exp_s = prev_exp_end
                        if exp_e < exp_s: exp_e = exp_s
                    else:
                        exp_s = ph_ps; exp_e = ph_pe
                    exp_phase_dates[ph] = (exp_s, exp_e)
                    prev_exp_end = exp_e
                # Render one row per phase
                for pi, ph in enumerate(phases):
                    ph_pct = dm_phase_avgs.get(ph, 0) or 0
                    ph_color = phase_colors[pi % len(phase_colors)]
                    ph_fill = phase_fills[pi % len(phase_fills)]
                    is_first = pi == 0
                    # Diameter label only on first row
                    if is_first:
                        ws.cell(row, 1, dk).font = Font(bold=True, size=11, color='2F5496'); ws.cell(row, 1).border = thin
                    else:
                        ws.cell(row, 1, '').border = thin
                    ws.cell(row, 2, ph.capitalize()).font = Font(size=9, color='666666'); ws.cell(row, 2).border = thin
                    pct_cell = ws.cell(row, 3)
                    if ph_pct >= 100: pct_cell.value = '\u2713'; pct_cell.font = Font(bold=True, size=10, color='27AE60')
                    elif ph_pct > 0: pct_cell.value = f'{ph_pct:.0f}%'; pct_cell.font = Font(bold=True, size=8, color=ph_color)
                    else: pct_cell.value = '-'; pct_cell.font = Font(size=8, color='AAAAAA')
                    pct_cell.alignment = center; pct_cell.border = thin
                    if ph in phase_dates and ph in exp_phase_dates:
                        ph_ps, ph_pe = phase_dates[ph]
                        exp_s, exp_e = exp_phase_dates[ph]
                        is_last_phase = pi == len(phases) - 1
                        for i, (ws_d, we_d) in enumerate(weeks):
                            col = 4 + i; cell = ws.cell(row, col); cell.border = thin
                            in_std = ph_ps <= we_d and ph_pe >= ws_d
                            in_exp = exp_s <= we_d and exp_e >= ws_d
                            is_saved_cell = has_expediting and in_std and not in_exp
                            is_today_col = today_col and col == today_col
                            is_forecast = is_last_phase and dm_started and dm_fc_end and overall_pct < 100 and dm_fc_end >= ws_d and dm_fc_end <= we_d
                            if in_exp:
                                cell.fill = ph_fill
                                if is_today_col: cell.value = f'{ph_pct:.0f}%'; cell.font = Font(bold=True, size=7, color='FFFFFF'); cell.alignment = center
                                elif we_d < today: cell.value = '\u2713'; cell.font = Font(bold=True, size=8, color='FFFFFF'); cell.alignment = center
                            elif is_saved_cell:
                                cell.fill = saved_fill
                            elif is_forecast:
                                cell.border = Border(left=Side('medium',color='E74C3C'),right=Side('medium',color='E74C3C'),top=Side('medium',color='E74C3C'),bottom=Side('medium',color='E74C3C'))
                            if is_today_col: cell.border = today_bar_border if in_exp else today_border
                    else:
                        for i in range(len(weeks)):
                            col = 4 + i; cell = ws.cell(row, col); cell.border = thin
                            if today_col and col == today_col: cell.border = today_border
                    row += 1
            row += 1
            ws.cell(row, 1, 'Legend:').font = Font(bold=True, size=9)
            legend_col = 2
            for pi, ph in enumerate(phases):
                lg = ws.cell(row, legend_col, ph.capitalize()); lg.fill = phase_fills[pi % len(phase_fills)]; lg.font = Font(size=8, color='FFFFFF'); lg.alignment = center
                legend_col += 1
            if has_expediting:
                lg = ws.cell(row, legend_col, 'Saved \u2713'); lg.fill = saved_fill; lg.font = Font(size=8, color='A9D18E'); lg.alignment = center
                legend_col += 1
            lg = ws.cell(row, legend_col, '[ ] Forecast'); lg.font = Font(size=8, color='E74C3C'); lg.alignment = center; lg.border = Border(left=Side('medium',color='E74C3C'),right=Side('medium',color='E74C3C'),top=Side('medium',color='E74C3C'),bottom=Side('medium',color='E74C3C'))
            legend_col += 1
            ws.cell(row, legend_col, '| Today |').font = Font(bold=True, size=8, color='FF0000')
            row += 2
        # Production Rate
        actual_weld = fc_data.get('actual_weld_ipd', 0) or 0
        actual_paint = fc_data.get('actual_paint_m2d', 0) or 0
        weld_cap = fc_data.get('welding_capability', 0) or 0
        paint_cap = fc_data.get('painting_capability', 0) or 0
        ws.cell(row, 1, "PRODUCTION RATE / \u751f\u4ea7\u7387").font = Font(bold=True, size=12, color='2F5496')
        row += 1
        rate_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        rate_items = [('Welding / \u710a\u63a5', actual_weld, weld_cap, 'linear inches/day')]
        # Only show surface rate if project has surface steps
        has_surface = any(s.get('hours_variable') == 'surface' for s in get_project_steps(project))
        if has_surface:
            surface_label = phases[-1].capitalize() if len(phases) > 1 else 'Surface'
            rate_items.append((f'{surface_label} / \u6d82\u88c5', actual_paint, paint_cap, 'm\u00b2/day'))
        for label, actual, target, unit in rate_items:
            c1 = ws.cell(row, 1, label); c1.font = Font(bold=True, size=10); c1.fill = rate_fill; c1.border = thin
            c2 = ws.cell(row, 2, actual); c2.font = Font(bold=True, size=12, color='27AE60' if actual >= target else 'E74C3C'); c2.alignment = center; c2.fill = rate_fill; c2.border = thin
            c3 = ws.cell(row, 3, f"/ {target}"); c3.font = Font(size=9, color='888888'); c3.fill = rate_fill; c3.border = thin
            c4 = ws.cell(row, 4, unit); c4.font = Font(size=9, color='888888'); c4.fill = rate_fill; c4.border = thin
            pct_of_target = min(actual / max(target, 1) * 100, 100) if target else 0
            c5 = ws.cell(row, 5, f"{pct_of_target:.0f}% of target"); c5.font = Font(size=9, color='27AE60' if actual >= target else 'E74C3C'); c5.fill = rate_fill; c5.border = thin
            row += 1
        row += 1
        # Results Summary
        ws.cell(row, 1, "RESULTS SUMMARY / \u7ed3\u679c\u6458\u8981").font = Font(bold=True, size=12, color='2F5496')
        row += 1
        res_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        res_dark = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
        if has_expediting and std_end and commit_end:
            fc_saved = (std_end - fc_end_d).days if fc_end_d else 0
            fc_diff_c = (commit_end - fc_end_d).days if fc_end_d else 0
            res_green = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            results = [
                ('Overall Progress / \u603b\u8fdb\u5ea6', f"{st['overall_pct']}%", f"{st['total']} spools \u00b7 {st['in_progress']} WIP", res_fill, '2F5496'),
                ('Production End / \u751f\u4ea7\u5b8c\u5de5', f"{str(std_end)} \u2192 {str(commit_end)}", 'Standard \u2192 Committed (Expediting)', res_fill, '4472C4'),
                ('Expediting / \u52a0\u6025\u627f\u8bfa', f"{total_saved} days saved", f"{wks_saved} weeks with expediting fee", res_green, '27AE60'),
                ('Forecast / \u751f\u4ea7\u9884\u6d4b', f"{fc_saved} days saved \u00b7 ends {fc_end_d or chr(8212)}", ('\u2713 ' + str(fc_diff_c) + 'd ahead of commitment' if fc_diff_c >= 0 else '\u2717 ' + str(abs(fc_diff_c)) + 'd behind commitment') if fc_end_d else '', res_green, '27AE60' if fc_diff_c >= 0 else 'E74C3C'),
            ]
        else:
            fc_end_d = date.fromisoformat(fc_data.get('overall_forecast_end')) if fc_data and fc_data.get('overall_forecast_end') else None
            results = [
                ('Overall Progress / \u603b\u8fdb\u5ea6', f"{st['overall_pct']}%", f"{st['total']} spools \u00b7 {st['in_progress']} WIP", res_fill, '2F5496'),
                ('Forecast End / \u9884\u6d4b\u5b8c\u5de5', str(fc_end_d or '\u2014'), 'Based on actual rate', res_fill, '4472C4'),
            ]
        for label, val, sub, fill, color in results:
            c1 = ws.cell(row, 1, label); c1.font = Font(bold=True, size=10); c1.fill = fill; c1.border = thin
            c2 = ws.cell(row, 2, val); c2.font = Font(bold=True, size=11, color=color); c2.fill = fill; c2.border = thin
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            c5 = ws.cell(row, 5, sub); c5.font = Font(size=9, color='888888'); c5.fill = fill; c5.border = thin
            row += 1
        c1 = ws.cell(row, 1, 'Actual End / \u5b9e\u9645\u5b8c\u5de5'); c1.font = Font(bold=True, size=10, color='FFFFFF'); c1.fill = res_dark; c1.border = thin
        actual_end_val = str(today) if st['completed'] >= st['total'] else '\u2014'
        c2 = ws.cell(row, 2, actual_end_val); c2.font = Font(bold=True, size=11, color='FFFFFF'); c2.fill = res_dark; c2.border = thin
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        c5 = ws.cell(row, 5, 'Production complete' if st['completed'] >= st['total'] else 'Shown when complete'); c5.font = Font(size=9, color='AAAAAA'); c5.fill = res_dark; c5.border = thin
        row += 2
        # Transit
        if has_expediting and commit_end:
            commit_arrival = commit_end + timedelta(days=transit_days)
            fc_arrival = fc_end_d + timedelta(days=transit_days) if fc_end_d else None
            ship_fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
            ws.cell(row, 1, "\U0001f6a2 SEA TRANSIT / \u6d77\u8fd0").font = Font(bold=True, size=12, color='002060')
            row += 1
            c1 = ws.cell(row, 1, f"~{transit_days} days transit"); c1.font = Font(bold=True, size=10, color='FFFFFF'); c1.fill = ship_fill; c1.border = thin
            c2 = ws.cell(row, 2, f"Committed arrival: {commit_arrival}"); c2.font = Font(size=10, color='FFFFFF'); c2.fill = ship_fill; c2.border = thin
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            if fc_arrival:
                c5 = ws.cell(row, 5, f"Forecast arrival: {fc_arrival}"); c5.font = Font(size=10, color='FFFFFF'); c5.fill = ship_fill; c5.border = thin
            row += 2
    # Today's Activity
    if rpt.get('today_activity'):
        ws.cell(row, 1, f"TODAY'S ACTIVITY ({rpt['date']}) / \u4eca\u65e5\u52a8\u6001").font = Font(bold=True, size=12, color='2F5496')
        row += 1
        steps_today = rpt.get('steps_completed_today', 0)
        released = rpt.get('released_today', 0)
        past_rt = rpt.get('past_rt', 0)
        spools_touched = len(set(a.get('spool_id','') for a in rpt['today_activity']))
        ws.cell(row, 1, f"{steps_today} steps").font = Font(bold=True, size=11, color='2F5496'); ws.cell(row, 1).border = thin
        ws.cell(row, 2, f"{spools_touched} spools").font = Font(bold=True, size=11, color='2F5496'); ws.cell(row, 2).border = thin
        ws.cell(row, 3, f"{released} released").font = Font(bold=True, size=11, color='27AE60'); ws.cell(row, 3).border = thin
        ws.cell(row, 4, f"{past_rt} past RT").font = Font(bold=True, size=11, color='4472C4'); ws.cell(row, 4).border = thin
        row += 1
        for col, h in enumerate(['Time','Spool','Action','Operator','Details'], 1):
            c = ws.cell(row, col, h); c.font = hf; c.fill = hfill; c.border = thin
        row += 1
        for a in rpt['today_activity'][:50]:
            ts = str(a.get('timestamp',''))
            ws.cell(row, 1, ts[11:19] if len(ts)>11 else ts).font = bf; ws.cell(row, 1).border = thin
            ws.cell(row, 2, a.get('spool_id','')).font = bf; ws.cell(row, 2).border = thin
            ws.cell(row, 3, a.get('action','')).font = bf; ws.cell(row, 3).border = thin
            ws.cell(row, 4, a.get('operator','')).font = bf; ws.cell(row, 4).border = thin
            ws.cell(row, 5, a.get('details','')).font = bf; ws.cell(row, 5).border = thin
            row += 1
    ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 12; ws.column_dimensions['C'].width = 8
    ws.freeze_panes = 'A7'
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False); wb.save(tmp.name)
    return send_file(tmp.name, as_attachment=True, download_name=f"{project}_report_{today.strftime('%Y%m%d')}.xlsx")

@app.route('/api/project/<project>/report/pdf')
@login_required
def api_report_pdf(project):
    """Download a professional PDF production report (landscape A4)."""
    import tempfile
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm, inch
    from reportlab.lib.colors import HexColor, white, black, Color
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
    from reportlab.platypus.flowables import HRFlowable

    rpt = generate_report_data(project)
    st = rpt['stats']; sched = rpt.get('schedule'); sett = rpt.get('settings', {})
    fc_data = rpt.get('forecast', {}) or {}
    fc_diams = fc_data.get('diameters', {}) if fc_data else {}
    phases = rpt.get('phase_order', ['fab'])
    phase_colors = ['#4472C4', '#ED7D31', '#8E44AD', '#27AE60']
    today = date.today()

    # Settings
    std_weeks = int(sett.get('standard_weeks', '9'))
    wks_saved = int(sett.get('committed_weeks_saved', '0'))
    days_saved_val = int(sett.get('committed_days_saved', '0'))
    total_saved = wks_saved * 7 + days_saved_val
    has_expediting = total_saved > 0
    transit_days = int(sett.get('sea_transit_days', '45'))

    # Colors
    BLUE = HexColor('#2F5496')
    DARK = HexColor('#404040')
    LIGHT_BLUE = HexColor('#D9E2F3')
    LIGHT_GREEN = HexColor('#E2EFDA')
    LIGHT_GRAY = HexColor('#F2F2F2')
    GREEN = HexColor('#27AE60')
    RED = HexColor('#E74C3C')
    ORANGE = HexColor('#F39C12')
    NAVY = HexColor('#002060')
    WHITE = white
    BLACK = black

    # Styles
    styles = getSampleStyleSheet()
    s_title = ParagraphStyle('RPT_Title', parent=styles['Title'], fontSize=18, textColor=BLUE, spaceAfter=2, alignment=TA_LEFT)
    s_heading = ParagraphStyle('RPT_H2', parent=styles['Heading2'], fontSize=13, textColor=BLUE, spaceBefore=6, spaceAfter=4, alignment=TA_LEFT)
    s_normal = ParagraphStyle('RPT_Normal', parent=styles['Normal'], fontSize=9, textColor=BLACK, leading=12)
    s_small = ParagraphStyle('RPT_Small', parent=styles['Normal'], fontSize=8, textColor=HexColor('#888888'), leading=10)
    s_center = ParagraphStyle('RPT_Center', parent=styles['Normal'], fontSize=9, textColor=BLACK, alignment=TA_CENTER, leading=11)
    s_center_bold = ParagraphStyle('RPT_CenterBold', parent=styles['Normal'], fontSize=9, textColor=BLACK, alignment=TA_CENTER, leading=11, fontName='Helvetica-Bold')
    s_cell = ParagraphStyle('RPT_Cell', parent=styles['Normal'], fontSize=8, textColor=BLACK, leading=10, alignment=TA_CENTER)
    s_cell_left = ParagraphStyle('RPT_CellL', parent=styles['Normal'], fontSize=8, textColor=BLACK, leading=10, alignment=TA_LEFT)
    s_cell_bold = ParagraphStyle('RPT_CellB', parent=styles['Normal'], fontSize=8, textColor=BLACK, leading=10, alignment=TA_CENTER, fontName='Helvetica-Bold')

    def p(text, style=s_normal):
        return Paragraph(str(text), style)

    def color_for_status(status):
        return {'on_time': GREEN, 'at_risk': ORANGE, 'delayed': RED}.get(status, HexColor('#95A5A6'))

    def status_label(status):
        return {'on_time': 'ON TIME', 'at_risk': 'AT RISK', 'delayed': 'DELAYED', 'not_started': 'NOT STARTED'}.get(status, status)

    page_w, page_h = landscape(A4)
    tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
    doc = SimpleDocTemplate(tmp.name, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    story = []

    # ── 1. Title Section ─────────────────────────────────────────────────────
    story.append(p('ENERXON \u2014 Production Report', s_title))
    story.append(Spacer(1, 2*mm))
    story.append(p(f'<b>Project:</b> {project} &nbsp;&nbsp;&nbsp; <b>Date:</b> {rpt["date"]}', s_normal))
    story.append(Spacer(1, 3*mm))
    s_big_pct = ParagraphStyle('RPT_BigPct', parent=styles['Normal'], fontSize=18, textColor=BLUE, fontName='Helvetica-Bold', leading=22)
    story.append(p(f'{st["overall_pct"]}% <font size="10" color="#888888">Overall Progress</font>', s_big_pct))
    story.append(Spacer(1, 2*mm))
    story.append(p(f'Total: {st["total"]} spools &nbsp;|&nbsp; Done: {st["completed"]} &nbsp;|&nbsp; In Progress: {st["in_progress"]} &nbsp;|&nbsp; Pending: {st["not_started"]}', s_small))
    story.append(Spacer(1, 8*mm))

    # ── 2. Schedule Status Table ─────────────────────────────────────────────
    if sched and sched.get('diameters'):
        story.append(p('SCHEDULE STATUS BY DIAMETER', s_heading))
        # Build headers dynamically
        phase_hdrs = [f'{ph.capitalize()} %' for ph in phases]
        headers = ['Diameter', 'Spools'] + phase_hdrs + ['Overall %', 'Diff', 'Status', 'Start', 'End', 'Forecast End']
        hdr_row = [p(f'<font color="white"><b>{h}</b></font>', s_cell) for h in headers]
        data_rows = [hdr_row]
        for d in sched['diameters']:
            dk = d['diameter']; fcd = fc_diams.get(dk, {})
            row_data = [p(f'<b>{dk}</b>', s_cell_left), p(str(d['spool_count']), s_cell)]
            for ph in phases:
                ph_val = (d.get('phase_avgs') or {}).get(ph, 0)
                row_data.append(p(f'<b>{ph_val}%</b>', s_cell_bold))
            row_data.append(p(f'<b>{d["actual_pct"]}%</b>', s_cell_bold))
            diff_val = d['diff']
            diff_str = f'+{diff_val}d' if diff_val > 0 else f'{diff_val}d' if diff_val < 0 else '0d'
            diff_color = '#27AE60' if diff_val > 0 else '#E74C3C' if diff_val < 0 else '#333333'
            row_data.append(p(f'<font color="{diff_color}">{diff_str}</font>', s_cell))
            s_color = {'on_time': '#27AE60', 'at_risk': '#F39C12', 'delayed': '#E74C3C'}.get(d['status'], '#95A5A6')
            row_data.append(p(f'<font color="{s_color}"><b>{status_label(d["status"])}</b></font>', s_cell))
            # Date columns — total start/end
            row_data.append(p(d.get('total_start', ''), s_cell))
            row_data.append(p(d.get('total_end', ''), s_cell))
            row_data.append(p(fcd.get('forecast_end', ''), s_cell))
            data_rows.append(row_data)

        num_cols = len(headers)
        avail_w = page_w - 30*mm
        # Proportional column widths
        base_widths = [42, 30] + [35]*len(phases) + [38, 32, 50, 48, 48, 48]
        total_base = sum(base_widths)
        col_widths = [w / total_base * avail_w for w in base_widths]

        t = Table(data_rows, colWidths=col_widths, repeatRows=1)
        t_style = [
            ('BACKGROUND', (0,0), (-1,0), BLUE),
            ('TEXTCOLOR', (0,0), (-1,0), WHITE),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, HexColor('#C0C0C0')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [WHITE, HexColor('#FAFBFD')]),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]
        # Highlight phase % cells >= 100 green
        phase_col_start = 2
        for ri, d in enumerate(sched['diameters'], 1):
            for pi, ph in enumerate(phases):
                ph_val = (d.get('phase_avgs') or {}).get(ph, 0)
                if ph_val >= 100:
                    t_style.append(('BACKGROUND', (phase_col_start+pi, ri), (phase_col_start+pi, ri), HexColor('#C6EFCE')))
            # Status cell color
            status_col = 2 + len(phases) + 2  # after phases, overall, diff
            s_bg = {'on_time': '#C6EFCE', 'at_risk': '#FCE4D6', 'delayed': '#FFC7CE'}.get(d['status'])
            if s_bg:
                t_style.append(('BACKGROUND', (status_col, ri), (status_col, ri), HexColor(s_bg)))
        t.setStyle(TableStyle(t_style))
        story.append(t)
        story.append(Spacer(1, 6*mm))

        # ── 3. Expediting Commitment Panel ───────────────────────────────────
        prod_start = None
        starts = [d['total_start'] for d in sched['diameters'] if d.get('total_start')]
        if starts:
            prod_start = date.fromisoformat(min(starts))
        if prod_start and has_expediting:
            std_end = prod_start + timedelta(days=std_weeks * 7 - 1)
            commit_end = std_end - timedelta(days=total_saved)
            fc_overall_end = fc_data.get('overall_forecast_end')
            fc_end_d = date.fromisoformat(fc_overall_end) if fc_overall_end else None
            fc_diff_commit = (commit_end - fc_end_d).days if fc_end_d else 0

            s_exp_card = ParagraphStyle('RPT_ExpCard', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, leading=20)
            exp_row = [
                p(f'<font color="#2F5496" size="14"><b>{prod_start.strftime("%d %b")}</b></font><br/><font size="8" color="#333333"><b>START</b></font>', s_exp_card),
                p(f'<font color="#888888" size="14"><s>{std_end.strftime("%d %b")}</s></font><br/><font size="8" color="#333333"><b>STANDARD END</b></font>', s_exp_card),
                p(f'<font color="#4472C4" size="14"><b>{commit_end.strftime("%d %b")}</b></font><br/><font size="8" color="#333333"><b>COMMITTED END</b></font>', s_exp_card),
                p(f'<font color="#4472C4" size="16"><b>{total_saved}d</b></font><br/><font size="8" color="#333333"><b>SAVED ({wks_saved} WK)</b></font>', s_exp_card),
                p(f'<font color="{"#27AE60" if fc_diff_commit >= 0 else "#E74C3C"}" size="14"><b>{fc_end_d.strftime("%d %b") if fc_end_d else chr(8212)}</b></font><br/><font size="8" color="#333333"><b>FORECAST {"&#10003;" if fc_diff_commit >= 0 else "&#10007;"}</b></font>', s_exp_card),
            ]
            exp_w = avail_w / 5
            exp_t = Table([exp_row], colWidths=[exp_w]*5)
            exp_t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), LIGHT_BLUE),
                ('GRID', (0,0), (-1,-1), 0.3, HexColor('#C0D4E8')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('TOPPADDING', (0,0), (-1,-1), 8),
                ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ]))
            story.append(KeepTogether([p('EXPEDITING COMMITMENT', s_heading), exp_t]))
            story.append(Spacer(1, 6*mm))
        elif prod_start:
            std_end = prod_start + timedelta(days=std_weeks * 7 - 1)
            commit_end = std_end; fc_end_d = None
        else:
            std_end = None; commit_end = None; fc_end_d = None; prod_start = None

        # ── 4. Production Gantt ──────────────────────────────────────────────
        gantt_elems = [p('PRODUCTION GANTT', s_heading)]  # Will be wrapped in KeepTogether
        if prod_start:
            fc_overall_end = fc_data.get('overall_forecast_end')
            if not fc_end_d and fc_overall_end:
                fc_end_d = date.fromisoformat(fc_overall_end)
            if has_expediting:
                num_weeks = std_weeks
            else:
                max_d = std_end or prod_start + timedelta(days=62)
                if fc_end_d and fc_end_d > max_d:
                    max_d = fc_end_d
                num_weeks = max(std_weeks, ((max_d - prod_start).days // 7) + 2)

            weeks = []
            for i in range(num_weeks):
                ws_date = prod_start + timedelta(days=i * 7)
                we_date = ws_date + timedelta(days=6)
                weeks.append((ws_date, we_date))

            ratio = (std_weeks - wks_saved) / std_weeks if has_expediting else 1.0

            # Build Gantt table
            gantt_hdrs = ['Diameter', 'Phase'] + [f'W{i+1}\n{ws.strftime("%d/%m")}-{we.strftime("%d/%m")}' for i, (ws, we) in enumerate(weeks)]
            gantt_data = []
            # Header row
            hdr = [p(f'<font color="white"><b>{h}</b></font>', s_cell) for h in gantt_hdrs]
            gantt_data.append(hdr)

            # Determine today column
            today_week_idx = None
            for i, (ws_d, we_d) in enumerate(weeks):
                if ws_d <= today <= we_d:
                    today_week_idx = i
                    break

            for dm_idx, d in enumerate(sched['diameters']):
                dk = d['diameter']; fcd = fc_diams.get(dk, {})
                overall_pct = fcd.get('overall_pct', 0) or 0
                dm_started = fcd.get('started', False)
                dm_fc_end_str = fcd.get('forecast_end', '')
                dm_fc_end = date.fromisoformat(dm_fc_end_str) if dm_fc_end_str else None
                dm_phase_avgs = fcd.get('phase_avgs', {}) or d.get('phase_avgs', {}) or {}

                # Parse schedule dates from phase_dates
                phase_dates = {}
                d_phase_dates = d.get('phase_dates', {})
                for ph in phases:
                    pd = d_phase_dates.get(ph, {})
                    if pd.get('start') and pd.get('end'):
                        phase_dates[ph] = (date.fromisoformat(pd['start']), date.fromisoformat(pd['end']))

                # Compute expedited dates per phase
                exp_phase_dates = {}
                prev_exp_end = None
                for pi, ph in enumerate(phases):
                    if ph not in phase_dates:
                        continue
                    ph_ps, ph_pe = phase_dates[ph]
                    if has_expediting and commit_end:
                        exp_s = prod_start + timedelta(days=int((ph_ps - prod_start).days * ratio))
                        exp_e = prod_start + timedelta(days=int((ph_pe - prod_start).days * ratio))
                        if exp_e > commit_end:
                            exp_e = commit_end
                        if prev_exp_end and exp_s < prev_exp_end:
                            exp_s = prev_exp_end
                        if exp_e < exp_s:
                            exp_e = exp_s
                    else:
                        exp_s = ph_ps; exp_e = ph_pe
                    exp_phase_dates[ph] = (exp_s, exp_e)
                    prev_exp_end = exp_e

                # One row per phase
                for pi, ph in enumerate(phases):
                    ph_pct = dm_phase_avgs.get(ph, 0) or 0
                    ph_hex = phase_colors[pi % len(phase_colors)]
                    is_first = pi == 0
                    is_last_phase = pi == len(phases) - 1
                    row = []
                    # Diameter label only on first phase row
                    if is_first:
                        row.append(p(f'<b>{dk}</b><br/><font size="6" color="#888888">{d["spool_count"]} spools</font>', s_cell_left))
                    else:
                        row.append(p('', s_cell))
                    # Phase label
                    row.append(p(f'<font color="#666666">{ph.capitalize()}</font>', s_cell))
                    # Week cells
                    for wi, (ws_d, we_d) in enumerate(weeks):
                        cell_text = ''
                        cell_bg = None
                        if ph in phase_dates and ph in exp_phase_dates:
                            ph_ps, ph_pe = phase_dates[ph]
                            exp_s, exp_e = exp_phase_dates[ph]
                            in_std = ph_ps <= we_d and ph_pe >= ws_d
                            in_exp = exp_s <= we_d and exp_e >= ws_d
                            is_saved_cell = has_expediting and in_std and not in_exp
                            is_forecast = is_last_phase and dm_started and dm_fc_end and overall_pct < 100 and dm_fc_end >= ws_d and dm_fc_end <= we_d

                            if in_exp:
                                cell_bg = HexColor(ph_hex)
                                if today_week_idx is not None and wi == today_week_idx:
                                    cell_text = f'{ph_pct:.0f}%'
                                elif we_d < today:
                                    cell_text = '\u2713'
                            elif is_saved_cell:
                                cell_bg = LIGHT_GREEN
                            if is_forecast:
                                if not in_exp: cell_bg = None
                                # Mark for dashed border (applied via TableStyle below)
                                if '_gantt_forecast' not in d: d['_gantt_forecast'] = []
                                d['_gantt_forecast'].append((pi, wi))
                        # Text color
                        if cell_text:
                            txt_color = 'white' if cell_bg and cell_bg != LIGHT_GREEN else '#333'
                            row.append(p(f'<font color="{txt_color}" size="7"><b>{cell_text}</b></font>', s_cell))
                        else:
                            row.append(p('', s_cell))
                        # Store bg info for styling later
                        if '_gantt_bgs' not in d:
                            d['_gantt_bgs'] = {}
                        d['_gantt_bgs'][(pi, wi)] = cell_bg
                    gantt_data.append(row)

            # Column widths for Gantt
            diam_col_w = 55
            phase_col_w = 32
            week_col_w = max(14, (avail_w - diam_col_w - phase_col_w) / max(num_weeks, 1))
            gantt_col_widths = [diam_col_w, phase_col_w] + [week_col_w] * num_weeks

            gantt_t = Table(gantt_data, colWidths=gantt_col_widths, repeatRows=1)
            gantt_style = [
                ('BACKGROUND', (0,0), (-1,0), DARK),
                ('TEXTCOLOR', (0,0), (-1,0), WHITE),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 7),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('ALIGN', (0,0), (0,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.3, HexColor('#E0E0E0')),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
                ('LEFTPADDING', (0,0), (-1,-1), 2),
                ('RIGHTPADDING', (0,0), (-1,-1), 2),
            ]
            # Today column highlight
            if today_week_idx is not None:
                tc = today_week_idx + 2  # offset for Diameter+Phase columns
                gantt_style.append(('BACKGROUND', (tc, 0), (tc, 0), HexColor('#C00000')))
                # Red borders on today column for all data rows
                for ri in range(1, len(gantt_data)):
                    gantt_style.append(('LINEAFTER', (tc, ri), (tc, ri), 1.5, RED))
                    gantt_style.append(('LINEBEFORE', (tc, ri), (tc, ri), 1.5, RED))

            # Apply phase bar backgrounds
            data_row_idx = 1
            for dm_idx, d in enumerate(sched['diameters']):
                for pi, ph in enumerate(phases):
                    for wi in range(num_weeks):
                        bg = d.get('_gantt_bgs', {}).get((pi, wi))
                        if bg:
                            gantt_style.append(('BACKGROUND', (wi+2, data_row_idx), (wi+2, data_row_idx), bg))
                    data_row_idx += 1
                # Merge diameter cells across phase rows
                if len(phases) > 1:
                    first_row = data_row_idx - len(phases)
                    gantt_style.append(('SPAN', (0, first_row), (0, first_row + len(phases) - 1)))

            # Apply forecast dashed red borders
            data_row_idx = 1
            for dm_idx, d in enumerate(sched['diameters']):
                for pi, ph in enumerate(phases):
                    for (fpi, fwi) in d.get('_gantt_forecast', []):
                        if fpi == pi:
                            gantt_style.append(('BOX', (fwi+2, data_row_idx), (fwi+2, data_row_idx), 1.5, RED))
                    data_row_idx += 1

            gantt_t.setStyle(TableStyle(gantt_style))
            # Clean up temp data
            for d in sched['diameters']:
                d.pop('_gantt_bgs', None)
                d.pop('_gantt_forecast', None)

            gantt_elems.append(gantt_t)

            # Legend
            legend_items = []
            for pi, ph in enumerate(phases):
                legend_items.append(f'<font color="{phase_colors[pi % len(phase_colors)]}">\u25a0</font> {ph.capitalize()}')
            if has_expediting:
                legend_items.append('<font color="#A9D18E">\u25a0</font> Saved')
            legend_items.append('<font color="#E74C3C">[ ]</font> Forecast')
            legend_items.append('<font color="#E74C3C">|</font> Today')
            gantt_elems.append(p('&nbsp;&nbsp;&nbsp;'.join(legend_items), s_small))
            story.append(KeepTogether(gantt_elems))
            story.append(Spacer(1, 6*mm))

        # ── 5. Production Rate ───────────────────────────────────────────────
        actual_weld = fc_data.get('actual_weld_ipd', 0) or 0
        actual_paint = fc_data.get('actual_paint_m2d', 0) or 0
        weld_cap = fc_data.get('welding_capability', 0) or 0
        paint_cap = fc_data.get('painting_capability', 0) or 0
        steps_def = get_project_steps(project)
        has_surface = any(s.get('hours_variable') == 'surface' for s in steps_def)

        rate_elems = [p('PRODUCTION RATE', s_heading)]
        rate_data = []
        weld_color = '#27AE60' if actual_weld >= weld_cap else '#E74C3C'
        weld_pct = min(actual_weld / max(weld_cap, 1) * 100, 100) if weld_cap else 0
        rate_data.append([
            p('<b>Welding</b>', s_cell_left),
            p(f'<font color="{weld_color}" size="12"><b>{actual_weld}</b></font>', s_cell),
            p(f'/ {weld_cap}', s_cell),
            p('linear inches/day', s_cell),
            p(f'<font color="{weld_color}">{weld_pct:.0f}% of target</font>', s_cell),
        ])
        if has_surface:
            surface_label = phases[-1].capitalize() if len(phases) > 1 else 'Surface'
            paint_color = '#27AE60' if actual_paint >= paint_cap else '#E74C3C'
            paint_pct = min(actual_paint / max(paint_cap, 1) * 100, 100) if paint_cap else 0
            rate_data.append([
                p(f'<b>{surface_label}</b>', s_cell_left),
                p(f'<font color="{paint_color}" size="12"><b>{actual_paint}</b></font>', s_cell),
                p(f'/ {paint_cap}', s_cell),
                p('m\u00b2/day', s_cell),
                p(f'<font color="{paint_color}">{paint_pct:.0f}% of target</font>', s_cell),
            ])
        rate_widths = [avail_w * f for f in [0.22, 0.15, 0.1, 0.2, 0.33]]
        rate_t = Table(rate_data, colWidths=rate_widths)
        rate_t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), LIGHT_GRAY),
            ('GRID', (0,0), (-1,-1), 0.5, HexColor('#D0D0D0')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        rate_elems.append(rate_t)
        story.append(KeepTogether(rate_elems))
        story.append(Spacer(1, 6*mm))

        # ── 6. Results Summary — Visual KPI cards ────────────────────────────
        if has_expediting and std_end and commit_end:
            if not fc_end_d and fc_data and fc_data.get('overall_forecast_end'):
                fc_end_d = date.fromisoformat(fc_data['overall_forecast_end'])
            fc_saved = (std_end - fc_end_d).days if fc_end_d else 0
            fc_diff_c = (commit_end - fc_end_d).days if fc_end_d else 0
            cards = [
                ('#2F5496', f'{st["overall_pct"]}%', 'PROGRESS', f'{st["total"]} spools | {st["completed"]} done'),
                ('#4472C4', f'{commit_end.strftime("%d %b %Y")}', 'COMMITTED END', f'{total_saved}d saved ({wks_saved} wk)'),
                ('#27AE60' if fc_diff_c >= 0 else '#E74C3C',
                 f'{fc_end_d.strftime("%d %b %Y") if fc_end_d else chr(8212)}',
                 'FORECAST END',
                 f'{"+" if fc_diff_c >= 0 else ""}{fc_diff_c}d vs commitment' if fc_end_d else ''),
                ('white', str(today) if st['completed'] >= st['total'] else chr(8212), 'ACTUAL END', 'Shown when complete'),
            ]
        else:
            fc_end_d_val = date.fromisoformat(fc_data.get('overall_forecast_end')) if fc_data and fc_data.get('overall_forecast_end') else None
            cards = [
                ('#2F5496', f'{st["overall_pct"]}%', 'PROGRESS', f'{st["total"]} spools | {st["completed"]} done'),
                ('#4472C4', fc_end_d_val.strftime('%d %b %Y') if fc_end_d_val else chr(8212), 'FORECAST END', 'Based on actual rate'),
                ('white', str(today) if st['completed'] >= st['total'] else chr(8212), 'ACTUAL END', 'Shown when complete'),
            ]
        # Build as a row of cards — big number on top, label below
        s_card = ParagraphStyle('RPT_Card', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, leading=20)
        card_row = []
        for ci, (color, val, label, sub) in enumerate(cards):
            is_dark = ci == len(cards) - 1  # last card = dark bg
            lbl_color = 'white' if is_dark else '#333333'
            sub_color = '#AAAAAA' if is_dark else '#888888'
            card_row.append(p(
                f'<font color="{color}" size="16"><b>{val}</b></font><br/>'
                f'<font color="{lbl_color}" size="8"><b>{label}</b></font><br/>'
                f'<font color="{sub_color}" size="7">{sub}</font>', s_card))
        card_w = avail_w / len(cards)
        card_t = Table([card_row], colWidths=[card_w]*len(cards))
        card_style = [
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 10), ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ('BACKGROUND', (0,0), (-2,-1), LIGHT_BLUE),
            ('BACKGROUND', (-1,0), (-1,-1), DARK),
            ('BOX', (0,0), (-1,-1), 0.5, HexColor('#C0D4E8')),
        ]
        for ci in range(len(cards)):
            card_style.append(('LINEBEFORE', (ci, 0), (ci, -1), 0.3, HexColor('#C0D4E8')))
        card_t.setStyle(TableStyle(card_style))
        story.append(KeepTogether([p('RESULTS SUMMARY', s_heading), card_t]))
        story.append(Spacer(1, 6*mm))

        # ── 7. Sea Transit ───────────────────────────────────────────────────
        if has_expediting and commit_end:
            if not fc_end_d and fc_data and fc_data.get('overall_forecast_end'):
                fc_end_d = date.fromisoformat(fc_data['overall_forecast_end'])
            commit_arrival = commit_end + timedelta(days=transit_days)
            fc_arrival = fc_end_d + timedelta(days=transit_days) if fc_end_d else None
            # Visual KPI cards for transit
            s_transit_card = ParagraphStyle('RPT_TransitCard', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, leading=20)
            transit_row = [
                p(f'<font color="white" size="16"><b>~{transit_days}d</b></font><br/><font size="8" color="#99BBDD"><b>SEA TRANSIT</b></font>', s_transit_card),
                p(f'<font color="white" size="14"><b>{commit_arrival.strftime("%d %b %Y")}</b></font><br/><font size="8" color="#99BBDD"><b>COMMITTED ARRIVAL</b></font>', s_transit_card),
                p(f'<font color="white" size="14"><b>{fc_arrival.strftime("%d %b %Y") if fc_arrival else chr(8212)}</b></font><br/><font size="8" color="#99BBDD"><b>FORECAST ARRIVAL</b></font>', s_transit_card),
            ]
            transit_t = Table([transit_row], colWidths=[avail_w/3]*3)
            transit_t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), NAVY),
                ('VALIGN', (0,0), (-1,0), 'BOTTOM'), ('VALIGN', (0,1), (-1,1), 'TOP'),
                ('TOPPADDING', (0,0), (-1,0), 8), ('BOTTOMPADDING', (0,0), (-1,0), 0),
                ('TOPPADDING', (0,1), (-1,1), 0), ('BOTTOMPADDING', (0,1), (-1,1), 6),
                ('BOX', (0,0), (-1,-1), 0.5, HexColor('#003366')),
            ]))
            story.append(KeepTogether([p('SEA TRANSIT', s_heading), transit_t]))

    doc.build(story)
    return send_file(tmp.name, as_attachment=True, download_name=f"{project}_report_{today.strftime('%Y%m%d')}.pdf",
                     mimetype='application/pdf')

# ── HTML Templates ───────────────────────────────────────────────────────────
COMMON_CSS = """*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,'PingFang SC','Microsoft YaHei',sans-serif;background:#f0f2f5;color:#333}
.header{background:linear-gradient(135deg,#2F5496,#1a3a6e);color:#fff;padding:16px 20px}
.header h1{font-size:20px}.header .sub{font-size:12px;opacity:.8;margin-top:4px}
.back{color:#fff;text-decoration:none;font-size:13px;opacity:.8}.back:hover{opacity:1}
.pbar-bg{background:#e8e8e8;border-radius:8px;height:12px;overflow:hidden}
.pbar-fill{height:100%;border-radius:8px;transition:width .5s}
.pct-green{color:#27ae60}.pct-yellow{color:#f39c12}.pct-red{color:#e74c3c}.pct-blue{color:#2F5496}
.btn{background:#2F5496;color:#fff;border:none;padding:8px 16px;border-radius:6px;cursor:pointer;font-size:13px;text-decoration:none;display:inline-block}
.btn:hover{background:#1a3a6e}
.line-badge{display:inline-block;width:22px;height:22px;border-radius:50%;color:#fff;text-align:center;line-height:22px;font-size:11px;font-weight:700}
.line-A{background:#2d8a4e}.line-B{background:#2F5496}.line-C{background:#c0392b}
.stats-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:12px;padding:16px}
.stat-card{background:#fff;border-radius:10px;padding:16px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.1)}
.stat-card .value{font-size:28px;font-weight:700;color:#2F5496}.stat-card .label{font-size:11px;color:#888;margin-top:4px}
@media(max-width:600px){.stats-grid{grid-template-columns:repeat(2,1fr)}}
/* ── ENERXON Chat Assistant Widget (shared across all pages) ───────────── */
.chat-btn-robot{position:fixed;top:12px;right:60px;z-index:900;width:44px;height:44px;border-radius:10px;background:#fff;border:1px solid rgba(255,255,255,.4);box-shadow:0 2px 8px rgba(0,0,0,.18);cursor:pointer;display:flex;align-items:center;justify-content:center;padding:0;transition:transform .15s,box-shadow .15s}
.chat-btn-robot:hover{transform:translateY(-1px);box-shadow:0 4px 12px rgba(0,0,0,.22)}
.chat-btn-robot svg{display:block;width:28px;height:28px}
.chat-btn-robot .chat-pulse{position:absolute;top:-3px;right:-3px;width:11px;height:11px;background:#27ae60;border:2px solid #fff;border-radius:50%}
.chat-backdrop{position:fixed;inset:0;background:rgba(0,0,0,.25);z-index:999;opacity:0;pointer-events:none;transition:opacity .25s}
.chat-backdrop.open{opacity:1;pointer-events:auto}
.chat-panel{position:fixed;top:0;right:-460px;width:420px;height:100vh;background:#fff;box-shadow:-4px 0 24px rgba(0,0,0,.12);display:flex;flex-direction:column;transition:right .28s cubic-bezier(.4,0,.2,1);z-index:1000;font-family:-apple-system,'PingFang SC','Microsoft YaHei',sans-serif}
.chat-panel.open{right:0}
.chat-panel-header{background:linear-gradient(135deg,#2F5496,#1a3a6e);color:#fff;padding:14px 16px;display:flex;align-items:center;justify-content:space-between}
.chat-panel-header .title{font-size:15px;font-weight:600}
.chat-panel-header .title-sub{font-size:11px;opacity:.85;margin-top:2px}
.chat-close{background:none;border:none;color:#fff;font-size:24px;cursor:pointer;opacity:.8;width:28px;height:28px;line-height:1;padding:0}
.chat-close:hover{opacity:1}
.chat-body{flex:1;overflow-y:auto;padding:16px;background:#f7f9fc}
.chat-welcome{padding:20px 8px;font-size:13px;color:#666;line-height:1.6;text-align:left}
.chat-welcome strong{color:#2F5496;font-size:14px;display:block;margin-bottom:4px}
.chat-msg{margin-bottom:14px;max-width:88%;word-wrap:break-word}
.chat-msg.user{margin-left:auto}
.chat-msg.assistant{margin-right:auto}
.chat-bubble{padding:10px 13px;border-radius:12px;font-size:13px;line-height:1.55;white-space:pre-wrap}
.chat-msg.user .chat-bubble{background:#2F5496;color:#fff;border-bottom-right-radius:3px}
.chat-msg.assistant .chat-bubble{background:#fff;color:#333;border:1px solid #e8ecf1;border-bottom-left-radius:3px}
.chat-bubble strong{font-weight:600}
.chat-bubble table{border-collapse:collapse;margin:6px 0;font-size:12px}
.chat-bubble td,.chat-bubble th{border:1px solid #d8dee8;padding:4px 8px;text-align:left}
.chat-bubble th{background:#eef4fc}
.chat-meta{font-size:10px;color:#999;margin-top:4px;padding:0 4px}
.chat-tool-badge{display:inline-block;background:#eef4fc;color:#4472C4;border:1px solid #d0dcef;border-radius:4px;padding:1px 6px;font-size:10px;margin:2px 4px 2px 0;font-family:monospace}
.chat-feedback{display:inline-flex;gap:6px;margin-top:6px;margin-left:2px}
.chat-feedback button{background:none;border:1px solid #e0e4eb;border-radius:6px;padding:3px 9px;cursor:pointer;font-size:13px;color:#888}
.chat-feedback button:hover{background:#f0f4fa;color:#2F5496}
.chat-feedback button.selected-up{background:#e8f5e9;color:#27ae60;border-color:#a8dbb1}
.chat-feedback button.selected-down{background:#fce4ec;color:#e74c3c;border-color:#f5b7c5}
.chat-input-bar{border-top:1px solid #e8ecf1;padding:12px;background:#fff;display:flex;gap:8px;align-items:flex-end}
.chat-input{flex:1;border:1px solid #d8dee8;border-radius:20px;padding:9px 14px;font-size:13px;font-family:inherit;resize:none;outline:none;max-height:80px;line-height:1.4}
.chat-input:focus{border-color:#2F5496}
.chat-send{background:#2F5496;color:#fff;border:none;width:36px;height:36px;border-radius:50%;cursor:pointer;font-size:16px;display:flex;align-items:center;justify-content:center;flex-shrink:0}
.chat-send:hover{background:#1a3a6e}
.chat-send:disabled{background:#c0c4cc;cursor:not-allowed}
.chat-thinking{display:inline-flex;gap:4px;align-items:center;padding:4px 0}
.chat-thinking span{width:6px;height:6px;background:#2F5496;border-radius:50%;animation:chatBounce 1.2s infinite ease-in-out}
.chat-thinking span:nth-child(2){animation-delay:.2s}
.chat-thinking span:nth-child(3){animation-delay:.4s}
@keyframes chatBounce{0%,80%,100%{transform:scale(.6);opacity:.5}40%{transform:scale(1);opacity:1}}
@media(max-width:600px){.chat-panel{width:100%;right:-100%}.chat-panel.open{right:0}.chat-btn-robot{top:10px;right:52px;width:40px;height:40px}.chat-btn-robot svg{width:24px;height:24px}}"""

HOME_HTML = """<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>ENERXON Tracker</title><style>""" + COMMON_CSS + """
.proj-list{padding:16px;display:grid;gap:12px}
.proj-card{background:#fff;border-radius:12px;padding:20px;box-shadow:0 2px 8px rgba(0,0,0,.08);cursor:pointer;transition:transform .15s}
.proj-card:hover{transform:translateY(-2px);box-shadow:0 4px 12px rgba(0,0,0,.12)}
.proj-card .id{font-size:20px;font-weight:700;color:#2F5496}.proj-card .count{font-size:13px;color:#888;margin-top:4px}
.proj-card .stats{display:flex;gap:16px;margin-top:12px;font-size:12px}
.proj-card .stats span{padding:3px 8px;border-radius:4px}
.proj-card .done{background:#e8f5e9;color:#27ae60}.proj-card .wip{background:#fff3e0;color:#f39c12}.proj-card .todo{background:#fce4ec;color:#e74c3c}
.empty{text-align:center;padding:60px 20px;color:#888;font-size:16px}
</style></head><body>
<div class="header"><div style="display:flex;align-items:center;gap:12px"><div><h1 style="margin:0">Production Tracker</h1><div class="sub" style="margin:0">\u751f\u4ea7\u8fdb\u5ea6\u8ffd\u8e2a\u7cfb\u7edf \u2014 Select a project / \u9009\u62e9\u9879\u76ee</div></div></div></div>
<div class="proj-list" id="projects"><div class="empty">Loading projects... / \u52a0\u8f7d\u4e2d...</div></div>
<script>
async function load(){
  const r = await fetch('/api/projects'); const projects = await r.json();
  if(!projects.length){ document.getElementById('projects').innerHTML='<div class="empty">No projects yet / \u6682\u65e0\u9879\u76ee<br><br>Use the API to import spools:<br><code>POST /api/import</code></div>'; return; }
  document.getElementById('projects').innerHTML = projects.map(p => `
    <div class="proj-card" onclick="location.href='/project/${p.project}'">
      <div class="id">${p.project}</div>
      <div class="count">${p.spool_count} spools</div>
      <div class="pbar-bg" style="margin-top:10px"><div class="pbar-fill" style="width:${p.overall_pct}%;background:${p.overall_pct>=100?'#27ae60':'#2F5496'}"></div></div>
      <div style="font-size:14px;font-weight:700;margin-top:6px;color:#2F5496">${p.overall_pct}%</div>
      <div class="stats">
        <span class="done">${p.completed} done</span>
        <span class="wip">${p.in_progress} in progress</span>
        <span class="todo">${p.not_started} pending</span>
      </div>
    </div>`).join('');
}
load(); setInterval(load, 30000);
</script></body></html>"""

# PROJECT_HTML, SPOOL_HTML, REPORT_HTML — preserved from original with dynamic hold/release
# These templates are loaded from the API which now returns is_hold_point/is_release flags

PROJECT_HTML = """<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{{ project }} — Tracker</title><style>""" + COMMON_CSS + """
.section{padding:0 16px 16px}.section h2{font-size:16px;color:#2F5496;margin:16px 0 8px}
.toolbar{display:flex;gap:8px;padding:8px 16px;flex-wrap:wrap}
.diam-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(130px,1fr));gap:10px;padding:0 16px}
.diam-card{background:#fff;border-radius:10px;padding:14px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.08);position:relative;border-top:3px solid #e8e8e8}
.diam-card .d{font-size:22px;font-weight:700;color:#2F5496}.diam-card .p{font-size:11px;color:#888}
.d-ahead{border-top-color:#27ae60}.d-ontrack{border-top-color:#4472C4}.d-atrisk{border-top-color:#f39c12}.d-behind{border-top-color:#e74c3c}.d-pending{border-top-color:#e8e8e8}
.pace-badge{position:absolute;top:6px;right:6px;font-size:7px;font-weight:700;padding:1px 5px;border-radius:4px;text-transform:uppercase;letter-spacing:.3px}
.pb-ahead{background:#e8f5e9;color:#27ae60}.pb-ontrack{background:#e3f2fd;color:#4472C4}.pb-atrisk{background:#fff3e0;color:#f39c12}.pb-behind{background:#fce4ec;color:#e74c3c}.pb-pending{background:#f5f5f5;color:#bbb}
.spool-row{display:flex;align-items:center;gap:10px;padding:10px 16px;background:#fff;margin:4px 16px;border-radius:8px;cursor:pointer;box-shadow:0 1px 2px rgba(0,0,0,.05)}
.spool-row:hover{background:#f8f9fb}.spool-row .info{flex:1}.spool-row .name{font-weight:600;font-size:14px;color:#2F5496}
.spool-row .meta{font-size:11px;color:#888}.spool-row .bar{width:80px}.spool-row .pct{font-size:14px;font-weight:700;min-width:45px;text-align:right}
.activity-item{padding:8px 0;border-bottom:1px solid #f0f2f5;font-size:12px}
.filter-row{display:flex;gap:8px;padding:8px 16px;flex-wrap:wrap}
.filter-row select,.filter-row input{padding:6px 10px;border:1px solid #ddd;border-radius:6px;font-size:12px}
.target-banner{display:flex;align-items:center;gap:14px;padding:10px 18px;margin:8px 16px;border-radius:10px;background:#fff;box-shadow:0 2px 8px rgba(0,0,0,.08);flex-wrap:wrap;border-left:4px solid #2F5496}
.tb-section{text-align:center;padding:0 12px;border-right:1px solid #eee;min-width:80px}.tb-section:last-child{border-right:none}
.tb-label{font-size:8px;color:#888;text-transform:uppercase;letter-spacing:.3px}.tb-value{font-size:18px;font-weight:700}.tb-sub{font-size:8px;color:#aaa}
.status-pill{display:inline-block;padding:4px 10px;border-radius:12px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.3px}
.sp-green{background:#e8f5e9;color:#27ae60}.sp-red{background:#fce4ec;color:#e74c3c}
.rate-strip{display:flex;align-items:center;gap:14px;padding:10px 18px;margin:0 16px 8px;border-radius:10px;background:#fff;box-shadow:0 2px 8px rgba(0,0,0,.08);flex-wrap:wrap;border-left:4px solid #4472C4}
.rs-title{font-size:11px;font-weight:700;color:#2F5496;white-space:nowrap}
.rs-item{display:flex;align-items:center;gap:6px;padding:0 10px;border-right:1px solid #eee}.rs-item:last-of-type{border-right:none}
.rs-lbl{font-size:8px;color:#888;text-transform:uppercase}.rs-val{font-size:20px;font-weight:700}
.rs-badge{padding:3px 10px;border-radius:8px;font-size:9px;font-weight:700;margin-left:auto;white-space:nowrap}
.rs-good{background:#e8f5e9;color:#27ae60}.rs-bad{background:#fce4ec;color:#e74c3c}
</style></head><body>
<div class="header">
  <div style="display:flex;align-items:center;gap:12px">
    <a class="back" href="/">\u2190 Home</a>
    <div style="flex:1"><h1>{{ project }}</h1><div class="sub" id="subtitle">Loading... / \u52a0\u8f7d\u4e2d...</div></div>
    <div onclick="toggleSettings()" style="cursor:pointer;font-size:20px;color:#aaa;padding:8px" title="Settings / \u8bbe\u7f6e">\u2699</div>
  </div>
</div>
<div id="settings-panel" style="display:none;background:#fff;margin:0 16px 8px;border-radius:10px;padding:16px;box-shadow:0 2px 8px rgba(0,0,0,.08);border-left:4px solid #888">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
    <h3 style="font-size:14px;color:#333;margin:0">Settings / \u8bbe\u7f6e</h3>
    <div onclick="toggleSettings()" style="cursor:pointer;font-size:18px;color:#aaa">\u2715</div>
  </div>
  <div id="settings-fields"></div>
  <div style="margin-top:16px;border-top:1px solid #eee;padding-top:12px">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
      <h4 style="font-size:12px;color:#333;margin:0">Shipments / \u53d1\u8fd0</h4>
      <button onclick="addShipmentSetting()" style="padding:3px 10px;font-size:10px;border:1px solid #2F5496;border-radius:4px;background:#fff;color:#2F5496;cursor:pointer">+ Add / \u65b0\u589e</button>
    </div>
    <div id="settings-shipments"></div>
  </div>
</div>
<div class="toolbar">
  <a class="btn" href="/api/project/{{ project }}/export">\U0001f4e5 Export Excel</a>
  <a class="btn" href="/project/{{ project }}/report" style="background:#27ae60">\U0001f4ca Report / \u62a5\u544a</a>
  <a class="btn" href="/api/project/{{ project }}/report/download" style="background:#ED7D31">\U0001f4ca Report Excel</a>
  <a class="btn" href="/api/project/{{ project }}/report/pdf" style="background:#E74C3C">\U0001f4c4 Report PDF</a>
</div>
<div id="target-banner"></div>
<div id="rate-strip"></div>
<div class="stats-grid" id="stats"></div>
<div id="diam" class="diam-grid"></div>
<div class="filter-row">
  <input id="q" placeholder="Search spool / \u641c\u7d22..." oninput="filter()">
  <select id="fd" onchange="filter()"><option value="">All Diameters</option></select>
  <select id="fl" onchange="filter()"><option value="">All Lines</option><option value="A">A</option><option value="B">B</option><option value="C">C</option></select>
  <select id="fs" onchange="filter()"><option value="">All Status</option><option value="done">Done / \u5b8c\u6210</option><option value="wip">WIP / \u8fdb\u884c\u4e2d</option><option value="todo">Pending / \u5f85\u5f00\u59cb</option></select>
</div>
<div id="list"></div>
<div id="act" class="section"></div>
<script>
const P='{{project}}';
let all=[];
async function load(){
  const [dr, sr] = await Promise.all([fetch(`/api/project/${P}/dashboard`), fetch(`/api/project/${P}/spools`)]);
  const st = await dr.json(); all = await sr.json();
  const fc = st.forecast||{}, sett = st.settings||{}, schd = st.schedule_data, pr = st.production_rate||{};
  const phases = st.phase_order || schd?.phase_order || ['fab'];
  const phaseColors = ['#4472C4', '#ED7D31', '#8E44AD', '#27AE60'];
  const toLocal = s => {const [y,m,d]=s.split('-');return new Date(+y,+m-1,+d);};
  const addDays = (dt,n) => new Date(dt.getFullYear(),dt.getMonth(),dt.getDate()+n);
  const fmt = d => d.toLocaleDateString('en',{day:'2-digit',month:'short'});
  document.getElementById('subtitle').textContent=`${st.total} spools \u00b7 ${st.overall_pct}% \u00b7 ${st.completed} done / \u5b8c\u6210 ${st.completed}`;
  document.getElementById('stats').innerHTML=[
    {v:st.total,l:'Total / \u603b\u6570'},{v:st.completed,l:'Done / \u5b8c\u6210',c:'#27ae60'},
    {v:st.in_progress,l:'WIP / \u8fdb\u884c\u4e2d',c:'#f39c12'},{v:st.not_started,l:'Pending / \u5f85\u5f00\u59cb',c:'#e74c3c'},
    {v:st.past_rt||0,l:'Past RT / \u8fc7RT',c:'#4472C4'}
  ].map(x=>`<div class="stat-card"><div class="value" style="color:${x.c||'#2F5496'}">${x.v}</div><div class="label">${x.l}</div></div>`).join('');

  // Expediting banner
  const stdWeeks = parseInt(sett.standard_weeks||'9');
  const wksSaved = parseInt(sett.committed_weeks_saved||'0');
  const daysSaved = parseInt(sett.committed_days_saved||'0');
  const totalSaved = wksSaved*7+daysSaved;
  const hasExpediting = totalSaved > 0;
  if(hasExpediting && schd && schd.diameters && schd.diameters.length){
    const starts = schd.diameters.map(x=>x.total_start).filter(x=>x).sort();
    if(starts.length){
      const psDate = toLocal(starts[0]);
      const stdEnd = addDays(psDate, stdWeeks*7 - 1);
      const commitEnd = addDays(stdEnd, -totalSaved);
      const fcEnd = fc.overall_forecast_end ? toLocal(fc.overall_forecast_end) : null;
      const today = new Date(); today.setHours(0,0,0,0);
      const daysToTarget = Math.round((commitEnd - today) / 86400000);
      const fcSaved = fcEnd ? Math.ceil((stdEnd - fcEnd) / 86400000) : 0;
      const fcDiff = fcEnd ? Math.ceil((commitEnd - fcEnd) / 86400000) : 0;
      const statusCls = fcDiff >= 0 ? 'tb-ok' : 'tb-warn';
      const pillCls = fcDiff >= 0 ? 'sp-green' : 'sp-red';
      const pillText = fcDiff >= 0 ? `\u2713 ${fcDiff}d ahead / \u8d85\u524d` : `\u2717 ${Math.abs(fcDiff)}d behind / \u843d\u540e`;
      document.getElementById('target-banner').innerHTML=`<div class="target-banner ${statusCls}">
        <div class="tb-section"><div class="tb-label">Expediting Target / \u52a0\u6025\u76ee\u6807</div><div class="tb-value" style="color:#4472C4">${fmt(commitEnd)}</div><div class="tb-sub">Committed / \u627f\u8bfa\u5b8c\u5de5</div></div>
        <div class="tb-section"><div class="tb-label">Forecast End / \u9884\u6d4b\u5b8c\u5de5</div><div class="tb-value" style="color:#27ae60">${fmt(fcEnd||new Date())}</div><div class="tb-sub">Based on rate / \u57fa\u4e8e\u8fdb\u5ea6</div></div>
        <div class="tb-section"><div class="tb-label">Days to Target / \u8ddd\u76ee\u6807</div><div class="tb-value" style="color:#2F5496">${daysToTarget}</div><div class="tb-sub">days left / \u5269\u4f59\u5929\u6570</div></div>
        <div class="tb-section"><div class="tb-label">Expediting Saved / \u52a0\u6025\u8282\u7701</div><div class="tb-value" style="color:#4472C4">${totalSaved}<span style="font-size:10px;font-weight:400"> days</span></div><div class="tb-sub">vs standard / \u8f83\u6807\u51c6</div></div>
        <div class="tb-section"><div class="tb-label">Forecast Saved / \u9884\u6d4b\u8282\u7701</div><div class="tb-value" style="color:#27ae60">${fcSaved}<span style="font-size:10px;font-weight:400"> days</span></div><div class="tb-sub">predicted / \u9884\u6d4b</div></div>
        <div class="tb-section" style="border-right:none;margin-left:auto"><div class="status-pill ${pillCls}">${pillText}</div><div style="font-size:8px;color:#888;margin-top:2px">vs commitment / \u8f83\u627f\u8bfa</div></div>
      </div>`;
      // Daily rate + bottleneck
      const avgRate = pr.avg_7day || 0;
      const trend = pr.trend || 0;
      const todaySteps = pr.today_steps || 0;
      const remaining = st.total - st.completed;
      const targetRate = daysToTarget > 0 ? (remaining / daysToTarget).toFixed(1) : '\u2014';
      const aboveTarget = avgRate >= parseFloat(targetRate);
      const trendArrow = trend >= 0 ? `<span style="color:#27ae60;font-size:10px;font-weight:600">\u25b2${trend}</span>` : `<span style="color:#e74c3c;font-size:10px;font-weight:600">\u25bc${Math.abs(trend)}</span>`;
      let bottleneck = null;
      if(schd && schd.diameters){
        let worstRatio = 999;
        schd.diameters.forEach(dm => {
          if(dm.actual_pct >= 100 || dm.status === 'not_started') return;
          const ratio = dm.actual_pct / Math.max(dm.expected_pct, 1);
          if(ratio < worstRatio){ worstRatio = ratio; bottleneck = dm; }
        });
      }
      let bnHtml = '';
      if(bottleneck && bottleneck.actual_pct < 100){
        const fcDiam = fc.diameters ? fc.diameters[bottleneck.diameter] : null;
        const fcEndDiam = fcDiam ? fcDiam.forecast_end : '\u2014';
        const neededRate = daysToTarget > 0 ? ((100 - bottleneck.actual_pct) / 100 * bottleneck.spool_count / daysToTarget).toFixed(1) : '\u2014';
        bnHtml = `<div style="background:#fff;border-radius:8px;padding:10px 16px;box-shadow:0 1px 2px rgba(0,0,0,.05);min-width:240px;border-left:4px solid #e74c3c;display:flex;flex-direction:column;justify-content:center">
          <div style="font-size:9px;color:#888;text-transform:uppercase;letter-spacing:.3px">Critical Path / \u5173\u952e\u8def\u5f84</div>
          <div style="display:flex;align-items:baseline;gap:5px;margin:2px 0"><span style="font-size:20px;font-weight:700;color:#e74c3c">${bottleneck.diameter}</span><span style="font-size:11px;color:#888">${bottleneck.spool_count} spools \u00b7 slowest / \u6700\u6162</span></div>
          <div style="font-size:10px;color:#666">Need / \u9700\u8981 <strong>${neededRate} spools/day / \u6bcf\u65e5</strong> to hit target / \u8fbe\u5230\u76ee\u6807</div>
          <div style="font-size:9px;color:#999;margin-top:2px">${bottleneck.actual_pct}% done \u00b7 Forecast / \u9884\u6d4b: ${fcEndDiam}</div>
        </div>`;
      }
      document.getElementById('rate-strip').innerHTML = `<div style="display:flex;gap:10px;margin:0 16px 8px;flex-wrap:wrap">
        <div class="rate-strip" style="flex:1;min-width:400px;margin:0">
          <div class="rs-title">\U0001f4ca Daily Rate / \u6bcf\u65e5\u751f\u4ea7\u7387</div>
          <div class="rs-item"><div><div class="rs-lbl">Target / \u76ee\u6807</div><div class="rs-val" style="color:#2F5496">${targetRate}</div></div><div class="rs-lbl">spools/day<br>\u6bcf\u65e5\u9700\u5b8c\u6210</div></div>
          <div class="rs-item"><div><div class="rs-lbl">7-day avg / 7\u5929\u5747\u503c</div><div style="display:flex;align-items:baseline;gap:3px"><div class="rs-val" style="color:${aboveTarget?'#27ae60':'#e74c3c'}">${avgRate}</div>${trendArrow}</div></div><div class="rs-lbl">spools/day<br>\u8f83\u4e0a\u5468</div></div>
          <div class="rs-item" style="border-right:none"><div><div class="rs-lbl">Steps today / \u4eca\u65e5\u6b65\u9aa4</div><div class="rs-val" style="color:#4472C4">${todaySteps}</div></div><div class="rs-lbl">completed<br>\u4eca\u65e5\u5b8c\u6210</div></div>
          <div class="rs-badge ${aboveTarget?'rs-good':'rs-bad'}">${aboveTarget?'\u2713 Above target / \u8d85\u8fc7\u76ee\u6807':'\u26a0 Below target / \u4f4e\u4e8e\u76ee\u6807'}</div>
        </div>
        ${bnHtml}
      </div>`;
    }
  }

  // Diameter cards
  const ds=Object.entries(st.by_diameter).sort((a,b)=>(parseInt(b[0])||0)-(parseInt(a[0])||0));
  document.getElementById('diam').innerHTML=ds.map(([d,v])=>{
    let cls='d-pending',badge='PENDING / \u5f85\u5f00\u59cb',badgeCls='pb-pending',paceText='',barColor='#ccc';
    if(schd && schd.diameters){
      const dm = schd.diameters.find(x=>x.diameter===d);
      if(dm){
        const diff = dm.actual_pct - dm.expected_pct;
        if(dm.status==='not_started'){cls='d-pending';badge='PENDING / \u5f85\u5f00\u59cb';badgeCls='pb-pending';barColor='#ccc';}
        else if(diff >= 5){cls='d-ahead';badge='AHEAD / \u8d85\u524d';badgeCls='pb-ahead';barColor='#27ae60';paceText=`+${Math.round(diff)}% ahead / \u8d85\u524d`;}
        else if(diff >= -5){cls='d-ontrack';badge='ON TRACK / \u8fbe\u6807';badgeCls='pb-ontrack';barColor='#4472C4';paceText=`${diff>=0?'+':''}${Math.round(diff)}% on pace / \u8fbe\u6807`;}
        else if(diff >= -15){cls='d-atrisk';badge='AT RISK / \u6709\u98ce\u9669';badgeCls='pb-atrisk';barColor='#f39c12';paceText=`${Math.round(diff)}% behind / \u843d\u540e`;}
        else{cls='d-behind';badge='BEHIND / \u843d\u540e';badgeCls='pb-behind';barColor='#e74c3c';paceText=`${Math.round(diff)}% behind / \u843d\u540e`;}
        if(dm.expected_pct===0 && dm.actual_pct>0){cls='d-ahead';badge='AHEAD / \u8d85\u524d';badgeCls='pb-ahead';barColor='#27ae60';paceText='Started early / \u63d0\u524d\u5f00\u59cb';}
      } else if(v.avg_pct > 0){cls='d-ontrack';badge='WIP';badgeCls='pb-ontrack';barColor='#4472C4';}
    } else if(v.avg_pct > 0){cls='d-ontrack';barColor='#4472C4';}
    else if(v.avg_pct >= 100){cls='d-ahead';barColor='#27ae60';}
    const phAvgs = v.phase_avgs || {};
    const phaseBars = phases.map((ph,pi) => {
      const pv = phAvgs[ph]||0;
      return `<div style="margin-top:${pi===0?6:2}px"><div style="font-size:8px;color:#aaa">${ph.charAt(0).toUpperCase()+ph.slice(1)}</div><div class="pbar-bg" style="height:8px"><div class="pbar-fill" style="width:${pv}%;background:${phaseColors[pi%phaseColors.length]};height:100%"></div></div></div>`;
    }).join('');
    return `<div class="diam-card ${cls}"><div class="pace-badge ${badgeCls}">${badge}</div><div class="d">${d}</div><div class="p">${v.total} spools</div>
      ${phaseBars}
      <div style="font-size:12px;margin-top:4px;font-weight:600;color:${barColor}">${v.avg_pct}%</div>
      ${paceText?`<div style="font-size:9px;color:#888;margin-top:2px">${paceText}</div>`:''}</div>`;
  }).join('');
  render(all);
  const diamsSet = [...new Set(all.map(s=>s.spool.main_diameter||'?'))].sort((a,b)=>(parseInt(b)||0)-(parseInt(a)||0));
  const fdEl = document.getElementById('fd');
  const curFd = fdEl.value;
  fdEl.innerHTML = '<option value="">All Diameters</option>';
  diamsSet.forEach(d=>{ const o=document.createElement('option'); o.value=d; o.textContent=d; fdEl.appendChild(o); });
  fdEl.value = curFd;
  if(st.recent_activity&&st.recent_activity.length)
    document.getElementById('act').innerHTML='<h2 style="font-size:16px;color:#2F5496;margin:8px 0">Recent / \u6700\u8fd1\u52a8\u6001</h2>'+
      st.recent_activity.slice(0,10).map(a=>`<div class="activity-item"><strong>${a.spool_id}</strong> \u2014 ${a.details||a.action} <span style="color:#aaa;font-size:11px">${a.timestamp||''}</span></div>`).join('');
}
function render(sp){
  document.getElementById('list').innerHTML=sp.map(s=>{
    const p=s.progress_pct,c=p>=100?'pct-green':p>0?'pct-yellow':'pct-red',bg=p>=100?'#27ae60':p>0?'#f39c12':'#e8e8e8',l=s.spool.line||'?';
    return`<div class="spool-row" onclick="location.href='/project/${P}/spool/${s.spool.spool_id}'">
      <span class="line-badge line-${l}">${l}</span>
      <div class="info"><div class="name">${s.spool.spool_id}</div><div class="meta">${s.spool.main_diameter||''} \u00b7 ${s.spool.iso_no||''}</div></div>
      <div class="bar"><div class="pbar-bg"><div class="pbar-fill" style="width:${p}%;background:${bg}"></div></div></div>
      <div class="pct ${c}">${p}%</div></div>`;}).join('');
}
function filter(){
  const q=document.getElementById('q').value.toLowerCase(),d=document.getElementById('fd').value,l=document.getElementById('fl').value,s=document.getElementById('fs').value;
  render(all.filter(x=>{
    if(q&&!x.spool.spool_id.toLowerCase().includes(q)&&!(x.spool.iso_no||'').toLowerCase().includes(q))return false;
    if(d&&(x.spool.main_diameter||'?')!==d)return false;
    if(l&&x.spool.line!==l)return false;
    if(s==='done'&&x.progress_pct<100)return false;
    if(s==='wip'&&(x.progress_pct<=0||x.progress_pct>=100))return false;
    if(s==='todo'&&x.progress_pct>0)return false;
    return true;
  }));
}
// ── Settings Panel ──
const SETTINGS_SCHEMA = [
  {key:'standard_weeks',        label_en:'Standard Delivery',label_cn:'标准交货',unit:'weeks / 周',type:'number',def:'9'},
  {key:'committed_weeks_saved', label_en:'Weeks Expedited',  label_cn:'加急节省周数',unit:'weeks / 周',type:'number',def:'0'},
  {key:'production_start',     label_en:'Production Start',  label_cn:'生产开始日期',unit:'',type:'date',def:''},
  {key:'welding_capability_ipd',label_en:'Welding Capacity', label_cn:'焊接产能',unit:'inch-dia/day',type:'number',def:'552'},
  {key:'surface_capability_m2d',label_en:'Surface Capacity', label_cn:'涂装产能',unit:'m²/day',type:'number',def:'91'},
];
let settingsOpen = false;
function toggleSettings(){
  settingsOpen = !settingsOpen;
  document.getElementById('settings-panel').style.display = settingsOpen ? 'block' : 'none';
  if(settingsOpen) loadSettings();
}
async function loadSettings(){
  const r = await fetch(`/api/project/${P}/dashboard`);
  const d = await r.json();
  const sett = d.settings || {};
  let h = '<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">';
  SETTINGS_SCHEMA.forEach(s => {
    const val = sett[s.key] || s.def;
    h += `<div style="display:flex;align-items:center;gap:6px;padding:4px 0">
      <div style="flex:1"><div style="font-size:11px;font-weight:600;color:#333">${s.label_en}</div><div style="font-size:9px;color:#888">${s.label_cn}</div></div>
      <input type="${s.type}" value="${val}" style="width:70px;padding:4px 6px;border:1px solid #ddd;border-radius:4px;font-size:12px;text-align:center"
        onchange="saveSetting('${s.key}',this.value)">
      <span style="font-size:9px;color:#aaa;min-width:50px">${s.unit}</span>
    </div>`;
  });
  h += '</div>';
  document.getElementById('settings-fields').innerHTML = h;
  const sr = await fetch(`/api/project/${P}/shipments`);
  const ships = await sr.json();
  let sh = '';
  if(!ships.length) sh = '<div style="text-align:center;padding:8px;color:#aaa;font-size:11px">No shipments / 未配置发运</div>';
  else {
    sh = `<table style="width:100%;border-collapse:collapse;font-size:11px"><tr style="background:#f0f2f5">
      <th style="padding:4px;border:1px solid #ddd">#</th>
      <th style="padding:4px;border:1px solid #ddd">Description / 描述</th>
      <th style="padding:4px;border:1px solid #ddd">ETD / 离港</th>
      <th style="padding:4px;border:1px solid #ddd">Transit / 运输</th>
      <th style="padding:4px;border:1px solid #ddd">ETA / 到达</th>
      <th style="padding:4px;border:1px solid #ddd"></th></tr>`;
    ships.forEach(s => {
      sh += `<tr data-ship="${s.shipment_number}">
        <td style="padding:3px;border:1px solid #ddd;font-weight:700;color:#2F5496;text-align:center">${s.shipment_number}</td>
        <td style="padding:3px;border:1px solid #ddd"><input value="${s.description||''}" style="width:100%;border:1px solid #eee;padding:2px 4px;font-size:11px;border-radius:3px" onchange="saveShipment(${s.shipment_number},{description:this.value})"></td>
        <td style="padding:3px;border:1px solid #ddd"><input type="date" value="${(s.etd||'').substring(0,10)}" style="border:1px solid #eee;padding:2px;font-size:10px;border-radius:3px" onchange="saveShipment(${s.shipment_number},{etd:this.value})"></td>
        <td style="padding:3px;border:1px solid #ddd"><input type="number" value="${s.transit_days||45}" style="width:50px;border:1px solid #eee;padding:2px;font-size:11px;text-align:center;border-radius:3px" onchange="saveShipment(${s.shipment_number},{transit_days:parseInt(this.value)})"></td>
        <td class="eta-cell" style="padding:3px;border:1px solid #ddd;font-weight:700;color:#003366;text-align:center">${s.eta||'—'}</td>
        <td style="padding:3px;border:1px solid #ddd;text-align:center"><span onclick="deleteShipmentS(${s.shipment_number})" style="color:#e74c3c;cursor:pointer;font-size:14px">×</span></td>
      </tr>`;
    });
    sh += '</table>';
  }
  document.getElementById('settings-shipments').innerHTML = sh;
}
async function saveSetting(key, val){
  await fetch(`/api/project/${P}/settings`, {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({[key]:val})});
  load();
}
async function addShipmentSetting(){
  const r = await fetch(`/api/project/${P}/shipments`);
  const ships = await r.json();
  const next = ships.length ? Math.max(...ships.map(s=>s.shipment_number)) + 1 : 1;
  await fetch(`/api/project/${P}/shipments`, {method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({shipment_number:next, description:'Shipment '+next, transit_days:45})});
  loadSettings(); // full reload needed after add
}
async function saveShipment(num, fields){
  fields.shipment_number = num;
  await fetch(`/api/project/${P}/shipments`, {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify(fields)});
  refreshShipmentsOnly();
}
async function refreshShipmentsOnly(){
  const sr = await fetch(`/api/project/${P}/shipments`);
  const ships = await sr.json();
  // Update only ETA cells without re-rendering inputs
  ships.forEach(s => {
    const row = document.querySelector(`[data-ship="${s.shipment_number}"]`);
    if(row) { const eta = row.querySelector('.eta-cell'); if(eta) eta.textContent = s.eta || '—'; }
  });
}
async function deleteShipmentS(num){
  if(!confirm('Delete shipment '+num+'? / 删除批次'+num+'?')) return;
  await fetch(`/api/project/${P}/shipments/${num}`, {method:'DELETE'});
  loadSettings(); // full reload needed after delete
}

load(); setInterval(load,30000);
</script></body></html>"""

SPOOL_HTML = """<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{{ spool_id }} — Checklist</title><style>""" + COMMON_CSS + """
.info-bar{background:#fff;padding:12px 16px;display:flex;flex-wrap:wrap;gap:16px;align-items:center;box-shadow:0 1px 2px rgba(0,0,0,.05)}
.info-item{font-size:12px}.info-item .lb{color:#888}.info-item .vl{font-weight:600}
.prog{text-align:center;padding:20px}.prog .big{font-size:48px;font-weight:700;color:#2F5496}.prog .lbl{font-size:13px;color:#888}
.op-input{padding:8px 16px}.op-input input{width:100%;padding:10px;border:1px solid #ddd;border-radius:8px;font-size:14px}
.checklist{padding:8px 16px 80px}
.step{background:#fff;border-radius:10px;margin-bottom:8px;overflow:hidden;box-shadow:0 1px 2px rgba(0,0,0,.05)}
.step.done{border-left:4px solid #27ae60}.step.pending{border-left:4px solid #e8e8e8}
.step-h{display:flex;align-items:center;padding:14px 16px;gap:12px;cursor:pointer}
.step-h .num{width:28px;height:28px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;flex-shrink:0}
.step.done .num{background:#27ae60;color:#fff}.step.pending .num{background:#e8e8e8;color:#888}
.step-h .text{flex:1}.step-h .en{font-size:13px;font-weight:600}.step-h .cn{font-size:11px;color:#888}
.step-h .chk{font-size:24px}
.step-meta{padding:0 16px 8px 56px;font-size:11px;color:#888}
.step-rem{padding:4px 16px 12px 56px}
.step-rem input{width:100%;padding:6px 10px;border:1px solid #ddd;border-radius:6px;font-size:12px}
.wbadge{font-size:10px;background:#f0f2f5;padding:2px 6px;border-radius:4px;color:#888}
.tab-bar{display:flex;padding:0 16px;gap:4px;background:#fff;box-shadow:0 1px 2px rgba(0,0,0,.05)}
.tab-btn{flex:1;padding:12px;text-align:center;font-size:13px;font-weight:600;cursor:pointer;border-bottom:3px solid transparent;color:#888;transition:all .2s}
.tab-btn.active{color:#2F5496;border-bottom-color:#2F5496}
.tab-content{display:none}.tab-content.active{display:block}
.qc-card{background:#fff;border-radius:10px;margin-bottom:8px;padding:14px 16px;box-shadow:0 1px 2px rgba(0,0,0,.05);display:flex;align-items:center;gap:12px;cursor:pointer;text-decoration:none;color:inherit;border-left:4px solid #e8e8e8}
.qc-card:hover{background:#f8f9fa}
.qc-card.qc-draft{border-left-color:#f39c12}.qc-card.qc-submitted{border-left-color:#2F5496}.qc-card.qc-approved{border-left-color:#27ae60}
.qc-icon{font-size:24px;width:36px;text-align:center;flex-shrink:0}
.qc-info{flex:1}.qc-info .en{font-size:13px;font-weight:600}.qc-info .cn{font-size:11px;color:#888}
.qc-info .sub-label{font-size:10px;color:#2F5496;font-weight:600}
.qc-status{font-size:10px;padding:3px 8px;border-radius:10px;font-weight:600;white-space:nowrap}
.qs-not_started{background:#f0f2f5;color:#888}.qs-draft{background:#FEF3CD;color:#856404}.qs-submitted{background:#D6E4F0;color:#2F5496}.qs-approved{background:#D5F5D5;color:#1B7A1B}
.qc-cat-header{font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;padding:12px 16px 4px;color:#888}
</style></head><body>
<div class="header">
  <a class="back" href="/project/{{ project }}">\u2190 {{ project }}</a>
  <h1>{{ spool_id }}</h1>
  <div class="sub" id="sub-info">Loading... / 加载中...</div>
</div>
<div class="info-bar" id="info-bar"></div>
<div id="ship-assign-wrap" style="padding:8px 16px;display:none">
  <div style="background:#fff;border-radius:10px;padding:10px 14px;box-shadow:0 1px 2px rgba(0,0,0,.05);display:flex;align-items:center;gap:10px">
    <span style="font-size:16px">\U0001f6a2</span>
    <div style="flex:1">
      <div style="font-size:11px;color:#888">Shipment / \u53d1\u8fd0</div>
      <select id="ship-select" onchange="assignShipment(this.value)" style="width:100%;padding:6px;border:1px solid #ddd;border-radius:6px;font-size:13px;margin-top:2px">
        <option value="">\u2014 Unassigned / \u672a\u5206\u914d \u2014</option>
      </select>
    </div>
  </div>
</div>
<div style="padding:8px 16px;display:none" id="dwg-wrap">
  <a class="btn" id="dwg-btn" target="_blank" style="width:100%;text-align:center;display:block;padding:10px;font-size:14px">\U0001f4c4 View Drawing / \u67e5\u770b\u56fe\u7eb8</a>
</div>
<div class="prog"><div class="big" id="prog-pct">--</div><div class="lbl">Progress / \u8fdb\u5ea6</div>
  <div class="pbar-bg" style="max-width:300px;margin:8px auto"><div class="pbar-fill" id="prog-bar" style="width:0%;background:#2F5496"></div></div>
</div>
<div class="op-input"><input id="operator" placeholder="Operator name / \u64cd\u4f5c\u5458\u59d3\u540d" value=""></div>
<div class="tab-bar">
  <div class="tab-btn active" onclick="switchTab('checklist')">Checklist / \u68c0\u67e5\u8868</div>
  <div class="tab-btn" onclick="switchTab('qc')">QC Reports / \u8d28\u68c0\u62a5\u544a</div>
</div>
<div class="tab-content active" id="tab-checklist">
  <div class="checklist" id="steps"></div>
</div>
<div class="tab-content" id="tab-qc">
  <div class="checklist" id="qc-reports"></div>
</div>
<script>
const P='{{project}}',S='{{spool_id}}';
let stepDefs=[], stepMap={};
async function load(){
  const r = await fetch(`/api/project/${P}/spool/${S}`);
  const d = await r.json();
  if(d.error){document.getElementById('steps').innerHTML=`<p style="color:red">${d.error}</p>`;return;}
  const sp=d.spool;
  stepDefs = d.step_definitions || [];
  document.getElementById('sub-info').textContent=`${sp.main_diameter||''} \u00b7 ${sp.iso_no||''} \u00b7 ${sp.marking||''}`;
  document.getElementById('info-bar').innerHTML=[
    ['Diameter / \u53e3\u5f84',sp.main_diameter],['ISO',sp.iso_no],['Line / \u7ebf',sp.line],['MK',sp.mk_number],['Marking / \u6807\u8bc6',sp.marking]
  ].filter(x=>x[1]).map(([l,v])=>`<div class="info-item"><span class="lb">${l}:</span> <span class="vl">${v}</span></div>`).join('');
  document.getElementById('prog-pct').textContent=d.progress_pct+'%';
  document.getElementById('prog-bar').style.width=d.progress_pct+'%';
  document.getElementById('prog-bar').style.background=d.progress_pct>=100?'#27ae60':d.progress_pct>0?'#2F5496':'#e8e8e8';
  stepMap={};
  (d.steps||[]).forEach(s=>stepMap[s.step_number]=s);
  // Build hold/release lookup from step_definitions
  const holdSteps = new Set(stepDefs.filter(s=>s.is_hold_point).map(s=>s.number));
  const releaseSteps = new Set(stepDefs.filter(s=>s.is_release).map(s=>s.number));
  document.getElementById('steps').innerHTML=stepDefs.map(sd=>{
    const st=stepMap[sd.number]||{completed:0};
    const done=st.completed;
    const badges=[];
    if(holdSteps.has(sd.number)) badges.push('<span class="wbadge" style="background:#EDE7F6;color:#5E35B1">\u2605 HOLD POINT</span>');
    if(releaseSteps.has(sd.number)) badges.push('<span class="wbadge" style="background:#E8F5E9;color:#2E7D32">\U0001f3c1 WITNESS</span>');
    return `<div class="step ${done?'done':'pending'}" id="step-${sd.number}">
      <div class="step-h" onclick="toggle(${sd.number})">
        <div class="num">${sd.number}</div>
        <div class="text"><div class="en">${sd.name_en} ${badges.join(' ')}</div><div class="cn">${sd.name_cn||''}</div></div>
        <div class="chk">${done?'\u2705':'\u2b1c'}</div>
      </div>
      ${done?`<div class="step-meta">By ${st.completed_by||'?'} \u00b7 ${(st.completed_at||'').substring(0,16)}</div>`:''}
      <div class="step-rem"><input placeholder="Remarks / \u5907\u6ce8" value="${st.remarks||''}" onchange="remark(${sd.number},this.value)"></div>
    </div>`;
  }).join('');
  // Drawing button
  try{
    const dr=await fetch(`/api/project/${P}/drawings/list`);
    const dl=await dr.json();
    if(dl.find(x=>x.spool_id===S)){
      document.getElementById('dwg-wrap').style.display='block';
      document.getElementById('dwg-btn').href=`/api/project/${P}/spool/${S}/drawing`;
    }
  }catch(e){}
  // Shipment assignment dropdown (shown only if project has shipments configured)
  try{
    const shr = await fetch(`/api/project/${P}/shipments`);
    const ships = await shr.json();
    if(ships && ships.length){
      const sel = document.getElementById('ship-select');
      const current = sp.shipment_number;
      ships.forEach(s => {
        const o = document.createElement('option');
        o.value = s.shipment_number;
        o.textContent = `SH-${String(s.shipment_number).padStart(3,'0')}${s.description?' \u00b7 '+s.description:''}${s.etd?' \u00b7 ETD '+String(s.etd).substring(0,10):''}`;
        if(current == s.shipment_number) o.selected = true;
        sel.appendChild(o);
      });
      document.getElementById('ship-assign-wrap').style.display = 'block';
    }
  }catch(e){}
}
async function assignShipment(v){
  const body = {shipment_number: v === '' ? null : parseInt(v)};
  await fetch(`/api/project/${P}/spool/${S}/shipment`, {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify(body)
  });
}
async function toggle(n){
  const st=stepMap[n]||{completed:0};
  const newVal=st.completed?0:1;
  const op=document.getElementById('operator').value||'';
  const rem=document.querySelector(`#step-${n} .step-rem input`);
  const r=await fetch(`/api/project/${P}/spool/${S}/step/${n}`,{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({completed:newVal,operator:op,remarks:rem?rem.value:''})});
  const d=await r.json();
  if(d.ok) load();
}
async function remark(n,val){
  const st=stepMap[n]||{completed:0};
  const op=document.getElementById('operator').value||'';
  await fetch(`/api/project/${P}/spool/${S}/step/${n}`,{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({completed:st.completed?1:0,operator:st.completed_by||op,remarks:val})});
}
// ── Tab switching ──
function switchTab(tab){
  document.querySelectorAll('.tab-btn').forEach((b,i)=>b.classList.toggle('active', (tab==='checklist'?i===0:i===1)));
  document.getElementById('tab-checklist').classList.toggle('active', tab==='checklist');
  document.getElementById('tab-qc').classList.toggle('active', tab==='qc');
  if(tab==='qc') loadQC();
}
const catLabels = {'fab':['Fabrication','\u5236\u9020'],'weld':['Welding','\u710a\u63a5'],'ndt':['NDT','\u65e0\u635f\u68c0\u6d4b'],'measurement':['Measurement','\u6d4b\u91cf'],'material':['Material','\u6750\u6599'],'surface':['Surface','\u8868\u9762\u5904\u7406']};
const catColors = {'fab':'#2F5496','weld':'#C0392B','ndt':'#8E44AD','measurement':'#27AE60','material':'#E67E22','surface':'#16A085'};
const statusLabels = {'not_started':'\u2014','draft':'Draft / \u8349\u7a3f','submitted':'Submitted / \u5df2\u63d0\u4ea4','approved':'Done / \u5b8c\u6210'};
const statusCssMap = {'not_started':'qs-not_started','draft':'qs-draft','submitted':'qs-submitted','approved':'qs-approved'};
async function loadQC(){
  // Auto-seed from DXF on first load (no-op if already seeded)
  await fetch(`/api/project/${P}/spool/${S}/qc/seed`,{method:'POST'});
  const r = await fetch(`/api/project/${P}/spool/${S}/qc`);
  const reports = await r.json();
  if(!reports.length){document.getElementById('qc-reports').innerHTML='<p style="padding:20px;color:#888;text-align:center">No QC reports for this project / \u65e0\u8d28\u68c0\u62a5\u544a</p>';return;}
  const total = reports.length;
  const done = reports.filter(r=>r.status==='approved').length;
  const draft = reports.filter(r=>r.status==='draft').length;
  let html=`<div style="display:flex;justify-content:space-between;align-items:center;padding:8px 0">
    <div style="font-size:11px;color:#888">${total} reports \u00b7 ${done} done \u00b7 ${draft} draft</div>
    <div style="display:flex;gap:6px">
      <a href="/api/project/${P}/spool/${S}/qc/export" style="padding:4px 10px;font-size:10px;border:1px solid #27ae60;border-radius:4px;background:#27ae60;color:#fff;text-decoration:none;cursor:pointer">Export PDF / \u5bfc\u51faPDF</a>
    </div></div>`;
  let lastCat='';
  reports.forEach(rp=>{
    if(rp.category!==lastCat){
      lastCat=rp.category;
      const cl=catLabels[rp.category]||[rp.category,''];
      html+=`<div class="qc-cat-header" style="color:${catColors[rp.category]||'#888'}">${cl[0]} / ${cl[1]}</div>`;
    }
    const cardCls = rp.status==='approved'?'s-done':rp.status==='draft'?'s-draft':'';
    const url = `/project/${P}/spool/${S}/qc/${rp.type}${rp.subtype?'?sub='+rp.subtype:''}`;
    const holdBadge = rp.is_hold ? ' <span style="background:#EDE7F6;color:#5E35B1;font-size:8px;font-weight:700;padding:1px 5px;border-radius:3px">HOLD</span>' : '';
    html+=`<a class="qc-card ${cardCls}" href="${url}">
      <div class="qc-icon">${rp.icon}</div>
      <div class="qc-info">
        <div class="en">${rp.name_en}${holdBadge}</div>
        <div class="cn">${rp.name_cn} \u00b7 ITP Step ${rp.itp_step}${rp.sub_label?' \u00b7 '+rp.sub_label:''}</div>
        <div style="font-size:8px;color:#aaa;font-family:monospace;margin-top:1px">${rp.rec_no}</div>
      </div>
      <div class="qc-status ${statusCssMap[rp.status]||'qs-not_started'}">${statusLabels[rp.status]||rp.status}</div>
    </a>`;
  });
  document.getElementById('qc-reports').innerHTML=html;
}
async function seedQC(){
  const r=await fetch(`/api/project/${P}/spool/${S}/qc/seed`,{method:'POST'});
  const d=await r.json();
  if(d.ok) loadQC();
  else alert(d.error||'Seed failed');
}
load();
if(new URLSearchParams(window.location.search).get('tab')==='qc') switchTab('qc');
</script></body></html>"""

QC_REPORT_HTML = """<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{{ spool_id }} — QC Report</title><style>""" + COMMON_CSS + """
.qc-header{padding:12px 16px;background:#fff;box-shadow:0 1px 2px rgba(0,0,0,.05)}
.qc-header h2{font-size:16px;margin:4px 0 2px}.qc-header .qc-sub{font-size:11px;color:#888}
.inspector-bar{background:#fff;padding:12px 16px;display:grid;grid-template-columns:1fr 1fr;gap:8px;box-shadow:0 1px 2px rgba(0,0,0,.05);margin-bottom:8px}
.inspector-bar label{font-size:10px;color:#888;display:block;margin-bottom:2px}
.inspector-bar input,.inspector-bar select{width:100%;padding:8px;border:1px solid #ddd;border-radius:6px;font-size:13px}
.qc-section{background:#fff;border-radius:10px;margin:8px 16px;padding:14px 16px;box-shadow:0 1px 2px rgba(0,0,0,.05)}
.qc-section h3{font-size:13px;font-weight:700;color:#2F5496;margin-bottom:10px;border-bottom:2px solid #D6E4F0;padding-bottom:6px}
.qc-row{display:grid;gap:8px;align-items:center;padding:8px 0;border-bottom:1px solid #f0f2f5}
.qc-row:last-child{border-bottom:none}
.qc-label{font-size:12px;font-weight:600;color:#333}.qc-label .cn{font-size:10px;color:#888;font-weight:400}
.qc-nominal{font-size:12px;color:#888;text-align:center}
.qc-input{padding:8px;border:1px solid #ddd;border-radius:6px;font-size:14px;text-align:center;width:100%;background:#FFFDE7;font-weight:600;color:#2F5496}
.qc-input:focus{border-color:#2F5496;outline:none;box-shadow:0 0 0 2px rgba(47,84,150,.15)}
.qc-input.qc-readonly{background:#f8f9fa;color:#666;font-weight:400}
.pass-fail{display:flex;gap:4px}
.pf-btn{flex:1;padding:10px 8px;border:2px solid #ddd;border-radius:8px;font-size:12px;font-weight:700;cursor:pointer;text-align:center;background:#fff;transition:all .15s}
.pf-btn.sel-pass{border-color:#27ae60;background:#e8f5e9;color:#1B7A1B}
.pf-btn.sel-fail{border-color:#e74c3c;background:#fce4ec;color:#c0392b}
.pf-btn:active{transform:scale(.97)}
.overall-result{margin:8px 16px;padding:16px;background:#fff;border-radius:10px;box-shadow:0 1px 2px rgba(0,0,0,.05);text-align:center}
.overall-result h3{font-size:14px;color:#333;margin-bottom:12px}
.result-btns{display:flex;gap:8px;justify-content:center}
.result-btns .pf-btn{padding:14px 24px;font-size:14px;min-width:100px}
.save-status{position:fixed;bottom:20px;right:20px;padding:8px 16px;border-radius:20px;font-size:12px;font-weight:600;opacity:0;transition:opacity .3s;z-index:100}
.save-ok{background:#27ae60;color:#fff}.save-err{background:#e74c3c;color:#fff}
.img-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(120px,1fr));gap:8px;margin-top:8px}
.img-thumb{position:relative;border-radius:8px;overflow:hidden;aspect-ratio:1;background:#f0f2f5}
.img-thumb img{width:100%;height:100%;object-fit:cover}
.img-thumb .del{position:absolute;top:4px;right:4px;background:rgba(0,0,0,.6);color:#fff;border:none;border-radius:50%;width:24px;height:24px;cursor:pointer;font-size:14px;line-height:24px;text-align:center}
.upload-btn{display:flex;align-items:center;justify-content:center;gap:8px;padding:12px;border:2px dashed #ccc;border-radius:8px;cursor:pointer;font-size:13px;color:#888;margin-top:8px}
.upload-btn:active{background:#f8f9fa}
.weld-card{background:#f8f9fa;border-radius:8px;padding:12px;margin-bottom:8px;border-left:3px solid #8E44AD}
.weld-card .weld-tag{font-size:14px;font-weight:700;color:#333}.weld-card .weld-size{font-size:11px;color:#888}
.weld-grid{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-top:8px}
</style></head><body>
<div class="header">
  <a class="back" href="/project/{{ project }}/spool/{{ spool_id }}?tab=qc">\u2190 {{ spool_id }}</a>
  <h1 id="rpt-title">QC Report</h1>
  <div class="sub" id="rpt-sub">Loading... / 加载中...</div>
</div>
<div class="inspector-bar">
  <div>
    <label>Inspector / \u68c0\u9a8c\u5458</label>
    <select id="inspector-select" style="width:100%;padding:8px;border:1px solid #ddd;border-radius:6px;font-size:13px" onchange="onInspectorSelect(this.value)">
      <option value="">— Select / \u9009\u62e9\u68c0\u9a8c\u5458 —</option>
    </select>
    <input id="inspector" type="hidden">
  </div>
  <div><label>Date / \u65e5\u671f</label><input id="insp-date" type="date"></div>
  <div style="grid-column:1/-1"><label>ITP</label><div id="itp-display" style="padding:8px;border:1px solid #ddd;border-radius:6px;font-size:13px;font-weight:600;color:#8E44AD;background:#f8f5ff"></div></div>
  <div style="grid-column:1/-1" id="sig-section" style="display:none">
    <label>Signature / \u7b7e\u540d</label>
    <div style="position:relative;border:1px solid #ddd;border-radius:6px;background:#fff;height:80px;touch-action:none" id="sig-container">
      <canvas id="sig-canvas" style="width:100%;height:100%;cursor:crosshair"></canvas>
      <div style="position:absolute;top:4px;right:4px;display:flex;gap:4px">
        <button onclick="clearSig()" style="padding:2px 8px;font-size:10px;border:1px solid #ddd;border-radius:4px;background:#fff;cursor:pointer">Clear / \u6e05\u9664</button>
        <button onclick="saveSig()" style="padding:2px 8px;font-size:10px;border:1px solid #27ae60;border-radius:4px;background:#27ae60;color:#fff;cursor:pointer">Save / \u4fdd\u5b58</button>
      </div>
      <img id="sig-preview" style="position:absolute;inset:0;width:100%;height:100%;object-fit:contain;display:none">
    </div>
  </div>
</div>
<input type="hidden" id="tpi" value=""><input type="hidden" id="tpi-date" value="">
<div id="report-body"></div>
<div class="overall-result">
  <h3>Overall Result / \u6700\u7ec8\u7ed3\u679c</h3>
  <div class="result-btns">
    <div class="pf-btn" onclick="setResult('ACC')" id="res-pass">✓ ACC 合格</div>
    <div class="pf-btn" onclick="setResult('REJ')" id="res-fail">✗ REJ 不合格</div>
  </div>
</div>
<div class="save-status" id="save-status"></div>
<script>
const P='{{project}}',S='{{spool_id}}',RT='{{report_type}}',SUB='{{report_subtype}}';
let reportData = {};
let saveTimer = null;
let MATERIAL_TYPE = '';

// ── Report type configs — per Danny's specs 2026-04-04 ──
const REPORT_FIELDS = {
  cutting:     { title:'Cutting Inspection / 切割检验', render: renderCutting },
  fitup:       { title:'Fit-up Inspection / 组对检验', render: renderFitup },
  welding_log: { title:'Welding Log / 焊接记录', render: renderWeldingLog },
  vt:          { title:'Visual Testing (VT) / 目视检验', render: renderVT },
  rt:          { title:'Radiographic Testing (RT) / 射线检测', render: renderRT },
  pt:          { title:'Penetrant Testing (PT) / 渗透检测', render: renderPT },
  mt:          { title:'Magnetic Particle (MT) / 磁粉检测', render: renderMT },
  pmi:         { title:'PMI Report / PMI检测报告', render: renderPMI },
  ferrite:     { title:'Ferrite Content / 铁素体含量', render: renderFerrite },
  dimensional: { title:'Dimensional Inspection / 尺寸检验', render: renderDimensional },
  ferroxyl:    { title:'Ferroxyl Test / 铁离子检测', render: renderFerroxyl, hasImages: true },
  dft:         { title:'Coating Inspection (DFT) / 涂层检验', render: renderDFT },
};

// ── Shared helpers ──
function inp(key,val,opts={}){const t=opts.type||'text',ph=opts.ph||'',st=opts.style||'';return `<input class="qc-input" type="${t}" ${t==='number'?'inputmode="decimal" step="any"':''} value="${val!=null?val:opts.def||''}" placeholder="${ph}" style="font-size:12px;padding:6px;${st}" onchange="setD('${key}',${t==='number'?"this.value?parseFloat(this.value):null":"this.value"});scheduleSave()">`;}
function setD(path,val){let o=reportData.data,parts=path.split('.');for(let i=0;i<parts.length-1;i++){if(!o[parts[i]])o[parts[i]]={};o=o[parts[i]];}o[parts[parts.length-1]]=val;}
function sH(t){return `<div class="qc-section"><h3>${t}</h3>`;}
function reRender(){const cfg=REPORT_FIELDS[RT];if(cfg)cfg.render(reportData.data);scheduleSave();}
function pfCell(key,val){const cls=val===true?'pass':val===false?'fail':'';const txt=val===true?'✓ ACC 合格':val===false?'✗ REJ 不合格':'—';return `<td style="padding:4px;border:1px solid #ddd;text-align:center;cursor:pointer;font-size:11px" class="${cls}" onclick="setD('${key}',${val===true?'false':(val===false?'null':'true')});reRender()">${txt}</td>`;}
function accRejCell(key,val){const cls=val==='ACC'?'pass':val==='REJ'?'fail':'';const txt=val==='ACC'?'✓ ACC 合格':val==='REJ'?'✗ REJ 不合格':'—';return `<td style="padding:4px;border:1px solid #ddd;text-align:center;cursor:pointer;font-size:11px" class="${cls}" onclick="var c=getDV('${key}');setD('${key}',c==='ACC'?'REJ':(c==='REJ'?null:'ACC'));reRender()">${txt}</td>`;}
function getDV(path){let o=reportData.data,parts=path.split('.');for(let i=0;i<parts.length;i++){if(!o)return null;o=o[parts[i]];}return o;}
let seedData = null;
async function seedIfNeeded(){
  if(seedData) return;
  try{const r=await fetch(`/api/project/${P}/spool/${S}/qc/seed`,{method:'POST'});seedData=await r.json();}catch(e){}
}

// ══════════════════════════════════════════════════════════════════════════════
// 1. CUTTING — Simple: cut lengths from DXF + pass/fail
// ══════════════════════════════════════════════════════════════════════════════
function renderCutting(data){
  const ref = data.drawing_ref || `${P}-${S}`;
  let html = sH('Drawing Reference / 图纸参考');
  html += `<div style="font-size:13px;font-weight:600;color:#2F5496;padding:4px 0">${ref}</div></div>`;
  const pieces = data.pieces || [];
  html += sH('Pipe Cut Inspection / 管件切割检查');
  if(!pieces.length){html += '<p style="color:#888;text-align:center;padding:12px">No cut data — Seed from DXF / 无切割数据，请从DXF导入</p>';}
  else {
    html += `<table style="width:100%;border-collapse:collapse;font-size:11px"><tr style="background:#f0f2f5">
      <th style="padding:6px;border:1px solid #ddd;text-align:left">Piece / 件号</th>
      <th style="padding:6px;border:1px solid #ddd">Size / 尺寸</th>
      <th style="padding:6px;border:1px solid #ddd">Nominal (mm) / 标称</th>
      <th style="padding:6px;border:1px solid #ddd">Actual (mm) / 实测</th>
      <th style="padding:6px;border:1px solid #ddd">Pass/Fail / 合格</th></tr>`;
    pieces.forEach((p,i) => {
      html += `<tr><td style="padding:4px 6px;border:1px solid #ddd;font-weight:600">${p.mark||'Cut '+(i+1)}</td>
        <td style="padding:4px;border:1px solid #ddd;text-align:center;font-size:10px">${p.size||''}</td>
        <td style="padding:4px;border:1px solid #ddd;text-align:center">${p.nominal?Math.round(p.nominal):''}</td>
        <td style="padding:4px;border:1px solid #ddd">${inp('pieces.'+i+'.actual',p.actual,{type:'number',ph:'mm / 毫米'})}</td>
        ${pfCell('pieces.'+i+'.pass',p.pass)}</tr>`;
    });
    html += '</table>';
  }
  html += '</div>';
  document.getElementById('report-body').innerHTML = html;
}

// ══════════════════════════════════════════════════════════════════════════════
// 2. FIT-UP — Geometric check pass/fail per joint (details in dimensional)
// ══════════════════════════════════════════════════════════════════════════════
function renderFitup(data){
  const welds = data.welds || [];
  let html = sH('Fit-up & Alignment — Geometric Check / 组对校准—几何检查');
  html += `<div class="acc-box"><strong>Standard / 标准:</strong> ASME B16.25 + WPS<br>Detailed measurements in Dimensional Report / 详细测量见尺寸报告</div>`;
  if(!welds.length){html += '<p style="color:#888;text-align:center;padding:12px">No weld data — Seed from DXF / 无焊缝数据，请从DXF导入</p>';}
  else {
    html += `<table style="width:100%;border-collapse:collapse;font-size:11px"><tr style="background:#f0f2f5">
      <th style="padding:6px;border:1px solid #ddd;text-align:left">Joint / 接头</th>
      <th style="padding:6px;border:1px solid #ddd">Size / 尺寸</th><th style="padding:6px;border:1px solid #ddd">Type / 类型</th>
      <th style="padding:6px;border:1px solid #ddd">Geometric Check<br>几何检查</th></tr>`;
    welds.forEach((w,i) => {
      html += `<tr><td style="padding:4px 6px;border:1px solid #ddd;font-weight:600">${w.weld_tag||'W'+(i+1)}</td>
        <td style="padding:4px;border:1px solid #ddd;text-align:center">${w.size||''}</td>
        <td style="padding:4px;border:1px solid #ddd;text-align:center">${w.weld_type||'BW'}</td>
        ${pfCell('welds.'+i+'.geometric_ok',w.geometric_ok)}</tr>`;
    });
    html += '</table>';
  }
  html += '</div>';
  html += sH('Remarks / 备注') + `<textarea class="qc-input" style="text-align:left;font-weight:400;height:50px;resize:vertical;background:#fff" onchange="setD('remarks',this.value);scheduleSave()">${data.remarks||''}</textarea></div>`;
  document.getElementById('report-body').innerHTML = html;
}

// ══════════════════════════════════════════════════════════════════════════════
// 3. WELDING LOG — WPS dropdown + tick that it was followed
// ══════════════════════════════════════════════════════════════════════════════
let wpsOptions = {};
async function loadWPS(){try{const r=await fetch(`/api/project/${P}/wps`);wpsOptions=await r.json();}catch(e){}}
function toggleWPS(k){
  if(!reportData.data.wps_numbers) reportData.data.wps_numbers=[];
  const idx=reportData.data.wps_numbers.indexOf(k);
  if(idx>=0) reportData.data.wps_numbers.splice(idx,1);
  else reportData.data.wps_numbers.push(k);
  reRender();
}
function renderWeldingLog(data){
  loadWPS().then(()=>renderWeldingLogInner(data));
}
function renderWeldingLogInner(data){
  const welds = data.welds || [];
  const isDuplex = MATERIAL_TYPE==='duplex';
  const selectedList = data.wps_numbers || (data.wps_number ? [data.wps_number] : []);
  if(!data.wps_numbers && data.wps_number) reportData.data.wps_numbers = [data.wps_number];
  let html = sH('WPS Applied / 应用的焊接工艺 <span style="font-size:10px;font-weight:400;color:#888">(select all that apply / 选择所有适用的)</span>');
  for(const [k,v] of Object.entries(wpsOptions)){
    const checked = selectedList.includes(k);
    html += `<div style="display:flex;align-items:flex-start;gap:8px;padding:6px 0;border-bottom:1px solid #f0f0f0;cursor:pointer" onclick="toggleWPS('${k}')">
      <div style="font-size:18px;min-width:24px;text-align:center">${checked?'☑':'☐'}</div>
      <div style="flex:1">
        <div style="font-size:12px;font-weight:600;color:${checked?'#2F5496':'#888'}">${k}</div>
        <div style="font-size:10px;color:#888">${v.label} · ${v.thickness}</div>
      </div></div>`;
  }
  // Show details for selected WPS
  selectedList.forEach(sel => {
    const wps = wpsOptions[sel];
    if(!wps) return;
    html += `<div style="background:#f8f9fa;border:1px solid #ddd;border-radius:6px;padding:10px;margin-top:8px;font-size:11px;border-left:3px solid #2F5496">
      <div style="font-weight:700;color:#2F5496;margin-bottom:4px">${sel}</div>
      <div><strong>Process / 工艺:</strong> ${wps.processes} · <strong>Type / 类型:</strong> ${wps.type} · <strong>Thickness / 厚度:</strong> ${wps.thickness}</div>
      <div><strong>Filler / 焊材:</strong> ${wps.filler} (${wps.filler_dia})</div>
      <div><strong>Shielding / 保护气:</strong> ${wps.shielding} · <strong>Backing / 背面气:</strong> ${wps.backing}</div>
      <div><strong>Preheat / 预热:</strong> ${wps.preheat} · <strong>Interpass / 层间:</strong> <span style="color:${isDuplex?'#e74c3c':'inherit'};font-weight:${isDuplex?'700':'400'}">${wps.interpass}</span></div>
      <div><strong>Root / 根层:</strong> ${wps.root}</div>
      <div><strong>Fill / 填充:</strong> ${wps.fill}</div>
      <div><strong>Cap / 盖面:</strong> ${wps.cap}</div>
      <div><strong>Heat Input / 热输入:</strong> ${wps.heat_input} · <strong>Position / 位置:</strong> ${wps.position}</div>
    </div>`;
  });
  if(selectedList.length > 0){
    html += `<div style="margin-top:8px;display:flex;align-items:center;gap:8px">
      <div class="pf-btn ${data.wps_followed?'sel-pass':''}" style="padding:10px 20px;font-size:13px" onclick="setD('wps_followed',!reportData.data.wps_followed);reRender()">
        ${data.wps_followed?'✓ WPS Followed / 已遵循焊接工艺':'☐ Confirm WPS Followed / 确认遵循焊接工艺'}</div></div>`;
  }
  html += '</div>';
  html += sH('Welds / 焊缝');
  if(!welds.length){html += '<p style="color:#888;text-align:center;padding:12px">No weld data — Seed from DXF / 无焊缝数据，请从DXF导入</p>';}
  else {
    html += `<table style="width:100%;border-collapse:collapse;font-size:10px"><tr style="background:#f0f2f5">
      <th style="padding:4px;border:1px solid #ddd;text-align:left">Weld / 焊缝</th><th style="padding:4px;border:1px solid #ddd">Size / 尺寸</th>
      <th style="padding:4px;border:1px solid #ddd">Welder ID / 焊工号</th><th style="padding:4px;border:1px solid #ddd">Date / 日期</th>
      <th style="padding:4px;border:1px solid #ddd">Completed / 完成</th></tr>`;
    welds.forEach((w,i) => {
      html += `<tr><td style="padding:4px;border:1px solid #ddd;font-weight:600">${w.weld_tag||'W'+(i+1)}</td>
        <td style="padding:4px;border:1px solid #ddd;text-align:center">${w.size||''}</td>
        <td style="padding:3px;border:1px solid #ddd">${inp('welds.'+i+'.welder_id',w.welder_id,{style:'padding:3px;font-size:10px'})}</td>
        <td style="padding:3px;border:1px solid #ddd">${inp('welds.'+i+'.date',w.date,{type:'date',style:'padding:3px;font-size:10px'})}</td>
        ${pfCell('welds.'+i+'.done',w.done)}</tr>`;
    });
    html += '</table>';
  }
  html += '</div>';
  html += sH('Remarks / 备注') + `<textarea class="qc-input" style="text-align:left;font-weight:400;height:50px;resize:vertical;background:#fff" onchange="setD('remarks',this.value);scheduleSave()">${data.remarks||''}</textarea></div>`;
  document.getElementById('report-body').innerHTML = html;
}

// ══════════════════════════════════════════════════════════════════════════════
// 4. VT — 12-item checklist per Danny's spec
// ══════════════════════════════════════════════════════════════════════════════
const VT_CHECKS = [
  {key:'id_correct',en:'Identification correct: spool no., line no., material, size, rating per drawing',cn:'标识正确：管段号、管线号、材料、尺寸、等级按图纸'},
  {key:'surface_suitable',en:'Surface condition suitable for VT: clean, accessible, adequate lighting',cn:'表面状况适合VT：清洁、可达、充足照明'},
  {key:'workmanship',en:'Workmanship acceptable: welds and base metal free from visible damage',cn:'工艺合格：焊缝和母材无可见损伤'},
  {key:'dimensions_align',en:'Dimensions / alignment correct: fit-up, orientation, flange alignment, branch location',cn:'尺寸/对中正确：组对、方向、法兰对中、支管位置'},
  {key:'weld_profile',en:'Weld profile acceptable: reinforcement, contour, transition, fillet size',cn:'焊缝轮廓合格：余高、轮廓、过渡、角焊缝尺寸'},
  {key:'no_cracks',en:'No visible cracks in weld or HAZ',cn:'焊缝或HAZ无可见裂纹'},
  {key:'no_fusion_defect',en:'No incomplete fusion / incomplete penetration where visible',cn:'无可见未熔合/未焊透'},
  {key:'no_surface_defect',en:'No undercut, overlap, arc strikes, craters, or excessive porosity',cn:'无咬边、搭接、弧击、弧坑或过量气孔'},
  {key:'no_damage',en:'No surface defects: gouges, dents, spatter, mechanical damage',cn:'无表面缺陷：刻痕、凹坑、飞溅、机械损伤'},
  {key:'attachments_ok',en:'Attachments / end prep acceptable: bevels, sockets, temporary attachments removed',cn:'附件/端口准备合格：坡口、承插口、临时附件已拆除'},
  {key:'acceptance_verified',en:'Acceptance verified to ASME B31.3 Table 341.3.2',cn:'按ASME B31.3表341.3.2验收合格'},
  {key:'vt_recorded',en:'VT performed per ASME Section V, Article 9 and results recorded',cn:'VT按ASME第V卷第9条执行并记录'},
];
function renderVT(data){
  let html = sH('Visual Testing Checklist / 目视检验清单');
  html += `<div class="acc-box"><strong>Standard / 标准:</strong> ASME B31.3 Table 341.3.2 / ASME Section V Art. 9</div>`;
  html += '<table style="width:100%;border-collapse:collapse;font-size:11px">';
  VT_CHECKS.forEach((c,i) => {
    const val = data[c.key];
    html += `<tr style="background:${i%2?'#fafafa':'#fff'}">
      <td style="padding:6px 8px;border:1px solid #eee;font-size:11px">${i+1}. ${c.en}<br><span style="font-size:10px;color:#888">${c.cn}</span></td>
      ${pfCell(c.key,val)}</tr>`;
  });
  html += '</table></div>';
  html += sH('Remarks / 备注') + `<textarea class="qc-input" style="text-align:left;font-weight:400;height:50px;resize:vertical;background:#fff" onchange="setD('remarks',this.value);scheduleSave()">${data.remarks||''}</textarea></div>`;
  document.getElementById('report-body').innerHTML = html;
}

// ══════════════════════════════════════════════════════════════════════════════
// 5. RT — Full equipment + per-weld × 3 films with defects + density
// ══════════════════════════════════════════════════════════════════════════════
function renderRT(data){
  const welds = data.welds || [];
  let html = sH('RT Equipment & Parameters / 射线设备及参数');
  html += `<div class="weld-grid" style="grid-template-columns:1fr 1fr">
    <div><label style="font-size:10px;color:#888">Instrument / 仪器型号</label>${inp('instrument',data.instrument,{ph:'XXGHz-1605 / 仪器型号'})}</div>
    <div><label style="font-size:10px;color:#888">Source Type / 射源</label>${inp('source_type',data.source_type,{def:'Ir-192'})}</div>
    <div><label style="font-size:10px;color:#888">Focus Size (mm) / 焦点尺寸</label>${inp('focus_size',data.focus_size,{ph:'1.0×2.4 / 焦点尺寸'})}</div>
    <div><label style="font-size:10px;color:#888">SFD (mm) / 焦距</label>${inp('sfd',data.sfd)}</div>
    <div><label style="font-size:10px;color:#888">IQI Type / 透度计</label>${inp('iqi_type',data.iqi_type,{def:'ASTM'})}</div>
    <div><label style="font-size:10px;color:#888">Required Sensitivity / 应识别丝号</label>${inp('iqi_required',data.iqi_required)}</div>
    <div><label style="font-size:10px;color:#888">Tube Voltage (kV) / 管电压</label>${inp('tube_voltage',data.tube_voltage,{type:'number'})}</div>
    <div><label style="font-size:10px;color:#888">Tube Current (mA) / 管电流</label>${inp('tube_current',data.tube_current,{type:'number'})}</div>
    <div><label style="font-size:10px;color:#888">Exposure Time / 曝光时间</label>${inp('exposure_time',data.exposure_time)}</div>
    <div><label style="font-size:10px;color:#888">Film Brand / 胶片</label>${inp('film_brand',data.film_brand)}</div>
    <div><label style="font-size:10px;color:#888">Density Range / 黑度范围</label>${inp('density_range',data.density_range,{def:'1.8-4.0'})}</div>
    <div><label style="font-size:10px;color:#888">Technique / 透照方式</label>${inp('technique',data.technique,{def:'SWSI'})}</div>
  </div>
  <div class="acc-box" style="margin-top:8px"><strong>Exam Standard / 检验标准:</strong> Art.2 of ASME V<br><strong>Acceptance / 验收标准:</strong> ASME B31.3</div>`;
  html += '</div>';
  html += sH('Film Results / 底片检查结果 <span style="font-size:10px;font-weight:400;color:#888">(3 films per weld)</span>');
  if(!welds.length){html += '<p style="color:#888;text-align:center;padding:12px">No weld data — Seed from DXF / 无焊缝数据，请从DXF导入<br><small>Hugo will upload RT results / Hugo将上传RT结果</small></p>';}
  else {
    welds.forEach(w=>{if(!w.films||w.films.length<3)w.films=[0,1,2].map(fi=>(w.films&&w.films[fi])||{film_id:'',density:'',bar:false,circular:false,crack:false,lop:false,lof:false,result:null});});
    const dc=[{k:'bar',en:'Bar',cn:'条缺'},{k:'circular',en:'Circ.',cn:'圆缺'},{k:'crack',en:'Crack',cn:'裂纹'},{k:'lop',en:'LOP',cn:'未焊透'},{k:'lof',en:'LOF',cn:'未熔合'}];
    html += `<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:10px"><tr style="background:#f0f2f5">
      <th style="padding:4px;border:1px solid #ddd" rowspan="2">Weld / 焊缝</th><th style="padding:4px;border:1px solid #ddd" rowspan="2">Film / 底片</th>
      <th style="padding:4px;border:1px solid #ddd" colspan="5">Defects / 缺陷</th>
      <th style="padding:4px;border:1px solid #ddd" rowspan="2">Density / 黑度</th><th style="padding:4px;border:1px solid #ddd" rowspan="2">Acc/Rej / 合格</th></tr>
      <tr style="background:#f0f2f5">${dc.map(d=>`<th style="padding:3px;border:1px solid #ddd;font-size:8px">${d.en}<br>${d.cn}</th>`).join('')}</tr>`;
    welds.forEach((w,wi)=>{
      w.films.forEach((f,fi)=>{
        html+=`<tr>${fi===0?`<td style="padding:4px;border:1px solid #ddd;font-weight:700;vertical-align:top" rowspan="3">${w.weld_tag||'W'+(wi+1)}<br><span style="font-size:9px;font-weight:400;color:#888">${w.size||''}</span></td>`:''}
          <td style="padding:3px;border:1px solid #ddd"><input class="qc-input" style="padding:2px;font-size:10px;width:45px" value="${f.film_id||''}" placeholder="${fi+1}" onchange="reportData.data.welds[${wi}].films[${fi}].film_id=this.value;scheduleSave()"></td>`;
        dc.forEach(d=>{const v=f[d.k];html+=`<td style="padding:2px;border:1px solid #ddd;text-align:center;cursor:pointer" class="${v?'fail':''}" onclick="reportData.data.welds[${wi}].films[${fi}].${d.k}=!reportData.data.welds[${wi}].films[${fi}].${d.k};reRender()">${v?'✗':'—'}</td>`;});
        html+=`<td style="padding:2px;border:1px solid #ddd"><input class="qc-input" style="padding:2px;font-size:10px;width:55px" value="${f.density||''}" placeholder="2.0-4.0 / 黑度" onchange="reportData.data.welds[${wi}].films[${fi}].density=this.value;scheduleSave()"></td>
          <td style="padding:2px;border:1px solid #ddd;text-align:center;cursor:pointer" class="${f.result==='ACC'?'pass':f.result==='REJ'?'fail':''}" onclick="var r=reportData.data.welds[${wi}].films[${fi}];r.result=r.result==='ACC'?'REJ':(r.result==='REJ'?null:'ACC');reRender()">${f.result==='ACC'?'✓ ACC':(f.result==='REJ'?'✗ REJ':'—')}</td></tr>`;
      });
    });
    html += '</table></div>';
  }
  html += '</div>';
  html += sH('Remarks / 备注') + `<textarea class="qc-input" style="text-align:left;font-weight:400;height:50px;resize:vertical;background:#fff" onchange="setD('remarks',this.value);scheduleSave()">${data.remarks||''}</textarea></div>`;
  document.getElementById('report-body').innerHTML = html;
}

// ══════════════════════════════════════════════════════════════════════════════
// 6. PT — Per weld pass/fail (ASME V Art. 6, chloride-free for Duplex)
// ══════════════════════════════════════════════════════════════════════════════
function renderPT(data){
  const welds = data.welds || [];
  const isDuplex = MATERIAL_TYPE==='duplex';
  let html = sH('PT Procedure / 渗透检测规程');
  if(isDuplex) html += `<div class="warn-box">⚠ Chloride-free penetrant mandatory for Duplex SS / 双相不锈钢必须使用无氯渗透剂 / 双相不锈钢必须使用无氯渗透剂</div>`;
  html += `<div class="acc-box"><strong>Standard / 标准:</strong> ASME Section V, Article 6<br><strong>Acceptance / 验收:</strong> ASME B31.3 Table 341.3.2</div></div>`;
  html += sH('Weld Examination / 焊缝检查');
  if(!welds.length){html += '<p style="color:#888;text-align:center;padding:12px">No weld data — Seed from DXF / 无焊缝数据，请从DXF导入</p>';}
  else {
    html += `<table style="width:100%;border-collapse:collapse;font-size:11px"><tr style="background:#f0f2f5">
      <th style="padding:6px;border:1px solid #ddd;text-align:left">Weld / 焊缝</th><th style="padding:6px;border:1px solid #ddd">Size / 尺寸</th>
      <th style="padding:6px;border:1px solid #ddd">Indication<br>显示</th><th style="padding:6px;border:1px solid #ddd">Acc/Rej / 合格</th></tr>`;
    welds.forEach((w,i) => {
      html += `<tr><td style="padding:4px 6px;border:1px solid #ddd;font-weight:600">${w.weld_tag||'W'+(i+1)}</td>
        <td style="padding:4px;border:1px solid #ddd;text-align:center">${w.size||''}</td>
        <td style="padding:3px;border:1px solid #ddd">${inp('welds.'+i+'.indication',w.indication,{def:'NRI',ph:'NRI 无可报告显示',style:'padding:3px;font-size:10px'})}</td>
        ${accRejCell('welds.'+i+'.result',w.result)}</tr>`;
    });
    html += '</table>';
  }
  html += '</div>';
  html += sH('Remarks / 备注') + `<textarea class="qc-input" style="text-align:left;font-weight:400;height:50px;resize:vertical;background:#fff" onchange="setD('remarks',this.value);scheduleSave()">${data.remarks||''}</textarea></div>`;
  document.getElementById('report-body').innerHTML = html;
}

// ══════════════════════════════════════════════════════════════════════════════
// 7. MT — Per weld pass/fail (ASME V Art. 7, 424 CS only)
// ══════════════════════════════════════════════════════════════════════════════
function renderMT(data){
  const welds = data.welds || [];
  let html = sH('MT Procedure / 磁粉检测规程');
  html += `<div class="acc-box"><strong>Standard / 标准:</strong> ASME Section V, Article 7<br><strong>Acceptance / 验收:</strong> ASME B31.3 Table 341.3.2</div></div>`;
  html += sH('Weld Examination / 焊缝检查');
  if(!welds.length){html += '<p style="color:#888;text-align:center;padding:12px">No weld data — Seed from DXF / 无焊缝数据，请从DXF导入</p>';}
  else {
    html += `<table style="width:100%;border-collapse:collapse;font-size:11px"><tr style="background:#f0f2f5">
      <th style="padding:6px;border:1px solid #ddd;text-align:left">Weld / 焊缝</th><th style="padding:6px;border:1px solid #ddd">Size / 尺寸</th>
      <th style="padding:6px;border:1px solid #ddd">Indication<br>显示</th><th style="padding:6px;border:1px solid #ddd">Acc/Rej / 合格</th></tr>`;
    welds.forEach((w,i) => {
      html += `<tr><td style="padding:4px 6px;border:1px solid #ddd;font-weight:600">${w.weld_tag||'W'+(i+1)}</td>
        <td style="padding:4px;border:1px solid #ddd;text-align:center">${w.size||''}</td>
        <td style="padding:3px;border:1px solid #ddd">${inp('welds.'+i+'.indication',w.indication,{ph:'NRI 无显示',style:'padding:3px;font-size:10px'})}</td>
        ${accRejCell('welds.'+i+'.result',w.result)}</tr>`;
    });
    html += '</table>';
  }
  html += '</div>';
  html += sH('Remarks / 备注') + `<textarea class="qc-input" style="text-align:left;font-weight:400;height:50px;resize:vertical;background:#fff" onchange="setD('remarks',this.value);scheduleSave()">${data.remarks||''}</textarea></div>`;
  document.getElementById('report-body').innerHTML = html;
}

// ══════════════════════════════════════════════════════════════════════════════
// 8. PMI (ASME B31.3 Sec 323.2.4) — 423 Duplex only
// ══════════════════════════════════════════════════════════════════════════════
function renderPMI(data){
  const items = data.items || [];
  let html = sH('PMI Instrument / PMI设备');
  html += `<div class="weld-grid" style="grid-template-columns:1fr 1fr">
    <div><label style="font-size:10px;color:#888">Instrument / 仪器</label>${inp('instrument_model',data.instrument_model)}</div>
    <div><label style="font-size:10px;color:#888">Serial No. / Cal. Date / 序列号/校准日期</label>${inp('instrument_serial',data.instrument_serial)}</div>
  </div>`;
  html += `<div class="acc-box"><strong>Standard / 标准:</strong> ASME B31.3 Sec 323.2.4<br><strong>Expected Grade / 期望牌号:</strong> UNS S32205 (22Cr-5Ni-3Mo)</div></div>`;
  html += sH('PMI Test Points / PMI检测点');
  if(!items.length){html += '<p style="color:#888;text-align:center;padding:12px">No data — Seed from DXF / 无数据</p>';}
  else {
    html += `<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:10px"><tr style="background:#f0f2f5">
      <th style="padding:4px;border:1px solid #ddd;text-align:left">Item / 项目</th><th style="padding:4px;border:1px solid #ddd">Location / 位置</th>
      <th style="padding:4px;border:1px solid #ddd">Cr%</th><th style="padding:4px;border:1px solid #ddd">Ni%</th><th style="padding:4px;border:1px solid #ddd">Mo%</th>
      <th style="padding:4px;border:1px solid #ddd">Grade OK / 牌号确认</th><th style="padding:4px;border:1px solid #ddd">Result / 结果</th></tr>`;
    items.forEach((it,i) => {
      html += `<tr><td style="padding:3px;border:1px solid #ddd;font-size:9px;font-weight:600">${it.item||'Point '+(i+1)}</td>
        <td style="padding:3px;border:1px solid #ddd">${inp('items.'+i+'.location',it.location,{ph:'Base/Weld/HAZ 母材/焊缝/热影响区',style:'padding:2px;font-size:9px;width:65px'})}</td>
        <td style="padding:3px;border:1px solid #ddd">${inp('items.'+i+'.cr',it.cr,{type:'number',style:'padding:2px;font-size:9px;width:45px'})}</td>
        <td style="padding:3px;border:1px solid #ddd">${inp('items.'+i+'.ni',it.ni,{type:'number',style:'padding:2px;font-size:9px;width:45px'})}</td>
        <td style="padding:3px;border:1px solid #ddd">${inp('items.'+i+'.mo',it.mo,{type:'number',style:'padding:2px;font-size:9px;width:45px'})}</td>
        ${pfCell('items.'+i+'.grade_ok',it.grade_ok)}
        ${accRejCell('items.'+i+'.result',it.result)}</tr>`;
    });
    html += '</table></div>';
  }
  html += '</div>';
  html += sH('Remarks / 备注') + `<textarea class="qc-input" style="text-align:left;font-weight:400;height:50px;resize:vertical;background:#fff" onchange="setD('remarks',this.value);scheduleSave()">${data.remarks||''}</textarea></div>`;
  document.getElementById('report-body').innerHTML = html;
}

// ══════════════════════════════════════════════════════════════════════════════
// 9. FERRITE — 3 readings per weld, auto average, auto pass/fail 30-65%
// ══════════════════════════════════════════════════════════════════════════════
function renderFerrite(data){
  const readings = data.readings || [];
  let html = sH('Ferrite Instrument / 铁素体测量设备');
  html += `<div class="weld-grid" style="grid-template-columns:1fr 1fr">
    <div><label style="font-size:10px;color:#888">Instrument / 仪器</label>${inp('instrument_model',data.instrument_model)}</div>
    <div><label style="font-size:10px;color:#888">Serial No. / Cal. Date / 序列号/校准日期</label>${inp('instrument_serial',data.instrument_serial)}</div>
  </div>`;
  html += `<div class="acc-box"><strong>Standard / 标准:</strong> ASTM A799/A800, AWS A4.2<br><strong>Acceptance Range / 验收范围:</strong> 30% – 65% (3 readings per weld → average)</div></div>`;
  html += sH('Ferrite Readings / 铁素体读数');
  if(!readings.length){html += '<p style="color:#888;text-align:center;padding:12px">No weld data — Seed from DXF / 无焊缝数据，请从DXF导入</p>';}
  else {
    html += `<table style="width:100%;border-collapse:collapse;font-size:11px"><tr style="background:#f0f2f5">
      <th style="padding:4px;border:1px solid #ddd;text-align:left">Weld / 焊缝</th>
      <th style="padding:4px;border:1px solid #ddd">R1 (%) / 读数1</th><th style="padding:4px;border:1px solid #ddd">R2 (%) / 读数2</th><th style="padding:4px;border:1px solid #ddd">R3 (%) / 读数3</th>
      <th style="padding:4px;border:1px solid #ddd">Average<br>平均</th><th style="padding:4px;border:1px solid #ddd">Result / 结果</th></tr>`;
    readings.forEach((rd,i) => {
      const avg = (rd.r1!=null&&rd.r2!=null&&rd.r3!=null)?((rd.r1+rd.r2+rd.r3)/3).toFixed(1):'';
      const pass = avg?(parseFloat(avg)>=30&&parseFloat(avg)<=65):null;
      html += `<tr><td style="padding:4px 6px;border:1px solid #ddd;font-weight:600">${rd.weld_tag||'W'+(i+1)} <span style="font-size:9px;color:#888">${rd.size||''}</span></td>
        <td style="padding:3px;border:1px solid #ddd">${inp('readings.'+i+'.r1',rd.r1,{type:'number',style:'padding:3px;font-size:10px;width:55px'})}</td>
        <td style="padding:3px;border:1px solid #ddd">${inp('readings.'+i+'.r2',rd.r2,{type:'number',style:'padding:3px;font-size:10px;width:55px'})}</td>
        <td style="padding:3px;border:1px solid #ddd">${inp('readings.'+i+'.r3',rd.r3,{type:'number',style:'padding:3px;font-size:10px;width:55px'})}</td>
        <td style="padding:4px;border:1px solid #ddd;text-align:center;font-weight:700;color:${pass===true?'#27ae60':pass===false?'#e74c3c':'#888'}">${avg?avg+'%':'—'}</td>
        <td style="padding:4px;border:1px solid #ddd;text-align:center" class="${pass===true?'pass':pass===false?'fail':''}">${pass===true?'ACC':pass===false?'REJ':'—'}</td></tr>`;
    });
    html += '</table>';
  }
  html += '</div>';
  html += sH('Remarks / 备注') + `<textarea class="qc-input" style="text-align:left;font-weight:400;height:50px;resize:vertical;background:#fff" onchange="setD('remarks',this.value);scheduleSave()">${data.remarks||''}</textarea></div>`;
  document.getElementById('report-body').innerHTML = html;
}

// ══════════════════════════════════════════════════════════════════════════════
// 10. DIMENSIONAL — Auto-fill from DXF, editable, flanges + fittings
// ══════════════════════════════════════════════════════════════════════════════
function renderDimensional(data){
  const ref = data.drawing_ref || `${P}-${S}`;
  let html = sH(`Overall Dimensions / 总体尺寸 <span style="font-size:10px;font-weight:400;color:#888">${ref}</span>`);
  const dims = [{k:'l',en:'Length / 长度'},{k:'w',en:'Width / 宽度'},{k:'h',en:'Height / 高度'}];
  html += `<table style="width:100%;border-collapse:collapse;font-size:11px"><tr style="background:#f0f2f5">
    <th style="padding:6px;border:1px solid #ddd;text-align:left">Dimension / 尺寸</th><th style="padding:6px;border:1px solid #ddd">Nominal (mm) / 标称</th>
    <th style="padding:6px;border:1px solid #ddd">Actual (mm) / 实测</th><th style="padding:6px;border:1px solid #ddd">Dev. / 偏差</th><th style="padding:6px;border:1px solid #ddd">OK / 合格</th></tr>`;
  dims.forEach(dim=>{
    const nom=data['nominal_'+dim.k+'_mm']||'';const act=data['actual_'+dim.k+'_mm']||'';
    const dev=(nom&&act)?(parseFloat(act)-parseFloat(nom)).toFixed(1):'';
    const ok=dev?Math.abs(parseFloat(dev))<=5:null;
    html += `<tr><td style="padding:4px 8px;border:1px solid #ddd;font-weight:600">${dim.en}</td>
      <td style="padding:4px;border:1px solid #ddd">${inp('nominal_'+dim.k+'_mm',nom,{type:'number',ph:'From drawing / 图纸标称'})}</td>
      <td style="padding:4px;border:1px solid #ddd">${inp('actual_'+dim.k+'_mm',act,{type:'number',ph:'Measured / 实测'})}</td>
      <td style="padding:4px;border:1px solid #ddd;text-align:center;color:${ok===false?'#e74c3c':'#333'}">${dev?dev+'mm':'—'}</td>
      <td style="padding:4px;border:1px solid #ddd;text-align:center" class="${ok===true?'pass':ok===false?'fail':''}">${ok===true?'✓':ok===false?'✗':'—'}</td></tr>`;
  });
  html += '</table></div>';
  const flanges = data.flanges||[];
  if(flanges.length){
    html += sH('Flange Alignment / 法兰校准 <span style="font-size:10px;font-weight:400">(±1.5mm bolt-hole, 0.5%OD max 2mm perp.)</span>');
    html += `<table style="width:100%;border-collapse:collapse;font-size:11px"><tr style="background:#f0f2f5">
      <th style="padding:6px;border:1px solid #ddd;text-align:left">Flange / 法兰</th><th style="padding:6px;border:1px solid #ddd">Size / 尺寸</th>
      <th style="padding:6px;border:1px solid #ddd">Bolt-Hole<br>螺栓孔</th><th style="padding:6px;border:1px solid #ddd">Perp.<br>垂直度</th></tr>`;
    flanges.forEach((fl,i)=>{
      html += `<tr><td style="padding:4px 8px;border:1px solid #ddd;font-size:10px">${fl.description||'Flange #'+(i+1)}</td>
        <td style="padding:4px;border:1px solid #ddd;text-align:center">${fl.size||''}</td>
        ${pfCell('flanges.'+i+'.bolt_hole_ok',fl.bolt_hole_ok)}
        ${pfCell('flanges.'+i+'.perp_ok',fl.perp_ok)}</tr>`;
    });
    html += '</table></div>';
  }
  const fittings = data.fittings||[];
  if(fittings.length){
    html += sH('Fittings Orientation / 管件方向');
    html += `<table style="width:100%;border-collapse:collapse;font-size:11px"><tr style="background:#f0f2f5">
      <th style="padding:6px;border:1px solid #ddd;text-align:left">Fitting / 管件</th><th style="padding:6px;border:1px solid #ddd">Size / 尺寸</th>
      <th style="padding:6px;border:1px solid #ddd">Direction OK<br>方向正确</th></tr>`;
    fittings.forEach((ft,i)=>{
      html += `<tr><td style="padding:4px 8px;border:1px solid #ddd;font-size:10px">${ft.description||ft.part_type||'Fitting'}</td>
        <td style="padding:4px;border:1px solid #ddd;text-align:center">${ft.size||''}</td>
        ${pfCell('fittings.'+i+'.direction_ok',ft.direction_ok)}</tr>`;
    });
    html += '</table></div>';
  }
  html += sH('Remarks / 备注') + `<textarea class="qc-input" style="text-align:left;font-weight:400;height:50px;resize:vertical;background:#fff" onchange="setD('remarks',this.value);scheduleSave()">${data.remarks||''}</textarea></div>`;
  document.getElementById('report-body').innerHTML = html;
}

// ══════════════════════════════════════════════════════════════════════════════
// 11. FERROXYL — Full duplex scope per Danny's spec
// ══════════════════════════════════════════════════════════════════════════════
function renderFerroxyl(data){
  let html = sH('Ferroxyl Test — Duplex S32205 / 铁离子检测');
  html += `<div style="background:#f8f9fa;border:1px solid #ddd;border-radius:6px;padding:10px;font-size:11px;margin-bottom:8px">
    <div><strong>Material / 材料:</strong> UNS S32205 / Duplex 2205</div>
    <div><strong>Standard / 标准:</strong> ASTM A380/A380M-2017</div>
    <div><strong>Method / 方法:</strong> Copper Sulfate Test / 硫酸铜试验 (Para 7.3.4)</div>
    <div><strong>Contact Time / 接触时间:</strong> Minimum 6 minutes / 最少6分钟</div>
    <div><strong>Acceptance / 验收:</strong> No copper deposition / 无铜沉积 / no copper plating</div>
    <div style="margin-top:6px;color:#8E44AD"><strong>Scope / 范围:</strong> 100% weld seams + adjacent HAZ + ground/brushed/repaired areas (post-pickling)</div>
  </div></div>`;
  const areas = data.areas || [{location:'Weld area + HAZ',copper_deposit:null},{location:'Ground/repaired areas',copper_deposit:null}];
  if(!data.areas) reportData.data.areas = areas;
  html += sH('Test Results / 检测结果');
  html += `<table style="width:100%;border-collapse:collapse;font-size:11px"><tr style="background:#f0f2f5">
    <th style="padding:6px;border:1px solid #ddd;text-align:left">Area / 检测区域</th>
    <th style="padding:6px;border:1px solid #ddd">Surface Condition<br>表面状态</th>
    <th style="padding:6px;border:1px solid #ddd">Copper Deposit<br>铜沉积</th>
    <th style="padding:6px;border:1px solid #ddd">Result / 结果</th></tr>`;
  areas.forEach((a,i)=>{
    const pass=a.copper_deposit===false;const fail=a.copper_deposit===true;
    html += `<tr><td style="padding:4px 8px;border:1px solid #ddd">${inp('areas.'+i+'.location',a.location,{style:'font-size:10px;padding:3px'})}</td>
      <td style="padding:3px;border:1px solid #ddd">${inp('areas.'+i+'.surface_cond',a.surface_cond,{ph:'Pickled/Passivated / 酸洗/钝化',style:'font-size:10px;padding:3px'})}</td>
      <td style="padding:4px;border:1px solid #ddd;text-align:center;cursor:pointer" class="${fail?'fail':pass?'pass':''}" onclick="var v=reportData.data.areas[${i}].copper_deposit;reportData.data.areas[${i}].copper_deposit=v===false?true:(v===true?null:false);reRender()">${pass?'✓ No / 无铜沉积':fail?'✗ Yes / 有铜沉积':'—'}</td>
      <td style="padding:4px;border:1px solid #ddd;text-align:center;font-size:11px" class="${pass?'pass':fail?'fail':''}">${pass?'✓ ACC 合格':fail?'✗ REJ 不合格':'—'}</td></tr>`;
  });
  html += '</table>';
  html += `<div style="margin-top:8px"><div class="img-grid" id="img-grid"></div>
    <label class="upload-btn"><input type="file" accept="image/*" multiple style="display:none" onchange="uploadImages(this.files)">📷 Upload Test Photos / 上传检测照片</label></div>`;
  html += '</div>';
  html += sH('Remarks / 备注') + `<textarea class="qc-input" style="text-align:left;font-weight:400;height:50px;resize:vertical;background:#fff" onchange="setD('remarks',this.value);scheduleSave()">${data.remarks||''}</textarea></div>`;
  document.getElementById('report-body').innerHTML = html;
  loadImages();
}

// ══════════════════════════════════════════════════════════════════════════════
// 12. DFT — Coating inspection (ISO 19840, 424 CS only)
// ══════════════════════════════════════════════════════════════════════════════
function renderDFT(data){
  let html = sH('Coating System / 涂层体系');
  html += `<div class="weld-grid" style="grid-template-columns:1fr 1fr">
    <div><label style="font-size:10px;color:#888">Coat Description / 涂层</label>${inp('coat_desc',data.coat_desc,{ph:'底漆/中间漆/面漆 Primer/Inter./Topcoat'})}</div>
    <div><label style="font-size:10px;color:#888">Product & Batch / 产品批号</label>${inp('product_batch',data.product_batch)}</div>
    <div><label style="font-size:10px;color:#888">Specified NDFT (µm) / 规定干膜厚度</label>${inp('spec_dft',data.spec_dft,{ph:'250-350 / 规定厚度'})}</div>
    <div><label style="font-size:10px;color:#888">Surface Prep / 表面处理</label>${inp('surface_prep',data.surface_prep,{ph:'ISO 8501-1 Sa 2.5 / 表面处理标准'})}</div>
  </div>`;
  html += `<div class="acc-box"><strong>Standard / 标准:</strong> ISO 19840 / ISO 4624 / ISO 8501<br><strong>Rule / 规则:</strong> No reading / 无读数 &lt; 80% NDFT, average ≥ NDFT</div></div>`;
  const readings = data.readings||[{point:'Spot 1',value:null},{point:'Spot 2',value:null},{point:'Spot 3',value:null},{point:'Spot 4',value:null},{point:'Spot 5',value:null}];
  if(!data.readings)reportData.data.readings=readings;
  html += sH('DFT Readings / 干膜厚度读数');
  html += `<table style="width:100%;border-collapse:collapse;font-size:11px"><tr style="background:#f0f2f5">
    <th style="padding:6px;border:1px solid #ddd;text-align:left">Point / 测点</th><th style="padding:6px;border:1px solid #ddd">Reading (µm) / 读数</th><th style="padding:6px;border:1px solid #ddd">Pass / 合格</th></tr>`;
  const specMin=data.spec_dft?parseFloat(data.spec_dft.split('-')[0])*0.8:null;
  const specNom=data.spec_dft?parseFloat(data.spec_dft.split('-')[0]):null;
  readings.forEach((rd,i)=>{
    const ok=(rd.value!=null&&specMin!=null)?rd.value>=specMin:null;
    html += `<tr><td style="padding:4px 8px;border:1px solid #ddd">${rd.point}</td>
      <td style="padding:3px;border:1px solid #ddd">${inp('readings.'+i+'.value',rd.value,{type:'number',ph:'µm / 微米'})}</td>
      <td style="padding:4px;border:1px solid #ddd;text-align:center" class="${ok===true?'pass':ok===false?'fail':''}">${ok===true?'ACC':ok===false?'REJ':'—'}</td></tr>`;
  });
  const vals=readings.filter(r=>r.value!=null).map(r=>r.value);
  const avg=vals.length?(vals.reduce((a,b)=>a+b,0)/vals.length).toFixed(0):'';
  const avgOk=(avg&&specNom)?parseFloat(avg)>=specNom:null;
  html += `<tr style="background:#f0f2f5;font-weight:700"><td style="padding:4px 8px;border:1px solid #ddd">Avg / Min / Max / 平均/最小/最大</td>
    <td style="padding:4px;border:1px solid #ddd;text-align:center">${avg?avg+' / '+Math.min(...vals)+' / '+Math.max(...vals)+' µm':'—'}</td>
    <td style="padding:4px;border:1px solid #ddd;text-align:center" class="${avgOk===true?'pass':avgOk===false?'fail':''}">${avgOk===true?'ACC':avgOk===false?'REJ':'—'}</td></tr>`;
  html += '</table></div>';
  html += sH('Remarks / 备注') + `<textarea class="qc-input" style="text-align:left;font-weight:400;height:50px;resize:vertical;background:#fff" onchange="setD('remarks',this.value);scheduleSave()">${data.remarks||''}</textarea></div>`;
  document.getElementById('report-body').innerHTML = html;
}

// ── IMAGE UPLOAD (Ferroxyl only) ──
async function loadImages(){const r=await fetch(`/api/project/${P}/spool/${S}/qc/${RT}/images`);const imgs=await r.json();const grid=document.getElementById('img-grid');if(!grid)return;grid.innerHTML=imgs.map(img=>`<div class="img-thumb"><img src="/api/project/${P}/qc/image/${img.id}" loading="lazy"><button class="del" onclick="delImage(${img.id})">×</button></div>`).join('');}
async function uploadImages(files){for(const f of files){const c=await compressImage(f,2000,0.85);const fd=new FormData();fd.append('file',c,f.name);fd.append('operator',document.getElementById('inspector').value);await fetch(`/api/project/${P}/spool/${S}/qc/${RT}/image`,{method:'POST',body:fd});}loadImages();}
async function delImage(id){if(!confirm('Delete? / 删除？'))return;await fetch(`/api/project/${P}/qc/image/${id}`,{method:'DELETE'});loadImages();}
function compressImage(file,maxDim,quality){return new Promise(r=>{const img=new Image();img.onload=()=>{let w=img.width,h=img.height;if(w>maxDim||h>maxDim){const ratio=Math.min(maxDim/w,maxDim/h);w=Math.round(w*ratio);h=Math.round(h*ratio);}const c=document.createElement('canvas');c.width=w;c.height=h;c.getContext('2d').drawImage(img,0,0,w,h);c.toBlob(b=>r(b),'image/jpeg',quality);};img.src=URL.createObjectURL(file);});}


// ── Inspector Registry + Signature Pad ──
let inspectorList = [];
let sigCtx = null, sigDrawing = false;
async function loadInspectors(){
  try{const r=await fetch('/api/inspectors');inspectorList=await r.json();}catch(e){}
  const sel=document.getElementById('inspector-select');
  const current=document.getElementById('inspector').value;
  sel.innerHTML='<option value="">— Select / 选择检验员 —</option>';
  inspectorList.forEach(ins=>{
    sel.innerHTML+=`<option value="${ins.name}" ${ins.name===current?'selected':''}>${ins.name}${ins.role?' ('+ins.role+')':''}${ins.has_signature?' ✓':''}</option>`;
  });
  sel.innerHTML+='<option value="__new__">+ New inspector / 新增检验员</option>';
  sel.innerHTML+='<option value="__delete__">- Delete inspector / 删除检验员</option>';
  if(current && !inspectorList.find(i=>i.name===current)){
    sel.innerHTML+=`<option value="${current}" selected>${current}</option>`;
  }
}
async function onInspectorSelect(val){
  if(val==='__new__'){
    const name=prompt('Inspector name / 检验员姓名:');
    if(!name){document.getElementById('inspector-select').value='';return;}
    await fetch('/api/inspectors',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name:name})});
    document.getElementById('inspector').value=name;
    await loadInspectors();
    document.getElementById('inspector-select').value=name;
    showSigPad(name);
    scheduleSave();
    return;
  }
  if(val==='__delete__'){
    const names=inspectorList.map(i=>`${i.id}: ${i.name}`).join('\\n');
    const id=prompt('Enter inspector number to delete / 输入要删除的检验员编号:\\n'+names);
    if(id){await fetch(`/api/inspectors/${parseInt(id)}`,{method:'DELETE'});}
    await loadInspectors();
    document.getElementById('inspector-select').value='';
    return;
  }
  document.getElementById('inspector').value=val;
  if(val) showSigPad(val);
  else document.getElementById('sig-section').style.display='none';
  scheduleSave();
}
async function showSigPad(name){
  const section=document.getElementById('sig-section');
  section.style.display='block';
  const ins=inspectorList.find(i=>i.name===name);
  if(ins && ins.has_signature){
    // Load existing signature
    const r=await fetch(`/api/inspectors/${ins.id}/signature`);
    const d=await r.json();
    if(d.signature_data){
      document.getElementById('sig-preview').src=d.signature_data;
      document.getElementById('sig-preview').style.display='block';
      document.getElementById('sig-canvas').style.display='none';
      return;
    }
  }
  // Show canvas for new signature
  document.getElementById('sig-preview').style.display='none';
  document.getElementById('sig-canvas').style.display='block';
  initSigCanvas();
}
function initSigCanvas(){
  const canvas=document.getElementById('sig-canvas');
  const container=document.getElementById('sig-container');
  canvas.width=container.clientWidth;
  canvas.height=container.clientHeight;
  sigCtx=canvas.getContext('2d');
  sigCtx.strokeStyle='#2F5496';sigCtx.lineWidth=2;sigCtx.lineCap='round';
  canvas.onpointerdown=e=>{sigDrawing=true;sigCtx.beginPath();sigCtx.moveTo(e.offsetX,e.offsetY);};
  canvas.onpointermove=e=>{if(!sigDrawing)return;sigCtx.lineTo(e.offsetX,e.offsetY);sigCtx.stroke();};
  canvas.onpointerup=()=>{sigDrawing=false;};
  canvas.onpointerleave=()=>{sigDrawing=false;};
}
function clearSig(){
  const canvas=document.getElementById('sig-canvas');
  sigCtx.clearRect(0,0,canvas.width,canvas.height);
  document.getElementById('sig-preview').style.display='none';
  canvas.style.display='block';
}
async function saveSig(){
  const canvas=document.getElementById('sig-canvas');
  const dataUrl=canvas.toDataURL('image/png');
  const name=document.getElementById('inspector').value;
  if(!name){alert('Select inspector first / 先选择检验员');return;}
  await fetch('/api/inspectors',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({name:name,signature_data:dataUrl})});
  document.getElementById('sig-preview').src=dataUrl;
  document.getElementById('sig-preview').style.display='block';
  canvas.style.display='none';
  await loadInspectors();
  showSave('Signature saved / 签名已保存',true);
}

async function loadReport(){
  const r = await fetch(`/api/project/${P}/spool/${S}/qc/${RT}?sub=${SUB}`);
  reportData = await r.json();
  if(!reportData.data || typeof reportData.data !== 'object') reportData.data = {};
  const inspName = reportData.inspector_name || localStorage.getItem('qc_inspector') || '';
  document.getElementById('inspector').value = inspName;
  // Report date = checklist step completion date, NOT today
  document.getElementById('insp-date').value = reportData.inspector_date || reportData.step_date || '';
  document.getElementById('tpi').value = reportData.tpi_name || '';
  document.getElementById('tpi-date').value = reportData.tpi_date || '';
  await loadInspectors();
  if(inspName) showSigPad(inspName);
  const cfg = REPORT_FIELDS[RT];
  if(cfg){
    document.getElementById('rpt-title').textContent = cfg.title;
    const recNo = reportData.rec_no || '';
    const stepDate = reportData.step_date || '';
    const itpDoc = reportData.itp || P+'-ITP-SPL-001';
    MATERIAL_TYPE = reportData.material_type || '';
    document.getElementById('rpt-sub').innerHTML = `${S} · ${P}${recNo ? ' · <span style="font-family:monospace;font-size:10px;color:#2F5496">'+recNo+'</span>' : ''}${stepDate ? ' · <span style="font-size:10px">Date / 日期: '+stepDate+'</span>' : ''}`;
    document.getElementById('itp-display').textContent = itpDoc;
    cfg.render(reportData.data);
  } else {
    document.getElementById('report-body').innerHTML = '<p style="padding:20px;color:#888">Report type not configured</p>';
  }
  updateResult(reportData.data.overall_result || null);
}
function updateResult(val){
  document.getElementById('res-pass').className = 'pf-btn' + (val==='ACC'?' sel-pass':'');
  document.getElementById('res-fail').className = 'pf-btn' + (val==='REJ'?' sel-fail':'');
}
function setResult(val){
  const cur = reportData.data.overall_result;
  reportData.data.overall_result = (cur===val) ? null : val;
  updateResult(reportData.data.overall_result);
  scheduleSave();
}
function scheduleSave(){ clearTimeout(saveTimer); saveTimer = setTimeout(doSave, 600); }
async function doSave(){
  const insp = document.getElementById('inspector').value;
  localStorage.setItem('qc_inspector', insp);
  const body = {
    report_subtype: SUB, status: 'draft',
    inspector_name: insp, inspector_date: document.getElementById('insp-date').value,
    tpi_name: document.getElementById('tpi').value, tpi_date: document.getElementById('tpi-date').value,
    data: reportData.data,
  };
  try{
    const r = await fetch(`/api/project/${P}/spool/${S}/qc/${RT}`, {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(body)});
    const d = await r.json();
    showSave(d.ok ? 'Saved / 已保存' : 'Error', d.ok);
  }catch(e){ showSave('Error', false); }
}
function showSave(msg, ok){
  const el = document.getElementById('save-status');
  el.textContent = msg; el.className = 'save-status ' + (ok ? 'save-ok' : 'save-err');
  el.style.opacity = '1'; setTimeout(()=>{ el.style.opacity = '0'; }, 1500);
}
document.querySelectorAll('.inspector-bar input').forEach(el => el.addEventListener('change', scheduleSave));

loadReport();
</script></body></html>"""

REPORT_HTML = """<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{{ project }} — Report</title><style>""" + COMMON_CSS + """
.report-card{background:#fff;border-radius:12px;padding:16px;margin:12px 16px;box-shadow:0 2px 8px rgba(0,0,0,.08)}
.report-card h3{font-size:16px;color:#2F5496;margin-bottom:12px}
.status-badge{display:inline-block;padding:6px 14px;border-radius:20px;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:.5px}
.status-on_time{background:#e8f5e9;color:#27ae60}.status-at_risk{background:#fff3e0;color:#f39c12}.status-delayed{background:#fce4ec;color:#e74c3c}.status-not_started{background:#f5f5f5;color:#95a5a6}
.status-completed{background:#e8f8f5;color:#16a085}
.status-ready{background:#e3f2fd;color:#1976d2}
.status-partial{background:#fff3e0;color:#f39c12}
.status-shipped{background:#e8f5e9;color:#27ae60}
.status-pending{background:#f5f5f5;color:#7f8c8d}
.status-unassigned{background:#fafafa;color:#bdbdbd}
/* Post-production card */
.pp-dual{display:grid;grid-template-columns:repeat(auto-fit,minmax(320px,1fr));gap:16px}
.pp-gate{display:flex;align-items:center;gap:16px;padding:14px;background:#fafbfc;border-radius:10px;border:1px solid #ecf0f1}
.pp-ring{width:104px;height:104px;position:relative;flex-shrink:0}
.pp-ring svg{transform:rotate(-90deg)}
.pp-ring-center{position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center;text-align:center}
.pp-ring-num{font-size:22px;font-weight:700;line-height:1}
.pp-ring-sep{font-size:9px;color:#bbb;margin:2px 0}
.pp-ring-tot{font-size:11px;color:#7f8c8d;font-weight:600}
.pp-body{flex:1;min-width:0}
.pp-title{font-size:13px;font-weight:700;color:#2c3e50}
.pp-title-cn{font-size:11px;color:#7f8c8d;margin-bottom:6px}
.pp-bar{background:#ecf0f1;border-radius:6px;height:8px;overflow:hidden}
.pp-bar-fill{height:100%;border-radius:6px}
.pp-caption{font-size:11px;color:#95a5a6;margin-top:4px}
/* Shipments card */
.ship-list{display:flex;flex-direction:column;gap:10px}
.ship-row{padding:12px 14px;background:#fafbfc;border-radius:10px;border-left:4px solid #bdc3c7;display:grid;grid-template-columns:1fr auto;gap:12px;align-items:center}
.ship-row.pending{border-left-color:#bdc3c7}
.ship-row.ready{border-left-color:#1976d2}
.ship-row.partial{border-left-color:#f39c12}
.ship-row.shipped{border-left-color:#27ae60}
.ship-row.unassigned{border-left-color:#e0e0e0}
.ship-head{font-size:13px;font-weight:700;color:#2c3e50}
.ship-meta{font-size:11px;color:#888;margin-top:2px}
.ship-counts{display:flex;gap:14px;margin-top:6px;font-size:11px}
.ship-counts .lbl{color:#95a5a6}
.ship-counts .val{color:#2c3e50;font-weight:700}
.ship-counts .val.done{color:#27ae60}
.ship-counts .val.pack{color:#16a085}
.ship-bars{display:flex;gap:4px;margin-top:6px}
.ship-bar{flex:1}
.ship-bar .lab{font-size:8px;color:#aaa;margin-bottom:2px}
.ship-bar-track{background:#ecf0f1;height:5px;border-radius:3px;overflow:hidden}
.ship-bar-fill{height:100%;border-radius:3px}
.ship-badge-wrap{text-align:right}
.ship-date{font-size:10px;color:#888;margin-top:4px}
.ship-actions button{font-size:11px;color:#fff;background:#16a085;border:none;padding:5px 10px;border-radius:6px;margin-top:6px;cursor:pointer}
.ship-actions button:hover{background:#0e6655}
.diam-status{display:flex;flex-direction:column;gap:8px}
.diam-row{display:flex;align-items:center;gap:12px;padding:8px 12px;background:#FAFBFD;border-radius:8px}
.d-name{font-size:20px;font-weight:700;color:#2F5496;min-width:50px}.d-info{flex:1}.d-status{min-width:100px;text-align:right}
.gantt-mini{overflow-x:auto}
.gantt-table{border-collapse:collapse;width:100%;font-size:10px}
.gantt-table th{background:#404040;color:#fff;padding:4px 3px;font-size:9px;white-space:nowrap;text-align:center;min-width:45px}
.gantt-table td{padding:0;height:28px;position:relative;border:1px solid #f0f2f5;text-align:center;min-width:45px}
.gantt-table .g-label{font-weight:600;padding:4px 6px;text-align:left;white-space:nowrap;background:#fff;border-right:2px solid #e8e8e8}
.g-bar{position:absolute;top:2px;left:2px;right:2px;bottom:2px;border-radius:3px}
.g-exp-fab{background:#4472C4}.g-exp-paint{background:#ED7D31}.g-saved{background:#E2EFDA;border:1px dashed #A9D18E}
.g-forecast{background:transparent;border:2px dashed #e74c3c}
.g-today-line{position:absolute;left:50%;top:0;bottom:0;width:3px;background:#e74c3c;transform:translateX(-50%);z-index:10}
.g-pct{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;font-size:8px;font-weight:700;color:#fff;z-index:5}
.wk-current{background:#C00000!important;color:#fff!important}
.wk-dates{font-size:7px;font-weight:400;opacity:.7}
.mini-prog{width:100%;height:4px;background:#e8e8e8;border-radius:2px;margin-top:4px;overflow:hidden}
.mini-prog-fill{height:100%;border-radius:2px;background:#4472C4}
.act-item{padding:8px 12px;border-bottom:1px solid #f0f2f5;display:flex;gap:10px;align-items:center;font-size:13px}
.act-item .time{color:#888;font-size:11px;min-width:55px}.act-item .spool{font-weight:600;color:#2F5496;min-width:70px}
.act-item .detail{flex:1;color:#555}
.legend{display:flex;gap:12px;margin-top:10px;flex-wrap:wrap;font-size:10px;color:#666}
.legend span{display:flex;align-items:center;gap:4px}
.legend .box{width:14px;height:10px;border-radius:2px;display:inline-block}
.expected-bar{background:#e8e8e8;border-radius:6px;height:8px;position:relative;overflow:visible;margin-top:4px}
.expected-fill{position:absolute;left:0;top:0;height:100%;border-radius:6px;opacity:0.3;background:#2F5496}
.actual-fill{position:absolute;left:0;top:0;height:100%;border-radius:6px}
.commit-panel{background:#fff;border-radius:10px;padding:14px 18px;margin:12px 16px;box-shadow:0 2px 8px rgba(0,0,0,.08);display:flex;align-items:center;gap:16px;flex-wrap:wrap;border-left:4px solid #2F5496}
.commit-panel h4{font-size:12px;color:#2F5496;margin:0;white-space:nowrap}
.cp-item{text-align:center;padding:0 12px;border-right:1px solid #eee}.cp-item:last-child{border-right:none}
.cp-label{font-size:8px;color:#888;text-transform:uppercase}.cp-val{font-size:16px;font-weight:700}
.results-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin:12px 16px}
.res-card{background:#fff;border-radius:12px;padding:20px 16px;text-align:center;border:2px dashed #d0d8e8}
.res-card.rc-blue{border-color:#4472C4;border-top:4px solid #2F5496}.res-card.rc-green{border-color:#27ae60;border-top:4px solid #27ae60}.res-card.rc-dark{border-color:#1F4E79;border-top:4px solid #1F4E79}
.res-card h5{font-size:11px;color:#666;text-transform:uppercase;letter-spacing:1px;font-weight:700;margin:0 0 8px}
.res-card .rv{font-size:28px;font-weight:700}.res-card .rs{font-size:11px;color:#999;margin-top:6px;line-height:1.4}
.res-full{grid-column:1/-1}
@media(max-width:768px){.results-grid{grid-template-columns:repeat(2,1fr)}}
.transit-strip{background:#fff;border-radius:10px;padding:12px 18px;margin:10px 16px;box-shadow:0 1px 4px rgba(0,0,0,.06);display:flex;align-items:center;gap:16px;flex-wrap:wrap;border-left:4px solid #003366}
.sa-card{border:1px solid #f0f0f0;border-radius:8px;margin-bottom:6px;overflow:hidden}
.sa-header{display:flex;align-items:center;padding:8px 12px;gap:10px;cursor:pointer;user-select:none}
.sa-header:hover{background:#FAFBFD}
.sa-spool{font-weight:700;color:#2F5496;font-size:12px;min-width:65px}
.sa-dia{font-size:9px;color:#888;background:#f0f2f5;padding:1px 6px;border-radius:8px}
.sa-prog{width:60px;height:5px;background:#e8e8e8;border-radius:3px;overflow:hidden}
.sa-prog-fill{height:100%;border-radius:3px}
.sa-range{flex:1;font-size:11px;color:#555}.sa-range strong{color:#2F5496}
.sa-by{font-size:10px;color:#888}.sa-time{font-size:9px;color:#aaa;min-width:40px;text-align:right}
.sa-expand{font-size:9px;color:#ccc;transition:transform .2s}
.sa-card.open .sa-expand{transform:rotate(90deg)}
.sa-milestone{font-size:9px;font-weight:700;padding:2px 7px;border-radius:8px}
.sa-ms-rt{background:#EDE7F6;color:#5E35B1}.sa-ms-done{background:#E8F5E9;color:#2E7D32}
.sa-detail{display:none;border-top:1px solid #f5f5f5;background:#FAFBFD}
.sa-card.open .sa-detail{display:block}
.sa-drow{display:flex;align-items:center;padding:5px 12px 5px 36px;gap:8px;font-size:11px;border-bottom:1px solid #f5f5f5}
.sa-drow:last-child{border-bottom:none}
.act-summary{display:flex;background:#F8F9FB;border-radius:8px;padding:8px 0;margin-bottom:12px;flex-wrap:wrap}
.as-item{flex:1;text-align:center;padding:4px 12px;min-width:80px}
.as-item:not(:last-child){border-right:1px solid #e8e8e8}
.as-val{font-size:18px;font-weight:700}.as-lbl{font-size:9px;color:#888}
.summary-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:10px;margin:12px 0}
.sum-card{text-align:center;padding:12px;border-radius:8px;background:#f8f9fa}
.sum-card .v{font-size:24px;font-weight:700}.sum-card .l{font-size:11px;color:#888;margin-top:2px}
@media print{.header{background:#2F5496!important;-webkit-print-color-adjust:exact;print-color-adjust:exact}.no-print{display:none!important}}
</style></head><body>
<div class="header">
  <a class="back no-print" href="/project/{{ project }}">\u2190 Back / \u8fd4\u56de</a>
  <h1>Production Report / \u751f\u4ea7\u62a5\u544a</h1>
  <div class="sub">{{ project }} \u2014 <span id="rpt-date"></span></div>
</div>
<div id="rpt-content"><div style="text-align:center;padding:40px;color:#888">Loading report... / \u52a0\u8f7d\u62a5\u544a\u4e2d...</div></div>
<div style="padding:16px;text-align:center" class="no-print">
  <a class="btn" href="/api/project/{{ project }}/report/download">\U0001f4e5 Excel Report / Excel\u62a5\u544a</a>
  <a class="btn" href="/api/project/{{ project }}/report/pdf" style="margin-left:8px;background:#E74C3C">\U0001f4c4 PDF Report / PDF\u62a5\u544a</a>
  <button class="btn" onclick="window.print()" style="margin-left:8px">\U0001f5a8 Print / \u6253\u5370</button>
</div>
<script>
const P='{{project}}';
const STATUS_LABELS = {on_time:'ON TIME / \u6309\u65f6',at_risk:'AT RISK / \u6709\u5ef6\u8fdf\u98ce\u9669',delayed:'DELAYED / \u5df2\u5ef6\u8fdf',not_started:'NOT STARTED / \u672a\u5f00\u59cb',completed:'\u2713 COMPLETED / \u5df2\u5b8c\u6210'};
const STATUS_COLORS = {on_time:'#27ae60',at_risk:'#f39c12',delayed:'#e74c3c',not_started:'#95a5a6',completed:'#16a085'};
const SHIP_STATUS_LABELS = {pending:'\u25cf PENDING / \u5f85\u52a0\u5de5',ready:'\u25cf READY TO SHIP / \u5f85\u53d1\u8fd0',partial:'\u25cf PARTIAL / \u90e8\u5206\u53d1\u8fd0',shipped:'\u2713 SHIPPED / \u5df2\u53d1\u8fd0',unassigned:'UNASSIGNED / \u672a\u5206\u914d'};
async function load(){
  const [r, shipR] = await Promise.all([fetch(`/api/project/${P}/report`), fetch(`/api/project/${P}/shipments`)]);
  const d = await r.json();
  const shipments = await shipR.json();
  document.getElementById('rpt-date').textContent = d.date;
  // Dynamic hold/release step detection from API
  const holdStepSet = new Set(d.hold_steps || []);
  const releaseStepSet = new Set(d.release_steps || []);
  let html = '';
  const st = d.stats, sch = d.schedule, sett = d.settings||{}, fc = d.forecast||{};
  const phases = d.phase_order || sch?.phase_order || ['fab'];
  const phaseColors = ['#4472C4', '#ED7D31', '#8E44AD', '#27AE60'];
  const overallStatus = sch ? sch.overall_status : 'not_started';
  const stdWeeks = parseInt(sett.standard_weeks||'9');
  const wksSaved = parseInt(sett.committed_weeks_saved||'0');
  const daysSaved = parseInt(sett.committed_days_saved||'0');
  const totalSaved = wksSaved*7+daysSaved;
  const hasExpediting = totalSaved > 0;
  const transitDays = parseInt(sett.sea_transit_days||'45');
  const toLocal = s => {const [y,m,d]=s.split('-');return new Date(+y,+m-1,+d);};
  const addDays = (dt,n) => new Date(dt.getFullYear(),dt.getMonth(),dt.getDate()+n);
  const fmt = dt => dt.toLocaleDateString('en',{day:'2-digit',month:'short',year:'numeric'});
  const fmtShort = dt => dt.toLocaleDateString('en',{day:'2-digit',month:'short'});

  // Finish date = max forecast_end across diameters when overall is completed
  let finishDate = null;
  if(overallStatus === 'completed' && fc.overall_forecast_end) finishDate = fc.overall_forecast_end;
  const finishLine = (overallStatus === 'completed' && finishDate)
    ? `<div style="font-size:12px;color:#16a085;margin-top:6px;font-weight:600">Production finished ${finishDate} / \u751f\u4ea7\u5b8c\u5de5\u65e5\u671f</div>`
    : '';
  html += `<div class="report-card">
    <div style="display:flex;align-items:center;gap:16px;flex-wrap:wrap">
      <div><h3 style="margin:0">Overall Status / \u603b\u4f53\u72b6\u6001</h3>
        <div class="status-badge status-${overallStatus}" style="margin-top:8px">${STATUS_LABELS[overallStatus]||overallStatus}</div>
        ${finishLine}</div>
      <div style="flex:1;min-width:200px"><div class="summary-grid">
        <div class="sum-card"><div class="v" style="color:${overallStatus==='completed'?'#16a085':'#2F5496'}">${st.overall_pct}%</div><div class="l">Progress / \u8fdb\u5ea6</div></div>
        <div class="sum-card"><div class="v">${st.total}</div><div class="l">Total / \u603b\u6570</div></div>
        <div class="sum-card"><div class="v" style="color:#27ae60">${st.completed}</div><div class="l">Done / \u5b8c\u6210</div>${d.past_rt?`<div style="font-size:11px;color:#4472C4;margin-top:2px"><strong>${d.past_rt}</strong> past RT / \u5df2\u8fc7RT</div>`:''}</div>
        <div class="sum-card"><div class="v" style="color:#f39c12">${st.in_progress}</div><div class="l">WIP / \u8fdb\u884c\u4e2d</div></div>
        <div class="sum-card"><div class="v" style="color:#e74c3c">${st.not_started}</div><div class="l">Pending / \u5f85\u5f00\u59cb</div></div>
      </div></div>
    </div></div>`;

  if(sch && sch.diameters && sch.diameters.length){
    html += `<div class="report-card"><h3>Schedule Status by Diameter / \u6309\u7ba1\u5f84\u8ba1\u5212\u72b6\u6001</h3><div class="diam-status">`;
    const fcDiams = fc.diameters || {};
    sch.diameters.forEach(dm => {
      const color = STATUS_COLORS[dm.status] || '#95a5a6';
      const fdi = fcDiams[dm.diameter] || {};
      const dmPhaseAvgs = dm.phase_avgs || fdi.phase_avgs || {};
      const phaseInfo = phases.map((ph,pi) => `${ph.charAt(0).toUpperCase()+ph.slice(1)}: <strong>${dmPhaseAvgs[ph]||0}%</strong>`).join(' \u00b7 ');
      const phaseBars = phases.map((ph,pi) => `<div style="flex:1"><div style="font-size:8px;color:#aaa">${ph.charAt(0).toUpperCase()+ph.slice(1)}</div><div class="expected-bar"><div class="actual-fill" style="width:${dmPhaseAvgs[ph]||0}%;background:${phaseColors[pi%phaseColors.length]};border-radius:6px"></div></div></div>`).join('');
      html += `<div class="diam-row" style="border-left:4px solid ${color}">
        <div class="d-name">${dm.diameter}</div>
        <div class="d-info">
          <div style="font-size:12px;color:#888">${dm.spool_count} spools \u00b7 ${phaseInfo} \u00b7 Overall / \u603b: <strong>${dm.actual_pct}%</strong></div>
          <div style="display:flex;gap:4px;margin-top:6px">${phaseBars}</div>
          <div style="font-size:10px;color:#aaa;margin-top:2px">${fdi.remaining_raf?'RAF: '+fdi.remaining_raf+' in':''} ${fdi.remaining_m2?'\u00b7 Surface: '+fdi.remaining_m2+' m\u00b2':''}</div>
        </div>
        <div class="d-status"><span class="status-badge status-${dm.status}" style="font-size:11px;padding:4px 10px">${STATUS_LABELS[dm.status]||dm.status}</span></div>
      </div>`;
    });
    html += `</div></div>`;

    let prodStart = null;
    const starts = sch.diameters.map(x=>x.total_start).filter(x=>x).sort();
    if(starts.length) prodStart = starts[0];
    if(prodStart){
      const psDate = toLocal(prodStart);
      const stdEnd = addDays(psDate, stdWeeks*7 - 1);
      const commitEnd = hasExpediting ? addDays(stdEnd, -totalSaved) : stdEnd;
      const fcEnd = fc.overall_forecast_end ? toLocal(fc.overall_forecast_end) : null;
      const today = new Date(); today.setHours(0,0,0,0);

      if(hasExpediting){
        const diffDays = fcEnd ? Math.ceil((commitEnd - fcEnd) / 86400000) : 0;
        html += `<div class="commit-panel"><h4>\u26a1 Expediting Commitment / \u52a0\u6025\u627f\u8bfa</h4>
          <div class="cp-item"><div class="cp-label">Start / \u5f00\u59cb</div><div class="cp-val" style="color:#2F5496">${fmtShort(psDate)}</div></div>
          <div class="cp-item"><div class="cp-label">Standard End / \u6807\u51c6\u5b8c\u5de5</div><div class="cp-val" style="color:#888;text-decoration:line-through">${fmtShort(stdEnd)}</div></div>
          <div class="cp-item"><div class="cp-label">Committed End / \u627f\u8bfa\u5b8c\u5de5</div><div class="cp-val" style="color:#4472C4">${fmtShort(commitEnd)}</div></div>
          <div class="cp-item"><div class="cp-label">Saved / \u8282\u7701</div><div class="cp-val" style="color:#4472C4">${totalSaved}d</div></div>
          <div class="cp-item" style="border-right:none"><div class="cp-label">Forecast / \u9884\u6d4b</div><div class="cp-val" style="color:${diffDays>=0?'#27ae60':'#e74c3c'}">${fcEnd?fmtShort(fcEnd):'\u2014'} ${diffDays>=0?'\u2713':'\u2717'}</div></div>
        </div>`;
      }

      // Gantt
      const numWeeks = hasExpediting ? stdWeeks : Math.max(stdWeeks, Math.ceil(((fcEnd||stdEnd).getTime() - psDate.getTime()) / 86400000 / 7) + 1);
      html += `<div class="report-card"><h3>Production Gantt / \u751f\u4ea7\u7518\u7279\u56fe</h3>
        <div class="gantt-mini"><table class="gantt-table"><thead><tr>
        <th style="min-width:70px;background:#404040">Diameter</th><th style="min-width:45px;background:#404040">Phase</th>`;
      const weeks = [];
      for(let i=0;i<numWeeks;i++){
        const ws = addDays(psDate, i*7);
        const we = addDays(ws, 6);
        const isCurrent = today>=ws && today<=we;
        weeks.push({start:ws,end:we,num:i+1,current:isCurrent});
        html += `<th${isCurrent?' class="wk-current"':''}>W${i+1}<br><span class="wk-dates">${fmtShort(ws)} \u2192 ${fmtShort(we)}</span></th>`;
      }
      html += `</tr></thead><tbody>`;
      const ratio = hasExpediting ? (stdWeeks - wksSaved) / stdWeeks : 1;
      const lastDiamIdx = sch.diameters.length - 1;


      sch.diameters.forEach((dm, dmIdx) => {
        const fdi = fcDiams[dm.diameter] || {};
        const dmPhaseAvgs = fdi.phase_avgs || dm.phase_avgs || {};
        const overallP = fdi.overall_pct||0;
        const dmFcEnd = fdi.forecast_end ? toLocal(fdi.forecast_end) : null;
        const dmStarted = fdi.started;
        // Parse schedule dates from phase_dates
        const phaseDates = {};
        const dmPD = dm.phase_dates || {};
        phases.forEach(ph => {
          const pd = dmPD[ph];
          if(pd && pd.start && pd.end) phaseDates[ph] = {s:toLocal(pd.start), e:toLocal(pd.end)};
        });
        // Compute expedited dates per phase
        const expDates = {};
        let prevExpEnd = null;
        phases.forEach(ph => {
          if(!phaseDates[ph]) return;
          const {s:phS, e:phE} = phaseDates[ph];
          let expS, expE;
          if(hasExpediting){
            const daysDiffS = Math.round((phS-psDate)/86400000);
            const daysDiffE = Math.round((phE-psDate)/86400000);
            expS = addDays(psDate, Math.round(daysDiffS*ratio));
            expE = addDays(psDate, Math.round(daysDiffE*ratio));
            if(expE > commitEnd) expE = commitEnd;
            if(expS > commitEnd) expS = commitEnd;
            if(prevExpEnd && expS < prevExpEnd) expS = prevExpEnd;
            if(expE < expS) expE = expS;
          } else { expS = phS; expE = phE; }
          expDates[ph] = {s:expS, e:expE};
          prevExpEnd = expE;
        });
        const isLast = dmIdx===lastDiamIdx;
        // One row per phase
        html += `<tr>`;
        html += `<td class="g-label" rowspan="${phases.length}" style="color:#2F5496;font-size:13px"><b>${dm.diameter}</b> <span style="font-size:8px;color:#888;font-weight:400">${dm.spool_count}sp · ${overallP}%</span></td>`;
        phases.forEach((ph, pi) => {
          if(pi > 0) html += `<tr>`;
          const phP = dmPhaseAvgs[ph]||0;
          const phColor = phaseColors[pi % phaseColors.length];
          html += `<td class="g-label" style="font-size:9px;color:#666">${ph.charAt(0).toUpperCase()+ph.slice(1)}</td>`;
          const isLastPhase = pi === phases.length - 1;
          weeks.forEach(w => {
            const isToday = w.current;
            let content = '';
            if(phaseDates[ph] && expDates[ph]){
              const {s:phS, e:phE} = phaseDates[ph];
              const {s:expS, e:expE} = expDates[ph];
              const inStd = phS<=w.end && phE>=w.start;
              const inExp = expS<=w.end && expE>=w.start;
              const isSaved = hasExpediting && inStd && !inExp;
              const isForecast = isLastPhase && dmStarted && dmFcEnd && overallP<100 && dmFcEnd>=w.start && dmFcEnd<=w.end;
              if(inExp){ content = `<div class="g-bar" style="background:${phColor};position:absolute;top:2px;left:2px;right:2px;bottom:2px;border-radius:3px"></div>`;
                if(isToday) content += `<div class="g-today-line"></div><div class="g-pct">${phP}%</div>`;
                else if(w.end < today) content += `<div class="g-pct" style="color:#fff">\u2713</div>`;
              } else if(isSaved){ content = `<div class="g-bar g-saved"></div>`; }
              if(isForecast) content += `<div class="g-bar g-forecast"></div>`;
              if(isToday && !inExp){
                if(phP > 0 && phP < 100) content = `<div class="g-bar" style="background:${phColor};position:absolute;top:2px;left:2px;right:2px;bottom:2px;border-radius:3px"></div><div class="g-today-line"></div><div class="g-pct">${phP}%</div>`;
                else content += `<div class="g-today-line"></div>`;
              }
            } else {
              if(isToday) content += `<div class="g-today-line"></div>`;
            }
            if(isToday && isLast && isLastPhase) content += `<div style="position:absolute;bottom:-13px;left:50%;transform:translateX(-50%);font-size:7px;color:#e74c3c;font-weight:700;z-index:11">TODAY</div>`;
            html += `<td>${content}</td>`;
          });
          html += `</tr>`;
        });
        html += `<tr><td colspan="${numWeeks+2}" style="height:2px;background:#f0f2f5;border:none"></td></tr>`;
      });

      html += `</tbody></table></div>
        <div class="legend">
          ${phases.map((ph,pi) => `<span><span class="box" style="background:${phaseColors[pi%phaseColors.length]}"></span> ${ph.charAt(0).toUpperCase()+ph.slice(1)}</span>`).join('')}
          ${hasExpediting?'<span><span class="box" style="background:#E2EFDA;border:1px solid #A9D18E"></span> Saved / \u8282\u7701 \u2713</span>':''}
          <span><span class="box" style="border:2px dashed #e74c3c;width:12px;height:8px;display:inline-block;border-radius:2px"></span> Forecast / \u9884\u6d4b</span>
          <span><span class="box" style="background:#e74c3c;width:3px"></span> Today / \u4eca\u5929</span>
        </div></div>`;

      // Rate
      const actualWeld = fc.actual_weld_ipd||0, actualPaint = fc.actual_paint_m2d||0;
      const weldCap = fc.welding_capability||0, paintCap = fc.painting_capability||0;
      html += `<div class="report-card"><h3>Production Rate / \u751f\u4ea7\u7387</h3>
        <div style="display:flex;gap:20px;flex-wrap:wrap">
          <div style="flex:1;min-width:200px">
            <div style="font-size:12px;color:#888;margin-bottom:4px">Welding / \u710a\u63a5 (linear inches/day)</div>
            <div style="display:flex;align-items:baseline;gap:8px"><span style="font-size:24px;font-weight:700;color:${actualWeld>=weldCap?'#27ae60':'#e74c3c'}">${actualWeld}</span><span style="color:#888;font-size:12px">/ ${weldCap} target</span></div>
            <div class="expected-bar" style="margin-top:4px"><div class="actual-fill" style="width:${Math.min(actualWeld/Math.max(weldCap,1)*100,100)}%;background:${actualWeld>=weldCap?'#27ae60':'#e74c3c'}"></div></div>
          </div>
          ${paintCap > 0 ? `<div style="flex:1;min-width:200px">
            <div style="font-size:12px;color:#888;margin-bottom:4px">${phases.length>1?phases[phases.length-1].charAt(0).toUpperCase()+phases[phases.length-1].slice(1):'Surface'} / \u6d82\u88c5 (m\u00b2/day)</div>
            <div style="display:flex;align-items:baseline;gap:8px"><span style="font-size:24px;font-weight:700;color:${actualPaint>=paintCap?'#27ae60':'#e74c3c'}">${actualPaint}</span><span style="color:#888;font-size:12px">/ ${paintCap} target</span></div>
            <div class="expected-bar" style="margin-top:4px"><div class="actual-fill" style="width:${Math.min(actualPaint/Math.max(paintCap,1)*100,100)}%;background:${actualPaint>=paintCap?'#27ae60':'#e74c3c'}"></div></div>
          </div>` : ''}
        </div></div>`;

      // Results
      const fcSaved = fcEnd ? Math.ceil((stdEnd - fcEnd) / 86400000) : 0;
      const fcDiffCommit = fcEnd && hasExpediting ? Math.ceil((commitEnd - fcEnd) / 86400000) : 0;
      html += `<div class="results-grid">
        <div class="res-card rc-blue"><h5>Overall Progress / \u603b\u8fdb\u5ea6</h5><div class="rv" style="color:#2F5496">${st.overall_pct}%</div><div class="rs">${st.total} spools \u00b7 ${st.in_progress} WIP \u00b7 ${st.not_started} pending</div></div>
        ${hasExpediting?`<div class="res-card rc-blue"><h5>Production End / \u751f\u4ea7\u5b8c\u5de5</h5>
          <div style="display:flex;align-items:center;justify-content:center;gap:8px;margin:4px 0">
            <span style="font-size:16px;color:#888;text-decoration:line-through">${fmtShort(stdEnd)}</span><span style="color:#888">\u2192</span>
            <span style="font-size:22px;font-weight:700;color:#4472C4">${fmtShort(commitEnd)}</span>
          </div><div class="rs">Standard \u2192 Committed (Expediting) / \u6807\u51c6\u2192\u627f\u8bfa</div></div>
        <div class="res-card rc-green"><h5>Expediting Commitment / \u52a0\u6025\u627f\u8bfa</h5><div class="rv" style="color:#27ae60">${totalSaved}<span style="font-size:14px;font-weight:400"> days</span></div><div class="rs">saved with expediting fee (${wksSaved} weeks) / \u52a0\u6025\u8282\u7701</div></div>
        <div class="res-card rc-green"><h5>Production Forecast / \u751f\u4ea7\u9884\u6d4b</h5><div class="rv" style="color:#27ae60">${fcSaved}<span style="font-size:14px;font-weight:400"> days</span></div><div class="rs">predicted savings \u00b7 ends ${fcEnd?fmtShort(fcEnd):'\u2014'}${fcDiffCommit>=0?' \u00b7 '+fcDiffCommit+' days ahead of commitment':''}</div></div>`
        :`<div class="res-card rc-blue"><h5>Forecast End / \u9884\u6d4b\u5b8c\u5de5</h5><div class="rv" style="color:#4472C4">${fcEnd?fmtShort(fcEnd):'\u2014'}</div><div class="rs">Based on actual rate / \u57fa\u4e8e\u5b9e\u9645\u8fdb\u5ea6</div></div>`}
        <div class="res-card rc-dark res-full"><h5>Actual End / \u5b9e\u9645\u5b8c\u5de5</h5><div class="rv" style="color:#2F5496">${st.completed>=st.total?fmtShort(new Date()):'\u2014'}</div><div class="rs">${st.completed>=st.total?'Production complete / \u751f\u4ea7\u5b8c\u6210':'Shown when production completes / \u751f\u4ea7\u5b8c\u6210\u540e\u663e\u793a'}</div></div>
      </div>`;

      // Transit — from shipments table (last shipment ETA)
      const lastShip = shipments.length ? shipments[shipments.length - 1] : null;
      const lastETA = lastShip ? lastShip.eta : null;
      const lastTransit = lastShip ? lastShip.transit_days : transitDays;
      const lastETD = lastShip && lastShip.etd ? lastShip.etd.substring(0,10) : null;
      html += `<div class="transit-strip">
        <div style="display:flex;align-items:center;gap:6px"><span style="font-size:18px">\U0001f6a2</span><div><div style="font-size:10px;color:#888;text-transform:uppercase">Last Shipment / \u6700\u540e\u4e00\u6279</div><div style="font-size:14px;font-weight:700;color:#003366">${shipments.length ? 'Shipment '+lastShip.shipment_number : '\u2014'}</div></div></div>
        <div style="width:1px;height:28px;background:#e0e0e0"></div>
        <div><div style="font-size:10px;color:#888;text-transform:uppercase">ETD / \u79bb\u6e2f</div><div style="font-size:14px;font-weight:700;color:#003366">${lastETD||'\u2014'}</div></div>
        <div style="width:1px;height:28px;background:#e0e0e0"></div>
        <div><div style="font-size:10px;color:#888;text-transform:uppercase">Transit / \u8fd0\u8f93</div><div style="font-size:14px;font-weight:700;color:#003366">${lastTransit} days</div></div>
        <div style="width:1px;height:28px;background:#e0e0e0"></div>
        <div><div style="font-size:10px;color:#888;text-transform:uppercase">ETA Destination / \u9884\u8ba1\u5230\u8fbe</div><div style="font-size:14px;font-weight:700;color:#27ae60">${lastETA||'\u2014'}</div></div>
      </div>`;
    }
  } else {
    html += `<div class="report-card"><h3>Schedule Status / \u8ba1\u5212\u72b6\u6001</h3>
      <p style="color:#888;padding:20px 0">No schedule configured.<br><code>POST /api/project/${P}/schedule</code></p></div>`;
  }

  // ── Post-Production Gates card (rings per counts_for_production=0 step) ──
  const postProd = d.post_production || [];
  if(postProd.length){
    html += `<div class="report-card"><h3>Post-Production Gates / \u751f\u4ea7\u540e\u5de5\u5e8f</h3><div class="pp-dual">`;
    const ringCircum = 2 * Math.PI * 47; // r=47 in SVG
    const ringColors = ['#16a085','#27ae60','#8E44AD','#2F5496'];
    postProd.forEach((pp, idx) => {
      const col = ringColors[idx % ringColors.length];
      const pct = pp.pct || 0;
      const dashLen = ringCircum * (pct/100);
      const emoji = idx === postProd.length - 1 && postProd.length > 1 ? '\U0001F6A2' : '\U0001F4E6';
      html += `<div class="pp-gate">
        <div class="pp-ring">
          <svg width="104" height="104" viewBox="0 0 104 104">
            <circle cx="52" cy="52" r="47" fill="none" stroke="#ecf0f1" stroke-width="10"/>
            <circle cx="52" cy="52" r="47" fill="none" stroke="${col}" stroke-width="10"
              stroke-dasharray="${dashLen.toFixed(1)} ${ringCircum.toFixed(1)}" stroke-linecap="round"/>
          </svg>
          <div class="pp-ring-center">
            <div class="pp-ring-num" style="color:${col}">${pp.completed}</div>
            <div class="pp-ring-sep">\u2500\u2500 of \u2500\u2500</div>
            <div class="pp-ring-tot">${pp.total}</div>
          </div>
        </div>
        <div class="pp-body">
          <div class="pp-title">${emoji} ${pp.name_en||'Step '+pp.step_number}</div>
          <div class="pp-title-cn">${pp.name_cn||''}</div>
          <div class="pp-bar"><div class="pp-bar-fill" style="width:${pct}%;background:${col}"></div></div>
          <div class="pp-caption">${pp.completed} / ${pp.total} spools \u00b7 ${pct}%${pp.last_date?' \u00b7 last '+pp.last_date:''}</div>
        </div>
      </div>`;
    });
    html += `</div></div>`;
  }

  // ── Shipments card (derived status from assigned spools + post-prod gates) ──
  const shipRows = d.shipment_status || [];
  if(shipRows.length){
    html += `<div class="report-card"><h3>Shipments / \u53d1\u8fd0\u8ba1\u5212 <span style="font-size:10px;font-weight:400;color:#888">(edit details in \u2699 Settings)</span></h3><div class="ship-list">`;
    shipRows.forEach(sh => {
      const showPacked = postProd.length > 1;
      const shippedPct = sh.assigned ? (sh.shipped/sh.assigned*100) : 0;
      const packedPct = sh.assigned ? (sh.packed/sh.assigned*100) : 0;
      const etaLine = sh.etd ? `ETD planned ${sh.etd}${sh.eta?' \u00b7 '+sh.transit_days+'d transit \u2192 ETA '+sh.eta:''}` : 'No ETD set';
      const bars = showPacked
        ? `<div class="ship-bars">
             <div class="ship-bar"><div class="lab">Packed</div><div class="ship-bar-track"><div class="ship-bar-fill" style="width:${packedPct}%;background:#16a085"></div></div></div>
             <div class="ship-bar"><div class="lab">Shipped</div><div class="ship-bar-track"><div class="ship-bar-fill" style="width:${shippedPct}%;background:#27ae60"></div></div></div>
           </div>`
        : `<div class="ship-bars"><div class="ship-bar"><div class="lab">Shipped</div><div class="ship-bar-track"><div class="ship-bar-fill" style="width:${shippedPct}%;background:#27ae60"></div></div></div></div>`;
      const countsLine = showPacked
        ? `<span><span class="lbl">Assigned</span> <span class="val">${sh.assigned}</span></span>
           <span><span class="lbl">Packed</span> <span class="val pack">${sh.packed}/${sh.assigned}</span></span>
           <span><span class="lbl">Shipped</span> <span class="val done">${sh.shipped}/${sh.assigned}</span></span>`
        : `<span><span class="lbl">Assigned</span> <span class="val">${sh.assigned}</span></span>
           <span><span class="lbl">Shipped</span> <span class="val done">${sh.shipped}/${sh.assigned}</span></span>`;
      const dateLine = sh.status === 'shipped' && sh.shipped_date
        ? `<div class="ship-date">Dispatched ${sh.shipped_date}</div>`
        : (sh.etd ? `<div class="ship-date">ETD ${sh.etd}</div>` : '');
      const canMark = sh.status === 'ready' || sh.status === 'partial';
      const actionBtn = canMark
        ? `<div class="ship-actions no-print"><button onclick="markShipped(${sh.shipment_number})">Mark as shipped \u2192</button></div>`
        : '';
      html += `<div class="ship-row ${sh.status}">
        <div>
          <div class="ship-head">\U0001F6A2 SH-${String(sh.shipment_number).padStart(3,'0')}${sh.description?' \u00b7 '+sh.description:''}</div>
          <div class="ship-meta">${sh.assigned} spools \u00b7 ${etaLine}</div>
          <div class="ship-counts">${countsLine}</div>
          ${bars}
        </div>
        <div class="ship-badge-wrap">
          <span class="status-badge status-${sh.status}" style="font-size:10px;padding:4px 10px">${SHIP_STATUS_LABELS[sh.status]||sh.status}</span>
          ${dateLine}
          ${actionBtn}
        </div>
      </div>`;
    });
    html += `</div></div>`;
  }

  // Today's activity — grouped by spool, using dynamic hold/release
  const completed = (d.today_activity||[]).filter(a=>a.action==='completed');
  const stepNames = d.step_names||{};
  const spoolPcts = d.spool_progress||{};
  const groups = {};
  completed.forEach(a => { if(!groups[a.spool_id]) groups[a.spool_id]=[]; groups[a.spool_id].push(a); });
  const spoolKeys = Object.keys(groups);

  html += `<div class="report-card"><h3>Today's Activity / \u4eca\u65e5\u52a8\u6001 <span style="font-weight:400;color:#888;font-size:12px">(${d.steps_completed_today} steps)</span></h3>`;
  html += `<div class="act-summary">
    <div class="as-item"><div class="as-val" style="color:#2F5496">${d.steps_completed_today}</div><div class="as-lbl">Steps / \u6b65\u9aa4</div></div>
    <div class="as-item"><div class="as-val" style="color:#4472C4">${spoolKeys.length}</div><div class="as-lbl">Spools / \u7ba1\u6bb5</div></div>
    <div class="as-item"><div class="as-val" style="color:#27ae60">${d.released_today||0}</div><div class="as-lbl">Released / \u653e\u884c</div></div>
    <div class="as-item"><div class="as-val" style="color:#5E35B1">${d.past_rt||0}</div><div class="as-lbl">Past RT / \u8fc7RT</div></div>
  </div>`;

  if(spoolKeys.length){
    const showLimit = 15;
    spoolKeys.forEach((sid,idx) => {
      const items = groups[sid];
      const pct = spoolPcts[sid]||0;
      const first = items[items.length-1], last = items[0];
      const range = items.length===1 ? (stepNames[first.step_number]||'Step '+first.step_number) : (stepNames[first.step_number]||'Step '+first.step_number)+' \u2192 '+(stepNames[last.step_number]||'Step '+last.step_number);
      const ts = (last.timestamp||'').substring(11,16);
      const hasRT = items.some(a=>holdStepSet.has(a.step_number));
      const hasRelease = items.some(a=>releaseStepSet.has(a.step_number));
      const barColor = pct>=100?'#27ae60':pct>60?'#4472C4':pct>0?'#f39c12':'#ccc';
      const hidden = idx>=showLimit ? ' style="display:none" data-extra' : '';
      html += `<div class="sa-card"${hidden} onclick="this.classList.toggle('open')">
        <div class="sa-header">
          <div class="sa-spool">${sid}</div>
          <div class="sa-prog"><div class="sa-prog-fill" style="width:${pct}%;background:${barColor}"></div></div>
          <div class="sa-range"><strong>${items.length}</strong> \u00b7 ${range}</div>
          ${hasRT?'<div class="sa-milestone sa-ms-rt">\u2605 RT</div>':''}
          ${hasRelease?'<div class="sa-milestone sa-ms-done">\U0001f3c1 RELEASED</div>':''}
          <div class="sa-by">${last.operator||''}</div><div class="sa-time">${ts}</div><div class="sa-expand">\u25b8</div>
        </div>
        <div class="sa-detail">${items.map(a=>{
          const sn=stepNames[a.step_number]||'Step '+a.step_number;
          const t=(a.timestamp||'').substring(11,19);
          const icon=releaseStepSet.has(a.step_number)?'\U0001f3c1':holdStepSet.has(a.step_number)?'\u2b50':'\u2705';
          return `<div class="sa-drow"><span style="color:#aaa;font-size:10px;min-width:50px">${t}</span><span>${icon}</span><span style="flex:1;color:#555">${sn}</span><span style="color:#888;font-size:10px">${a.operator||''}</span></div>`;
        }).join('')}</div></div>`;
    });
    if(spoolKeys.length>showLimit) html += `<div style="text-align:center;padding:10px;color:#4472C4;font-size:12px;font-weight:600;cursor:pointer" onclick="document.querySelectorAll('[data-extra]').forEach(e=>e.style.display='');this.style.display='none'">Show all ${spoolKeys.length} spools / \u663e\u793a\u5168\u90e8 \u25be</div>`;
  } else {
    html += `<div style="text-align:center;padding:20px;color:#aaa">No activity today / \u4eca\u5929\u6682\u65e0\u52a8\u6001</div>`;
  }
  html += `</div>`;
  document.getElementById('rpt-content').innerHTML = html;
}
async function markShipped(num){
  const today = new Date().toISOString().substring(0,10);
  const date = prompt('Dispatch date (YYYY-MM-DD) / \u53d1\u8fd0\u65e5\u671f:', today);
  if(!date) return;
  const r = await fetch(`/api/project/${P}/shipment/${num}/mark-shipped`, {
    method:'POST', headers:{'Content-Type':'application/json'},
    body: JSON.stringify({date: date, operator: 'shipment-bulk'})
  });
  const res = await r.json();
  if(res.ok){ alert(`\u2713 Marked ${res.updated} spools as shipped on ${res.date}`); load(); }
  else alert('Error: ' + (res.error || 'unknown'));
}
load();
</script></body></html>"""

# ── Chat Widget Bundle — injected into every page template ───────────────────
# Generalised: project ID is read from the URL at runtime (/project/<id>/...).
# No hardcoded project strings. Appears on landing, project, spool, QC, report.

CHAT_WIDGET_HTML = """
<!-- ENERXON Chat Assistant -->
<button class="chat-btn-robot" id="cxChatBtn" title="ENERXON Assistant" type="button">
<svg viewBox="0 0 32 32" xmlns="http://www.w3.org/2000/svg">
<line x1="16" y1="2" x2="16" y2="6" stroke="#2F5496" stroke-width="2" stroke-linecap="round"/>
<circle cx="16" cy="2.5" r="1.5" fill="#E63946"/>
<rect x="5" y="6" width="22" height="18" rx="4" ry="4" fill="#2F5496"/>
<rect x="7.5" y="9" width="17" height="12" rx="2" ry="2" fill="#fff"/>
<circle cx="12" cy="14.5" r="2" fill="#2F5496"/>
<circle cx="20" cy="14.5" r="2" fill="#2F5496"/>
<circle cx="12.6" cy="13.9" r="0.6" fill="#fff"/>
<circle cx="20.6" cy="13.9" r="0.6" fill="#fff"/>
<path d="M11 18 Q16 20 21 18" stroke="#2F5496" stroke-width="1.4" fill="none" stroke-linecap="round"/>
<rect x="3" y="12" width="2" height="6" rx="1" fill="#2F5496"/>
<rect x="27" y="12" width="2" height="6" rx="1" fill="#2F5496"/>
<rect x="13" y="24" width="6" height="2" fill="#2F5496"/>
<rect x="8" y="26" width="16" height="3" rx="1" fill="#E63946"/>
</svg>
<span class="chat-pulse"></span>
</button>
<div class="chat-backdrop" id="cxChatBackdrop"></div>
<div class="chat-panel" id="cxChatPanel">
  <div class="chat-panel-header">
    <div style="display:flex;align-items:center;gap:10px">
      <svg width="32" height="32" viewBox="0 0 32 32" xmlns="http://www.w3.org/2000/svg" style="flex-shrink:0">
        <line x1="16" y1="2" x2="16" y2="6" stroke="#fff" stroke-width="2" stroke-linecap="round"/>
        <circle cx="16" cy="2.5" r="1.5" fill="#E63946"/>
        <rect x="5" y="6" width="22" height="18" rx="4" ry="4" fill="#fff"/>
        <rect x="7.5" y="9" width="17" height="12" rx="2" ry="2" fill="#2F5496"/>
        <circle cx="12" cy="14.5" r="2" fill="#fff"/>
        <circle cx="20" cy="14.5" r="2" fill="#fff"/>
        <path d="M11 18 Q16 20 21 18" stroke="#fff" stroke-width="1.4" fill="none" stroke-linecap="round"/>
        <rect x="3" y="12" width="2" height="6" rx="1" fill="#fff"/>
        <rect x="27" y="12" width="2" height="6" rx="1" fill="#fff"/>
        <rect x="13" y="24" width="6" height="2" fill="#fff"/>
        <rect x="8" y="26" width="16" height="3" rx="1" fill="#E63946"/>
      </svg>
      <div>
        <div class="title">ENERXON \u52a9\u624b</div>
        <div class="title-sub">Production & Quality Assistant \u00b7 <span id="cxChatProjLabel"></span></div>
      </div>
    </div>
    <button class="chat-close" id="cxChatClose" type="button" aria-label="Close">\u00d7</button>
  </div>
  <div class="chat-body" id="cxChatBody">
    <div class="chat-welcome" id="cxChatWelcome">
      <strong>\u5173\u4e8e\u672c\u9879\u76ee\uff0c\u95ee\u6211\u4efb\u4f55\u95ee\u9898</strong>
      <span style="color:#888">Ask me anything about this project.</span>
    </div>
  </div>
  <div class="chat-input-bar">
    <textarea class="chat-input" id="cxChatInput" placeholder="\u95ee\u6211\u4efb\u4f55\u95ee\u9898 / Ask me anything..." rows="1"></textarea>
    <button class="chat-send" id="cxChatSend" type="button" title="Send">\u27a4</button>
  </div>
</div>
<script>
(function(){
  // Derive project ID from URL path. Works on /project/<id>, /project/<id>/spool/..., etc.
  // No hardcoded IDs — generalised.
  function cxGetProject(){
    var m = window.location.pathname.match(/\\/project\\/([^\\/]+)/);
    return m ? decodeURIComponent(m[1]) : null;
  }
  var cxProject = cxGetProject();
  var panel = document.getElementById('cxChatPanel');
  var backdrop = document.getElementById('cxChatBackdrop');
  var btn = document.getElementById('cxChatBtn');
  var closeBtn = document.getElementById('cxChatClose');
  var body = document.getElementById('cxChatBody');
  var welcome = document.getElementById('cxChatWelcome');
  var input = document.getElementById('cxChatInput');
  var send = document.getElementById('cxChatSend');
  var projLabel = document.getElementById('cxChatProjLabel');

  // Hide the widget entirely when not on a project page (landing has no project scope)
  if(!cxProject){ btn.style.display='none'; return; }
  projLabel.textContent = cxProject;

  var historyKey = 'cxChat_'+cxProject;
  var history = [];
  try{ history = JSON.parse(sessionStorage.getItem(historyKey)||'[]'); }catch(e){ history=[]; }

  function saveHistory(){ try{ sessionStorage.setItem(historyKey, JSON.stringify(history.slice(-20))); }catch(e){} }

  // Minimal safe renderer — preserves newlines, escapes HTML, bolds **text**, supports simple GFM tables
  function escapeHtml(s){ return s.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }
  function renderText(txt){
    var esc = escapeHtml(txt);
    esc = esc.replace(/\\*\\*([^*]+)\\*\\*/g,'<strong>$1</strong>');
    // Simple table detection
    var lines = esc.split('\\n');
    var out = []; var i = 0;
    while(i < lines.length){
      if(/^\\|.+\\|\\s*$/.test(lines[i]) && i+1<lines.length && /^\\|[\\s\\-:|]+\\|\\s*$/.test(lines[i+1])){
        var headers = lines[i].split('|').slice(1,-1).map(function(s){return s.trim();});
        i += 2;
        var rows = [];
        while(i < lines.length && /^\\|.+\\|\\s*$/.test(lines[i])){
          rows.push(lines[i].split('|').slice(1,-1).map(function(s){return s.trim();}));
          i++;
        }
        out.push('<table><tr>'+headers.map(function(h){return '<th>'+h+'</th>';}).join('')+'</tr>'+
          rows.map(function(r){return '<tr>'+r.map(function(c){return '<td>'+c+'</td>';}).join('')+'</tr>';}).join('')+'</table>');
      } else {
        out.push(lines[i]); i++;
      }
    }
    return out.join('\\n');
  }

  function addMessage(role, text, tools, logId){
    if(welcome){ welcome.style.display='none'; }
    var msg = document.createElement('div');
    msg.className = 'chat-msg '+role;
    var bubble = document.createElement('div');
    bubble.className = 'chat-bubble';
    bubble.innerHTML = renderText(text);
    msg.appendChild(bubble);
    if(role === 'assistant'){
      if(tools && tools.length){
        var meta = document.createElement('div');
        meta.className = 'chat-meta';
        tools.forEach(function(t){
          var b = document.createElement('span'); b.className='chat-tool-badge'; b.textContent=t; meta.appendChild(b);
        });
        msg.appendChild(meta);
      }
      if(logId){
        var fb = document.createElement('div');
        fb.className = 'chat-feedback';
        var up = document.createElement('button'); up.type='button'; up.textContent='\U0001F44D'; up.title='helpful';
        var dn = document.createElement('button'); dn.type='button'; dn.textContent='\U0001F44E'; dn.title='not helpful';
        function sendFb(val, btnEl, other){
          fetch('/api/chat/feedback', {method:'POST', headers:{'Content-Type':'application/json'}, body: JSON.stringify({log_id: logId, feedback: val})});
          btnEl.classList.add(val==='up'?'selected-up':'selected-down');
          other.classList.remove('selected-up','selected-down');
        }
        up.addEventListener('click', function(){ sendFb('up', up, dn); });
        dn.addEventListener('click', function(){ sendFb('down', dn, up); });
        fb.appendChild(up); fb.appendChild(dn);
        msg.appendChild(fb);
      }
    }
    body.appendChild(msg);
    body.scrollTop = body.scrollHeight;
  }

  function addThinking(){
    var msg = document.createElement('div');
    msg.className = 'chat-msg assistant';
    msg.id = 'cxChatThinking';
    msg.innerHTML = '<div class="chat-bubble"><div class="chat-thinking"><span></span><span></span><span></span></div></div>';
    body.appendChild(msg);
    body.scrollTop = body.scrollHeight;
  }
  function removeThinking(){ var t = document.getElementById('cxChatThinking'); if(t) t.remove(); }

  function openChat(){ panel.classList.add('open'); backdrop.classList.add('open'); setTimeout(function(){input.focus();},300); }
  function closeChat(){ panel.classList.remove('open'); backdrop.classList.remove('open'); }

  btn.addEventListener('click', openChat);
  closeBtn.addEventListener('click', closeChat);
  backdrop.addEventListener('click', closeChat);

  input.addEventListener('input', function(){ input.style.height='auto'; input.style.height=Math.min(input.scrollHeight,80)+'px'; });
  input.addEventListener('keydown', function(e){ if(e.key==='Enter' && !e.shiftKey){ e.preventDefault(); send.click(); } });

  // Restore prior conversation
  history.forEach(function(h){
    if(h.role==='user' || h.role==='assistant') addMessage(h.role, h.content, h.tools||null, h.log_id||null);
  });

  send.addEventListener('click', function(){
    var msg = input.value.trim();
    if(!msg) return;
    addMessage('user', msg);
    history.push({role:'user', content: msg});
    input.value = ''; input.style.height='auto';
    send.disabled = true;
    addThinking();
    fetch('/api/chat', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({project: cxProject, message: msg, history: history.slice(0,-1)})
    })
    .then(function(r){ return r.json(); })
    .then(function(d){
      removeThinking();
      send.disabled = false;
      if(d.error){ addMessage('assistant', 'Error: '+d.error); return; }
      var reply = d.reply || '(no reply)';
      var tools = d.tools_used || [];
      var logId = d.log_id || null;
      addMessage('assistant', reply, tools, logId);
      history.push({role:'assistant', content: reply, tools: tools, log_id: logId});
      saveHistory();
    })
    .catch(function(err){
      removeThinking();
      send.disabled = false;
      addMessage('assistant', 'Network error: '+err.message);
    });
  });
})();
</script>"""

# Inject into every page template. One loop, one rule — any new template added
# later just needs to include </body> to pick up the widget automatically.
for _tpl in ('HOME_HTML', 'PROJECT_HTML', 'SPOOL_HTML', 'QC_REPORT_HTML', 'REPORT_HTML'):
    if _tpl in globals() and '</body>' in globals()[_tpl]:
        globals()[_tpl] = globals()[_tpl].replace('</body>', CHAT_WIDGET_HTML + '\n</body>', 1)

# ── Init & Run ────────────────────────────────────────────────────────────────
try: init_db(); print("DB initialized")
except Exception as e: print(f"DB init: {e}")

if __name__ == '__main__':
    app.run(host=os.environ.get('HOST','0.0.0.0'), port=int(os.environ.get('PORT',5000)))
