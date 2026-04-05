#!/usr/bin/env python3
"""
ENERXON Tracker — New Project Deployment Script (Generalised)
Parses spool matrix + route cards, calculates joints/surface, uploads everything.
Step definitions are project-specific (no hardcoded ITP steps).

Usage:
  python3 deploy_project.py \
    --project ENJOB25XXXXX \
    --matrix "/path/to/Matrix.xlsx" \
    --steps-file "/path/to/steps_424.json" \
    --route-cards "/path/to/Combined/" \
    --schedule-start 2026-04-01 \
    --standard-weeks 9 \
    --welding-ipd 1000 \
    --surface-m2d 91 \
    --expediting-weeks 0 \
    --spools-per-day '{"32":1.5,"24":1.7,"18":3.0,"16":2.0,"8":1.5,"2":2.5,"1":2.5}' \
    --secondary-phase-days 13 \
    [--drawings "/path/to/drawings/"] \
    [--transit-days 45]
"""

import argparse
import json
import math
import os
import re
import sys
from collections import defaultdict
from datetime import date, timedelta

import openpyxl
import requests

TRACKER_URL = "https://enerxon-china-tracker.onrender.com"
PASSWORD = "Enerxon@china"

# Outer diameter lookup (inches -> mm)
OD_MM = {
    0.5: 21.3, 0.75: 26.7, 1: 33.4, 1.5: 48.3, 2: 60.3, 3: 88.9, 4: 114.3,
    6: 168.3, 8: 219.1, 10: 273.1, 12: 323.9, 14: 355.6, 16: 406.4,
    18: 457.2, 20: 508.0, 24: 609.6, 28: 711.2, 30: 762.0, 32: 812.8,
    36: 914.4, 42: 1066.8, 48: 1219.2,
}


def login(session):
    """Login to tracker and return session."""
    r = session.post(f"{TRACKER_URL}/login", data={"password": PASSWORD})
    return session


def parse_spool_matrix(matrix_file):
    """Parse spool matrix Excel to extract spool data."""
    wb = openpyxl.load_workbook(matrix_file, data_only=True)

    spool_sheet = None
    for name in wb.sheetnames:
        if 'spool' in name.lower() or 'matrix' in name.lower():
            spool_sheet = wb[name]
            break
    if not spool_sheet:
        spool_sheet = wb[wb.sheetnames[-1]]

    ws = spool_sheet
    spools = []
    current_iso = None

    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(r, c).value or '').upper()
            if 'SPOOL' in val and 'DWG' in val:
                header_row = r
                break
        if header_row:
            break

    if not header_row:
        print(f"  WARNING: Could not find header row in {matrix_file}")
        wb.close()
        return spools

    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(header_row, c).value or '').upper()
        if 'N\u00b0' in val or val == 'N': col_map['num'] = c
        elif 'ISO' in val: col_map['iso'] = c
        elif 'SPOOL' in val and 'DWG' in val: col_map['spool'] = c
        elif 'MARK' in val: col_map['marking'] = c
        elif 'CUSTOMER' in val or 'CODE' in val and 'SPOOL' in val: col_map['mk'] = c

    for r in range(header_row + 1, ws.max_row + 1):
        num_val = ws.cell(r, col_map.get('num', 2)).value
        iso_val = ws.cell(r, col_map.get('iso', 3)).value
        spool_val = ws.cell(r, col_map.get('spool', 4)).value
        marking_val = ws.cell(r, col_map.get('marking', 9)).value
        mk_val = ws.cell(r, col_map.get('mk', 8)).value

        if num_val and iso_val:
            current_iso = str(iso_val).strip()

        if spool_val and str(spool_val).strip():
            spool_str = str(spool_val).strip()
            spl_match = re.search(r'(SPL-\d+[A-B]?)', spool_str, re.IGNORECASE)
            straight_match = re.search(r'STRAIGHT\s*PIPE', spool_str, re.IGNORECASE)

            if spl_match or straight_match:
                if spl_match:
                    spool_id = spl_match.group(1).upper()
                    spool_type = 'SPOOL'
                else:
                    len_match = re.search(r'L[=\s]*(\d+)', spool_str)
                    length = len_match.group(1) if len_match else '0'
                    spool_id = f"STRAIGHT PIPE L={length}mm"
                    spool_type = 'STRAIGHT'

                spools.append({
                    'spool_id': spool_id,
                    'spool_full': spool_str,
                    'iso_no': current_iso or '',
                    'marking': str(marking_val or '').strip(),
                    'mk_number': str(mk_val or '').strip(),
                    'spool_type': spool_type,
                })

    wb.close()
    return spools


def parse_route_cards(route_cards_dir):
    """Parse all route card PDFs to extract joints, surface area per spool."""
    try:
        import pdfplumber
    except ImportError:
        print("ERROR: pdfplumber not installed. Run: pip install pdfplumber")
        sys.exit(1)

    spool_data = {}

    for root, dirs, files in os.walk(route_cards_dir):
        folder_name = os.path.basename(root)
        line = 'A'
        if folder_name in ['A', 'B', 'C']:
            line = folder_name

        parent = os.path.basename(os.path.dirname(root))
        diameter = 0
        dim_match = re.search(r'(\d+)\s*INCH', parent, re.IGNORECASE)
        if dim_match:
            diameter = int(dim_match.group(1))

        for fname in files:
            if not fname.endswith('.pdf'):
                continue
            if '_Line' not in fname and '_in_' not in fname:
                continue

            fpath = os.path.join(root, fname)
            spool_dir = os.path.basename(root)

            if 'STRAIGHT' in spool_dir.upper():
                continue

            try:
                pdf = pdfplumber.open(fpath)
                text = pdf.pages[0].extract_text() or ""
                pdf.close()
            except Exception as e:
                print(f"  WARNING: Could not read {fpath}: {e}")
                continue

            spl_match = re.search(r'(SPL-\d+[A-B]?)', spool_dir, re.IGNORECASE)
            if not spl_match:
                continue
            spool_id = spl_match.group(1).upper()

            if diameter == 0:
                dm = re.search(r'(\d+)"', text[:500])
                if dm:
                    diameter = int(dm.group(1))

            pipes = len(re.findall(r'P-\d+\s+PIPE', text))
            fittings = sum(int(m.group(1)) for m in re.finditer(r'FIT-\d+\s+.*?\s+\d+"\s+(\d+)', text))
            flanges = sum(int(m.group(1)) for m in re.finditer(r'FLG-\d+\s+FLG.*?\s+\d+"\s+(\d+)', text))
            joints = max(pipes - 1, 0) + fittings + flanges
            raf = joints * diameter

            surface = 0
            for m in re.finditer(r'P-\d+\s+PIPE.*?(\d+)"\s+([\d,]+)\s*mm', text):
                d_inch = int(m.group(1))
                length = int(m.group(2).replace(',', ''))
                od = OD_MM.get(d_inch, d_inch * 25.4)
                surface += math.pi * (od / 1000) * (length / 1000)

            for m in re.finditer(r'FLG-\d+\s+FLG.*?(\d+)"\s+(\d+)', text):
                d_inch = int(m.group(1))
                qty = int(m.group(2))
                od = OD_MM.get(d_inch, d_inch * 25.4)
                od_m = od / 1000
                surface += (math.pi * (od_m / 2) ** 2 * 2 + math.pi * od_m * 0.05) * qty

            for m in re.finditer(r'FIT-\d+\s+(.*?)\s+(\d+)"\s+(\d+)', text):
                desc = m.group(1)
                d_inch = int(m.group(2))
                qty = int(m.group(3))
                od = OD_MM.get(d_inch, d_inch * 25.4)
                od_m = od / 1000
                mult = 3.5 if 'TEE' in desc.upper() else 2.5 if 'ELL' in desc.upper() else 2.0
                surface += (math.pi * od_m * od_m * mult) * qty

            has_branches = 1 if re.search(r'TEE|OLET|Branch', text, re.IGNORECASE) else 0

            spool_data[spool_id] = {
                'diameter': diameter,
                'line': line,
                'joint_count': joints,
                'raf_inches': raf,
                'surface_m2': round(surface, 4),
                'has_branches': has_branches,
            }

    return spool_data


def build_schedule(start_date, standard_weeks, diameter_counts, secondary_phase_days, diam_order, phase_order=None):
    """Build standard schedule: proportional distribution across standard_weeks.
    One diameter at a time, largest first. Each gets (count/total) share of total days."""
    if not phase_order:
        phase_order = ['fab']
    total_spools = sum(diameter_counts.values())
    total_std_days = standard_weeks * 7
    first_phase = phase_order[0]
    # For multi-phase: first phase total days = total - secondary_phase_days (last diameter paint must finish by end)
    first_phase_total_days = total_std_days - secondary_phase_days if len(phase_order) > 1 else total_std_days
    schedule = []
    current_start = start_date
    remaining_days = first_phase_total_days
    active_diams = [d for d in diam_order if d in diameter_counts]

    for i, diam in enumerate(active_diams):
        cnt = diameter_counts[diam]
        if i == len(active_diams) - 1:
            fab_days = remaining_days  # last diameter gets remainder
        else:
            fab_days = max(1, round(first_phase_total_days * cnt / total_spools))
        remaining_days -= fab_days
        fab_end = current_start + timedelta(days=fab_days)
        dk = f'{diam}"'
        schedule.append({
            'diameter': dk, 'task_type': first_phase,
            'description': f'{first_phase.capitalize()} {dk} ({cnt} spools)',
            'planned_start': str(current_start), 'planned_end': str(fab_end),
            'spool_count': cnt,
        })
        prev_end = fab_end
        for ph in phase_order[1:]:
            ph_start = prev_end + timedelta(days=1)
            ph_end = ph_start + timedelta(days=secondary_phase_days)
            schedule.append({
                'diameter': dk, 'task_type': ph,
                'description': f'{ph.capitalize()} {dk}',
                'planned_start': str(ph_start), 'planned_end': str(ph_end),
                'spool_count': cnt,
            })
            prev_end = ph_end
        current_start = fab_end  # next diameter starts when this one's first phase ends

    return schedule


def deploy(args):
    """Main deployment function."""
    print(f"\n{'='*60}")
    print(f"  DEPLOYING PROJECT: {args.project}")
    print(f"{'='*60}\n")

    session = requests.Session()
    login(session)

    # Parse spools_per_day
    spd = json.loads(args.spools_per_day) if args.spools_per_day else {}
    secondary_phase_days = args.secondary_phase_days

    # ── Step 1: Parse Spool Matrix ──────────────────────────────
    print("Step 1: Parsing spool matrix...")
    spools = parse_spool_matrix(args.matrix)
    print(f"  Found {len(spools)} spools in matrix")

    # ── Step 2: Parse Route Cards ───────────────────────────────
    print("\nStep 2: Parsing route cards...")
    route_data = {}
    if args.route_cards and os.path.isdir(args.route_cards):
        route_data = parse_route_cards(args.route_cards)
        print(f"  Parsed {len(route_data)} route cards")
    else:
        print("  No route cards folder provided, skipping")

    # ── Step 3: Upload step definitions ─────────────────────────
    if args.steps_file:
        print("\nStep 3: Uploading step definitions...")
        with open(args.steps_file) as f:
            steps = json.load(f)
        r = session.post(f"{TRACKER_URL}/api/project/{args.project}/steps", json=steps)
        print(f"  Result: {r.json()}")
    else:
        print("\nStep 3: No --steps-file provided, using default 424 steps")

    # ── Step 4: Merge data + determine diameters/lines ──────────
    print("\nStep 4: Merging spool data...")
    import_data = []
    diameter_counts = defaultdict(int)

    for spool in spools:
        sid = spool['spool_id']
        rd = route_data.get(sid, {})

        diameter = rd.get('diameter', 0)
        line = rd.get('line', 'A')
        has_branches = rd.get('has_branches', 0)
        spool_type = spool.get('spool_type', 'SPOOL')

        main_diameter = f'{diameter}"' if diameter > 0 else '?'

        import_item = {
            'spool_id': sid,
            'spool_full': spool.get('spool_full', ''),
            'iso_no': spool.get('iso_no', ''),
            'marking': spool.get('marking', ''),
            'mk_number': spool.get('mk_number', ''),
            'main_diameter': main_diameter,
            'line': line,
            'sequence': len(import_data) + 1,
            'project': args.project,
            'has_branches': has_branches,
            'spool_type': spool_type,
        }
        import_data.append(import_item)

        if diameter > 0:
            diameter_counts[str(diameter)] += 1

    print(f"  {len(import_data)} spools ready for import")
    print(f"  Diameters: {dict(diameter_counts)}")

    # ── Step 5: Import spools ───────────────────────────────────
    print("\nStep 5: Importing spools to tracker...")
    r = session.post(f"{TRACKER_URL}/api/import", json=import_data)
    result = r.json()
    print(f"  Result: {result}")

    # ── Step 6: Import joint counts ─────────────────────────────
    if route_data:
        print("\nStep 6: Importing joint counts...")
        joints_upload = {sid: {'joint_count': rd['joint_count'], 'raf_inches': rd['raf_inches']}
                        for sid, rd in route_data.items()}
        r = session.post(f"{TRACKER_URL}/api/project/{args.project}/joints", json=joints_upload)
        print(f"  Result: {r.json()}")

        print("\nStep 7: Importing surface areas...")
        surface_upload = {sid: rd['surface_m2'] for sid, rd in route_data.items()}
        r = session.post(f"{TRACKER_URL}/api/project/{args.project}/surface", json=surface_upload)
        print(f"  Result: {r.json()}")

    # ── Step 8: Create schedule ─────────────────────────────────
    print("\nStep 8: Creating production schedule...")
    start = date.fromisoformat(args.schedule_start)
    # Derive diameter order from actual data (descending)
    diam_order = sorted(diameter_counts.keys(), key=lambda x: float(x) if x.replace('.','').isdigit() else 0, reverse=True)
    # Load phase order from step definitions
    phase_order = ['fab']
    if args.steps_file:
        with open(args.steps_file) as f:
            steps = json.load(f)
        seen = set()
        phase_order = []
        for s in steps:
            p = s.get('phase', 'fab')
            if p not in seen:
                seen.add(p)
                phase_order.append(p)
    schedule = build_schedule(start, args.standard_weeks, diameter_counts, secondary_phase_days, diam_order, phase_order)
    r = session.post(f"{TRACKER_URL}/api/project/{args.project}/schedule", json=schedule)
    print(f"  Result: {r.json()}")

    # ── Step 9: Set project settings ────────────────────────────
    print("\nStep 9: Setting project settings...")
    settings = {
        'standard_weeks': str(args.standard_weeks),
        'welding_capability_ipd': str(args.welding_ipd),
        'surface_capability_m2d': str(args.surface_m2d),
        'committed_weeks_saved': str(args.expediting_weeks),
        'committed_days_saved': '0',
        'sea_transit_days': str(args.transit_days),
        'spools_per_day': json.dumps(spd),
        'secondary_phase_days': str(secondary_phase_days),
    }
    r = session.post(f"{TRACKER_URL}/api/project/{args.project}/settings", json=settings)
    print(f"  Result: {r.json()}")

    # ── Step 10: Upload drawings (optional) ─────────────────────
    if args.drawings and os.path.isdir(args.drawings):
        print("\nStep 10: Uploading spool drawings...")
        uploaded = 0
        for root, dirs, files in os.walk(args.drawings):
            for fname in files:
                if not fname.lower().endswith('.pdf'):
                    continue
                spl_match = re.search(r'(SPL-\d+[A-B]?)', fname, re.IGNORECASE)
                if not spl_match:
                    continue
                sid = spl_match.group(1).upper()
                fpath = os.path.join(root, fname)
                with open(fpath, 'rb') as f:
                    r = session.post(
                        f"{TRACKER_URL}/api/project/{args.project}/spool/{sid}/drawing",
                        files={'drawing': (fname, f, 'application/pdf')}
                    )
                    if r.ok:
                        uploaded += 1
                if uploaded % 20 == 0 and uploaded > 0:
                    print(f"    Uploaded {uploaded} drawings...")
        print(f"  Uploaded {uploaded} drawings")

    # ── Step 11: Chat Assistant Knowledge File Check ────────────
    # The tracker chat assistant reads knowledge/<project>_qc_knowledge.md
    # on demand. If the file is missing, the 10 knowledge tools return a
    # friendly "no knowledge base" message and the widget still works for
    # live data tools. We warn here so the deployer knows to add the file
    # before QC staff start asking standards/procedure questions.
    print("\nStep 11: Checking chat assistant knowledge file...")
    knowledge_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                   'knowledge', f'{args.project}_qc_knowledge.md')
    if os.path.exists(knowledge_path):
        size_kb = os.path.getsize(knowledge_path) // 1024
        print(f"  [OK] Found knowledge/{args.project}_qc_knowledge.md ({size_kb} KB)")
        print(f"       Chat assistant will answer quality/standards questions from this file.")
    else:
        print(f"  [WARN] No knowledge file at knowledge/{args.project}_qc_knowledge.md")
        print(f"         The chat assistant live-data tools will work, but standards,")
        print(f"         WPS, NDT, ITP, and acceptance-criteria questions will return")
        print(f"         'No QC knowledge base is configured for this project'.")
        print(f"         To enable full chat coverage, create the markdown file with")
        print(f"         ## section headings covering: Project Overview, Material &")
        print(f"         Metallurgy, Welding (WPS registry), NDT, ITP Flow, Finishing,")
        print(f"         Cutting & Fit-up, Documentation, Q&A, Glossary. Commit and")
        print(f"         redeploy — no code change required.")

    # ── Summary ─────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  DEPLOYMENT COMPLETE: {args.project}")
    print(f"{'='*60}")
    print(f"  Spools imported: {len(import_data)}")
    print(f"  Route cards parsed: {len(route_data)}")
    print(f"  Schedule: {len(schedule)} entries")
    print(f"  Settings: {settings}")
    print(f"  Knowledge file: {'present' if os.path.exists(knowledge_path) else 'MISSING (chat assistant will have limited answers)'}")
    print(f"\n  View at: {TRACKER_URL}/project/{args.project}")
    print(f"  Report:  {TRACKER_URL}/project/{args.project}/report")
    print(f"  Password: {PASSWORD}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Deploy new project to ENERXON Tracker')
    parser.add_argument('--project', required=True, help='Project ID (e.g., ENJOB25XXXXX)')
    parser.add_argument('--matrix', required=True, help='Path to spool matrix Excel file')
    parser.add_argument('--steps-file', help='JSON file with ITP step definitions for this project')
    parser.add_argument('--route-cards', help='Path to route cards folder (Combined/)')
    parser.add_argument('--schedule-start', required=True, help='Production start date (YYYY-MM-DD)')
    parser.add_argument('--standard-weeks', type=int, default=9, help='Standard production weeks (default: 9)')
    parser.add_argument('--welding-ipd', type=float, default=1000, help='Welding capability (linear inches/day)')
    parser.add_argument('--surface-m2d', type=float, default=91, help='Surface treatment capability (m\u00b2/day, 0 if no surface phase)')
    parser.add_argument('--expediting-weeks', type=int, default=0, help='Expediting weeks saved (0 = no expediting)')
    parser.add_argument('--transit-days', type=int, default=45, help='Sea transit days (default: 45)')
    parser.add_argument('--drawings', help='Path to spool drawings folder (optional)')
    parser.add_argument('--spools-per-day', default='{}', help='JSON: spools/day per diameter, e.g. \'{"32":1.5,"24":1.7}\'')
    parser.add_argument('--secondary-phase-days', type=int, default=13, help='Days for secondary phase per diameter (default: 13)')

    args = parser.parse_args()
    deploy(args)
